#!/usr/bin/env python3

from __future__ import annotations

import argparse
import bisect
import csv
import json
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles.numbers import BUILTIN_FORMATS, is_date_format
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, from_excel


PARTNER_PATTERNS = [
    ("altpay", "Altpay"),
    ("athena", "Athena"),
    ("bhn", "BHN"),
    ("capi", "Capi"),
    ("cellpay", "Cellpay"),
    ("clearshift", "Clearshift"),
    ("everflow", "Everflow"),
    ("factura", "Factura"),
    ("finastra", "Finastra"),
    ("goldstack", "Goldstack"),
    ("fulfil", "Fulfil"),
    ("gme_remit", "GME_Remit"),
    ("gme remit", "GME_Remit"),
    ("graph finance", "Graph Finance"),
    ("graph", "Graph Finance"),
    ("halorecruiting", "Halorecruiting"),
    ("jazz cash", "Jazz Cash"),
    ("jazz", "Jazz Cash"),
    ("lightnet", "LightNet"),
    ("magaya", "Magaya"),
    ("multigate", "MultiGate"),
    ("nibss", "NIBSS ( TurboTech)"),
    ("nium", "Nium"),
    ("ohent", "OhentPay"),
    ("oson", "Oson"),
    ("q2", "Q2"),
    ("repay", "Repay"),
    ("shepherd", "Shepherd"),
    ("stampli", "Stampli"),
    ("skydo", "Skydo"),
    ("blindpay", "Blindpay"),
    ("yeepay", "Yeepay"),
    ("nuvion", "Nuvion"),
    ("maplewave", "Maplewave"),
    ("triple-a", "TripleA"),
    ("triplea", "TripleA"),
    ("nomadglobal", "Nomad"),
    ("nsave", "Nsave"),
    ("lianlian", "LianLian"),
    ("whish", "Whish"),
    ("remittances hub", "Remittanceshub"),
    ("remittanceshub", "Remittanceshub"),
    ("triple-a", "TripleA"),
    ("triplea", "TripleA"),
    ("vg pay", "VG Pay"),
    ("vgpay", "VG Pay"),
]

PARTNER_ALIASES = {
    "altpay": "Altpay",
    "athena": "Athena",
    "bhn": "BHN",
    "blindpay": "Blindpay",
    "capi": "Capi",
    "cellpay": "Cellpay",
    "clearshift": "Clearshift",
    "everflow": "Everflow",
    "factura": "Factura",
    "finastra": "Finastra",
    "fulfil": "Fulfil",
    "gme_remit": "GME_Remit",
    "gme remit": "GME_Remit",
    "goldstack": "Goldstack",
    "graph": "Graph Finance",
    "graph finance": "Graph Finance",
    "halorecruiting": "Halorecruiting",
    "jazz cash": "Jazz Cash",
    "jazz": "Jazz Cash",
    "lightnet": "LightNet",
    "lianlian": "LianLian",
    "magaya": "Magaya",
    "maplewave": "Maplewave",
    "multigate": "MultiGate",
    "nibss": "NIBSS ( TurboTech)",
    "nium": "Nium",
    "nomad": "Nomad",
    "nsave": "Nsave",
    "nuvion": "Nuvion",
    "ohent": "OhentPay",
    "ohentpay": "OhentPay",
    "q2": "Q2",
    "remittanceshub": "Remittanceshub",
    "remittances hub": "Remittanceshub",
    "shepherd": "Shepherd",
    "skydo": "Skydo",
    "stampli": "Stampli",
    "triplea": "TripleA",
    "triple-a": "TripleA",
    "vg pay": "VG Pay",
    "vgpay": "VG Pay",
    "whish": "Whish",
    "yeepay": "Yeepay",
}
STAMPLI_FX_PARTNER = "Stampli"
STAMPLI_COMPANY_MARKUP_BPS = 0.004
STAMPLI_MARKUP_AMOUNT_TOLERANCE = 0.011
STAMPLI_MARKUP_RATE_TOLERANCE = 0.0001
FX_VARIABLE_SPREAD_TIERS = [
    (0, 1_000_000, 0.0015),
    (1_000_001, 5_000_000, 0.0012),
    (5_000_001, 10_000_000, 0.0010),
]
FX_DEFAULT_PAYMENT_FEE = 9.0
FX_PAYMENT_FEE_LABEL = "SWIFT - SHA"
FX_LOCAL_PAYMENT_FEES = {
    "AUD": 2.0,
    "CAD": 1.0,
    "CZK": 2.0,
    "DKK": 1.0,
    "EUR": 1.0,
    "GBP": 1.0,
    "HKD": 2.0,
    "HRK": 3.0,
    "HUF": 2.0,
    "IDR": 2.0,
    "INR": 1.0,
    "MYR": 1.0,
    "NOK": 1.0,
    "PHP": 3.0,
    "PLN": 1.0,
    "RON": 2.0,
    "SEK": 1.0,
    "SGD": 2.0,
    "USD": 0.4,
}
EEA_COUNTRY_CODES = {
    "AT",
    "BE",
    "BG",
    "HR",
    "CY",
    "CZ",
    "DK",
    "EE",
    "FI",
    "FR",
    "DE",
    "GR",
    "HU",
    "IS",
    "IE",
    "IT",
    "LV",
    "LI",
    "LT",
    "LU",
    "MT",
    "NL",
    "NO",
    "PL",
    "PT",
    "RO",
    "SK",
    "SI",
    "ES",
    "SE",
}
US_COUNTRY_CODES = {"US", "USA"}
UK_COUNTRY_CODES = {"GB", "UK"}
CA_COUNTRY_CODES = {"CA", "CAN"}
AU_COUNTRY_CODES = {"AU", "AUS"}
STREAM_XLSX_THRESHOLD_BYTES = 25_000_000
XLSX_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


@dataclass(frozen=True)
class Paths:
    offline_txns: Path
    offline_reversals: Path
    offline_accounts: Path | None
    revenue_txns: list[Path]
    revenue_summary: Path | None
    stampli_credit_complete_all: Path | None
    stampli_domestic_revenue: Path | None
    stampli_usd_abroad_revenue: Path | None
    stampli_fx_share: Path | None
    stampli_fx_reversals: Path | None


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def month_key(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    raw = text(value)
    if not raw:
        return ""
    for fmt in ("%Y-%m", "%Y-%m-%d", "%b-%y", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    if len(raw) >= 7 and raw[4] == "-":
        return raw[:7]
    return raw


def money(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = text(value).replace("$", "").replace(",", "")
    return float(raw) if raw else 0.0


def boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return text(value).lower() in {"true", "yes", "1", "rtp"}


def iso_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = text(value)
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return raw


def slugify(value: str) -> str:
    lowered = text(value).lower()
    chars = [char if char.isalnum() else "-" for char in lowered]
    slug = "".join(chars)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-") or "export"


@lru_cache(maxsize=2048)
def xlsx_column_index(cell_ref: str) -> int:
    value = 0
    for char in str(cell_ref or ""):
        if not ("A" <= char <= "Z"):
            break
        value = value * 26 + (ord(char) - 64)
    return max(value - 1, 0)


def load_xlsx_shared_strings(workbook: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in workbook.namelist():
        return []
    with workbook.open("xl/sharedStrings.xml") as handle:
        root = ET.parse(handle).getroot()
    values: list[str] = []
    for item in root.findall(f"{XLSX_NS}si"):
        values.append("".join(text_node.text or "" for text_node in item.findall(f".//{XLSX_NS}t")))
    return values


def workbook_uses_mac_epoch(workbook: zipfile.ZipFile) -> bool:
    if "xl/workbook.xml" not in workbook.namelist():
        return False
    with workbook.open("xl/workbook.xml") as handle:
        root = ET.parse(handle).getroot()
    workbook_properties = root.find(f"{XLSX_NS}workbookPr")
    return str(workbook_properties.attrib.get("date1904") or "").lower() in {"1", "true"} if workbook_properties is not None else False


def load_xlsx_date_style_indexes(workbook: zipfile.ZipFile) -> set[int]:
    if "xl/styles.xml" not in workbook.namelist():
        return set()
    with workbook.open("xl/styles.xml") as handle:
        root = ET.parse(handle).getroot()
    format_codes: dict[int, str] = {}
    num_formats = root.find(f"{XLSX_NS}numFmts")
    if num_formats is not None:
        for number_format in num_formats.findall(f"{XLSX_NS}numFmt"):
            try:
                format_codes[int(number_format.attrib.get("numFmtId") or "0")] = str(number_format.attrib.get("formatCode") or "")
            except ValueError:
                continue
    date_styles: set[int] = set()
    cell_formats = root.find(f"{XLSX_NS}cellXfs")
    if cell_formats is None:
        return date_styles
    for index, cell_format in enumerate(cell_formats.findall(f"{XLSX_NS}xf")):
        try:
            number_format_id = int(cell_format.attrib.get("numFmtId") or "0")
        except ValueError:
            continue
        format_code = format_codes.get(number_format_id) or BUILTIN_FORMATS.get(number_format_id) or ""
        if format_code and is_date_format(format_code):
            date_styles.add(index)
    return date_styles


def coerce_xlsx_number(value: str) -> Any:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return int(number) if number.is_integer() else number


def xlsx_cell_value(
    cell: ET.Element,
    shared_strings: list[str],
    date_style_indexes: set[int],
    use_mac_epoch: bool,
) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text_node.text or "" for text_node in cell.findall(f".//{XLSX_NS}t"))
    raw_value = cell.find(f"{XLSX_NS}v")
    if raw_value is None or raw_value.text is None:
        return ""
    value = raw_value.text
    if cell_type == "s":
        try:
            return shared_strings[int(value)]
        except (ValueError, IndexError):
            return value
    if cell_type == "b":
        return value == "1"
    try:
        style_index = int(cell.attrib.get("s") or "0")
    except ValueError:
        style_index = 0
    if style_index in date_style_indexes:
        try:
            return from_excel(float(value), epoch=MAC_EPOCH if use_mac_epoch else WINDOWS_EPOCH)
        except (TypeError, ValueError):
            return value
    if cell_type in (None, "n"):
        return coerce_xlsx_number(value)
    return value


def iter_large_xlsx_rows(path: Path):
    with zipfile.ZipFile(path) as workbook:
        shared_strings = load_xlsx_shared_strings(workbook)
        date_style_indexes = load_xlsx_date_style_indexes(workbook)
        use_mac_epoch = workbook_uses_mac_epoch(workbook)
        with workbook.open("xl/worksheets/sheet1.xml") as handle:
            iterator = ET.iterparse(handle, events=("end",))
            headers: list[str] | None = None
            for _, element in iterator:
                if element.tag != f"{XLSX_NS}row":
                    continue
                row_values: dict[int, Any] = {}
                for cell in element.findall(f"{XLSX_NS}c"):
                    reference = str(cell.attrib.get("r") or "")
                    if not reference:
                        continue
                    row_values[xlsx_column_index(reference)] = xlsx_cell_value(
                        cell,
                        shared_strings,
                        date_style_indexes,
                        use_mac_epoch,
                    )
                if headers is None:
                    max_index = max(row_values.keys(), default=-1)
                    headers = [""] * (max_index + 1)
                    for index, value in row_values.items():
                        headers[index] = text(value)
                    element.clear()
                    continue
                row = {
                    header: row_values.get(index)
                    for index, header in enumerate(headers)
                    if header and row_values.get(index) not in (None, "")
                }
                if any(value not in (None, "") for value in row.values()):
                    yield row
                element.clear()


def matches_period(month: str, period: str | None) -> bool:
    return not period or month == period


def title_case_funding(value: Any) -> str:
    normalized = text(value).lower()
    mapping = {
        "bank": "Bank",
        "wallet": "Wallet",
        "card": "Card",
        "credit": "Credit",
        "debit": "Debit",
        "cash": "Cash",
    }
    return mapping.get(normalized, normalized.title() if normalized else "")


def normalize_partner_name(raw_partner: Any) -> str:
    cleaned = text(raw_partner)
    if not cleaned:
        return ""
    alias = PARTNER_ALIASES.get(cleaned.lower())
    if alias:
        return alias
    for prefix, partner in PARTNER_ALIASES.items():
        if cleaned.lower().startswith(prefix):
            return partner
    return cleaned


def choose_fx_variable_spread_rate(period_volume: float) -> float:
    for min_volume, max_volume, rate in FX_VARIABLE_SPREAD_TIERS:
        if period_volume >= min_volume and period_volume <= max_volume:
            return rate
    return FX_VARIABLE_SPREAD_TIERS[-1][2]


def choose_fx_payment_fee(payee_currency: str) -> tuple[float, str]:
    normalized = text(payee_currency).upper()
    if normalized in FX_LOCAL_PAYMENT_FEES:
        return FX_LOCAL_PAYMENT_FEES[normalized], f"Local payment fee ({normalized})"
    return FX_DEFAULT_PAYMENT_FEE, FX_PAYMENT_FEE_LABEL


def choose_stampli_mid_market_usd(
    row: dict[str, Any],
    month: str,
    avg_ratio_by_month_currency: dict[tuple[str, str], float],
) -> tuple[float, bool]:
    payee_currency = text(
        row_value_first(
            row,
            "** Payment For Sales DV ** Payee Amount Currency",
            "Currency (Payee Amount Currency)",
            "Payee Amount Currency",
            patterns=("payeeamountcurrency",),
        )
    ).upper()
    payee_amount = money(
        row_value_first(
            row,
            "** Payment For Sales DV ** Payee Amount Number",
            "Foreign Currency Amount (Payee Amount Number)",
            patterns=("payeeamountnumber",),
        )
    )
    mid_market_usd = round(
        money(
            row_value_first(
                row,
                "Payment USD Equivalent Amount",
                "** Payment For Sales DV ** Total USD Amount Number",
                "** Payment For Sales DV ** USD Amount Number",
                patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"),
            )
        ),
        2,
    )
    if mid_market_usd > 0:
        return mid_market_usd, False
    ratio = avg_ratio_by_month_currency.get((month, payee_currency), 0.0)
    if payee_amount > 0 and ratio > 0:
        return round(payee_amount * ratio, 2), True
    return 0.0, False


def calculate_stampli_markup_from_feed_row(
    row: dict[str, Any],
    month: str,
    avg_ratio_by_month_currency: dict[tuple[str, str], float],
) -> dict[str, Any]:
    mid_market_usd, used_period_average = choose_stampli_mid_market_usd(row, month, avg_ratio_by_month_currency)
    transaction_volume_usd = round(
        money(
            row_value_first(
                row,
                "** Payment For Sales DV ** Total USD Amount Number",
                "** Payment For Sales DV ** USD Amount Number",
                "Payment USD Equivalent Amount",
                patterns=("totalusdamountnumber", "usdamountnumber", "paymentusdequivalentamount"),
            )
        ),
        2,
    )
    if mid_market_usd <= 0:
        mid_market_usd = transaction_volume_usd
    usd_debited = round(
        money(
            row_value_first(
                row,
                "USD Amount Debited to the Customer",
                "** Payment For Sales DV ** Total USD Amount Number",
                "** Payment For Sales DV ** USD Amount Number",
                patterns=("usdamountdebited", "totalusdamountnumber", "usdamountnumber"),
            )
        ),
        2,
    )
    provided_amount = round(money(row.get("Stampli Markup Amount")), 2)
    customer_markup_pct = round(
        money(
            row_value_first(
                row,
                "Customer Markup (%)",
                "Partner Revenue Share Variable Payer Markup Rate",
                patterns=("customermarkup", "payermarkuprate"),
            )
        ),
        6,
    )
    stampli_buy_rate_pct = round(
        money(
            row_value_first(
                row,
                "Stampli Buy Rate (%)",
                "Partner Revenue Share Variable Ini Partner Buy Rate",
                patterns=("stamplibuyrate", "inipartnerbuyrate"),
            )
        ),
        6,
    )
    provided_markup_pct = round(
        money(
            row_value_first(
                row,
                "Stampli Markup (%)",
                "Partner Revenue Share Variable Ini Partner Revenue Share Rate",
                patterns=("stamplimarkup", "inipartnerrevenuesharerate"),
            )
        ),
        6,
    )
    derived_markup_pct = round(customer_markup_pct - stampli_buy_rate_pct, 6) if (customer_markup_pct or stampli_buy_rate_pct) else 0.0
    effective_markup_pct = provided_markup_pct or derived_markup_pct
    markup_base_usd = transaction_volume_usd or mid_market_usd
    calculated_amount = round(markup_base_usd * effective_markup_pct, 2) if markup_base_usd > 0 and effective_markup_pct else 0.0
    gross_markup = round(usd_debited - mid_market_usd, 2) if usd_debited > 0 and mid_market_usd > 0 else 0.0

    amount_matches_sheet = False
    pct_matches_components = False
    if calculated_amount > 0 and provided_amount > 0:
        amount_matches_sheet = abs(calculated_amount - provided_amount) <= STAMPLI_MARKUP_AMOUNT_TOLERANCE
    if provided_markup_pct and derived_markup_pct:
        pct_matches_components = abs(provided_markup_pct - derived_markup_pct) <= STAMPLI_MARKUP_RATE_TOLERANCE

    if provided_amount > 0 and (calculated_amount <= 0 or amount_matches_sheet):
        used_amount = provided_amount
    elif calculated_amount > 0:
        used_amount = calculated_amount
    else:
        used_amount = provided_amount

    return {
        "midMarketUsd": mid_market_usd,
        "usedPeriodAverage": used_period_average,
        "usdDebited": usd_debited,
        "grossMarkup": gross_markup,
        "customerMarkupPct": customer_markup_pct,
        "stampliBuyRatePct": stampli_buy_rate_pct,
        "providedMarkupPct": provided_markup_pct,
        "derivedMarkupPct": derived_markup_pct,
        "effectiveMarkupPct": effective_markup_pct,
        "providedAmount": provided_amount,
        "calculatedAmount": calculated_amount,
        "usedAmount": round(used_amount, 2),
        "amountMatchesSheet": amount_matches_sheet,
        "pctMatchesComponents": pct_matches_components,
    }


def infer_partner(row: dict[str, Any]) -> str:
    for key in [
        "**  Initiator Customer Account ** Partner Group Source",
        "Partner Offline Billing PARTNER",
        "Partner Group With Bank",
    ]:
        partner = normalize_partner_name(row.get(key))
        if partner:
            return partner
    for pattern in [
        "partnergroupsource",
        "partnergroupwithbank",
        "partnerofflinebillingpartner",
    ]:
        partner = normalize_partner_name(row_value_by_patterns(row, pattern))
        if partner:
            return partner
    haystack = " ".join(
        [
            text(row.get("Payer Email")) or text(row_value_by_patterns(row, "payeremail", "payeraccountprimaryemail")),
            text(row.get("Payee Email")) or text(row_value_by_patterns(row, "payeeemail", "payeeaccountprimaryemail")),
            text(row.get("Payer Business Name")) or text(row_value_by_patterns(row, "payerbusinessname", "payeraccountname")),
            text(row.get("Payee Business Name")) or text(row_value_by_patterns(row, "payeebusinessname", "payeeaccountname")),
        ]
    ).lower()
    for needle, partner in PARTNER_PATTERNS:
        if needle in haystack:
            return partner
    return ""


def normalize_revenue_partner(raw_partner: Any) -> str:
    cleaned = text(raw_partner)
    if not cleaned:
        return ""
    return normalize_partner_name(cleaned.split("|")[0].strip())


def normalize_header(value: Any) -> str:
    return "".join(char.lower() for char in text(value) if char.isalnum())


def row_value_by_patterns(row: dict[str, Any], *patterns: str) -> Any:
    normalized_patterns = tuple(pattern.lower() for pattern in patterns if pattern)
    best_value = None
    best_score = -1
    for key, value in row.items():
        normalized_key = normalize_header(key)
        for pattern in normalized_patterns:
            score = -1
            if normalized_key == pattern:
                score = 4
            elif normalized_key.endswith(pattern):
                score = 3
            elif f"{pattern}id" in normalized_key:
                score = 2
            elif pattern in normalized_key:
                score = 1
            if score > best_score:
                best_score = score
                best_value = value
    return best_value


def row_value_first(row: dict[str, Any], *keys: str, patterns: tuple[str, ...] = ()) -> Any:
    for key in keys:
        if key and row.get(key) not in (None, ""):
            return row.get(key)
    if patterns:
        matched = row_value_by_patterns(row, *patterns)
        if matched not in (None, ""):
            return matched
    return None


def parse_dateish_from_row(row: dict[str, Any], *keys: str, patterns: tuple[str, ...] = ()) -> date | None:
    return parse_dateish(row_value_first(row, *keys, patterns=patterns))


def iso_value_from_row(row: dict[str, Any], *keys: str, patterns: tuple[str, ...] = ()) -> str:
    return iso_value(row_value_first(row, *keys, patterns=patterns))


def month_key_from_row(row: dict[str, Any], *keys: str, patterns: tuple[str, ...] = ()) -> str:
    return month_key(row_value_first(row, *keys, patterns=patterns))


def extract_est_revenue(row: dict[str, Any]) -> float:
    return money(
        row.get("Est Revenue")
        or row.get("Estimated Revenue")
        or row_value_by_patterns(row, "estrevenue", "estimatedrevenue", "estrev")
    )


def normalize_fixed_txn_type(raw_type: Any) -> str:
    value = text(raw_type)
    normalized = value.lower().replace("_", " ").replace("-", " ")
    normalized = " ".join(normalized.split())
    mapping = {
        "domestic": "Domestic",
        "Domestic": "Domestic",
        "usd abroad": "USD Abroad",
        "USD Abroad": "USD Abroad",
        "fx": "FX",
        "FX": "FX",
        "cad domestic": "CAD Domestic",
        "CAD Domestic": "CAD Domestic",
        "gbp domestic": "GBP Domestic",
        "eur domestic": "EUR Domestic",
        "aud domestic": "AUD Domestic",
        "incoming us": "Payin",
        "payin": "Payin",
        "payout": "Payout",
    }
    return mapping.get(value, mapping.get(normalized, value))


def normalize_speed(is_rtp: bool, faster_ach: bool) -> str:
    if is_rtp:
        return "RTP"
    if faster_ach:
        return "FasterACH"
    return "Standard"


def normalize_processing_method(txn_type: str, speed_flag: str, methods: set[str]) -> str:
    if txn_type in ("FX", "USD Abroad"):
        return "Wire"
    if txn_type == "CAD Domestic" or "eft" in methods:
        return "EFT"
    if speed_flag == "RTP":
        return "RTP"
    return "ACH"


def normalize_country_code(value: Any) -> str:
    raw = text(value).upper()
    if raw in {"UNITED STATES", "US"}:
        return "US"
    if raw in {"UNITED KINGDOM", "GREAT BRITAIN", "GB", "UK"}:
        return "GB"
    return raw


def is_eea_country(value: Any) -> bool:
    return normalize_country_code(value) in EEA_COUNTRY_CODES


def derive_contract_txn_type(
    raw_txn_type: Any,
    payer_ccy: Any = "",
    payee_ccy: Any = "",
    payer_country: Any = "",
    payee_country: Any = "",
    payment_type: Any = "",
) -> str:
    normalized_raw = normalize_fixed_txn_type(raw_txn_type)
    if normalized_raw in {
        "Domestic",
        "USD Abroad",
        "FX",
        "CAD Domestic",
        "GBP Domestic",
        "EUR Domestic",
        "AUD Domestic",
        "Payin",
        "Payout",
    }:
        return normalized_raw

    normalized_payment_type = normalize_fixed_txn_type(payment_type)
    if normalized_payment_type in {"Payin", "Payout"}:
        return normalized_payment_type

    payer_ccy_text = text(payer_ccy).upper()
    payee_ccy_text = text(payee_ccy).upper()
    payer_country_code = normalize_country_code(payer_country)
    payee_country_code = normalize_country_code(payee_country)

    if payer_ccy_text and payee_ccy_text and payer_ccy_text != payee_ccy_text:
        return "FX"
    if payee_ccy_text == "USD" and payee_country_code and payee_country_code not in US_COUNTRY_CODES:
        return "USD Abroad"
    if payer_ccy_text == payee_ccy_text == "CAD" and payer_country_code in CA_COUNTRY_CODES and payee_country_code in CA_COUNTRY_CODES:
        return "CAD Domestic"
    if payer_ccy_text == payee_ccy_text == "GBP" and payer_country_code in UK_COUNTRY_CODES and payee_country_code in UK_COUNTRY_CODES:
        return "GBP Domestic"
    if payer_ccy_text == payee_ccy_text == "EUR" and is_eea_country(payer_country_code) and is_eea_country(payee_country_code):
        return "EUR Domestic"
    if payer_ccy_text == payee_ccy_text == "AUD" and payer_country_code in AU_COUNTRY_CODES and payee_country_code in AU_COUNTRY_CODES:
        return "AUD Domestic"
    return "Domestic"


def parse_dateish(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value)
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S %Z", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    if len(raw) >= 10:
        try:
            return datetime.fromisoformat(raw[:19].replace("Z", "+00:00")).date()
        except ValueError:
            pass
        try:
            return date.fromisoformat(raw[:10])
        except ValueError:
            pass
    return None


def normalize_registered_account_rows(rows: list[dict[str, Any]], *, prefer_time_created: bool = False) -> list[dict[str, Any]]:
    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        partner = normalize_partner_name(
            row.get("Partner Name")
            or row.get("Partner Group Source")
            or row.get("Partner Group With Bank")
            or row_value_by_patterns(row, "partnername", "partnergroupsource", "partnergroupwithbank")
        )
        account_id = text(
            row.get("Account Id")
            or row.get("Account ID")
            or row.get("ACCOUNT_ID")
            or row.get("customer_account.id")
            or row_value_by_patterns(row, "accountid", "customeraccountid")
        )
        if prefer_time_created:
            join_date = (
                parse_dateish_from_row(
                    row,
                    "Time Created Date",
                    "Time Created Time",
                    "TIME_CREATED_date",
                    patterns=("timecreateddate", "timecreatedtime"),
                )
            )
        else:
            join_date = (
                parse_dateish_from_row(
                    row,
                    "Join Date Time",
                    patterns=("joindatetime",),
                )
            )
        type_defn = text(row.get("Type Defn") or row_value_by_patterns(row, "typedefn"))
        status = text(row.get("Status") or row.get("STATUS") or row_value_by_patterns(row, "status"))
        if not partner or not account_id or not join_date:
            continue
        normalized_rows.append(
            {
                "Partner Name": partner,
                "Account Id": account_id,
                "Join Date Time": join_date.isoformat(),
                "Type Defn": type_defn,
                "Status": status,
            }
        )
    return normalized_rows


def merge_registered_account_rows(*row_sets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for rows in row_sets:
        for row in rows:
            key = (text(row.get("Partner Name")), text(row.get("Account Id")))
            if not key[0] or not key[1]:
                continue
            existing = merged.get(key)
            if not existing:
                merged[key] = dict(row)
                continue
            candidate = dict(existing)
            for field in ("Join Date Time", "Type Defn", "Status"):
                if not text(candidate.get(field)) and text(row.get(field)):
                    candidate[field] = row.get(field)
            merged[key] = candidate
    return list(merged.values())


def build_vba_transaction_activity(rows: list[dict[str, Any]]) -> dict[str, list[date]]:
    activity: dict[str, list[date]] = defaultdict(list)
    for row in rows:
        account_id = text(
            row.get("Account Id")
            or row.get("ACCOUNT_ID")
            or row_value_by_patterns(row, "accountid")
        )
        if not account_id:
            continue
        for candidate in (
            row_value_first(
                row,
                "CC Completed Time",
                "Citi Trx Completed Time",
                "Customer Virtual Bank Account Payments CC COMPLETED AT Date",
                "Customer Virtual Bank Transaction Report Citi COMPLETED AT Date",
                patterns=(
                    "cccompletedtime",
                    "cititrxcompletedtime",
                    "customervirtualbankaccountpaymentscccompletedatdate",
                    "customervirtualbanktransactionreportciticompletedatdate",
                ),
            ),
        ):
            completed_date = parse_dateish(candidate)
            if completed_date:
                activity[account_id].append(completed_date)
    return {account_id: sorted(set(days)) for account_id, days in activity.items()}


def first_existing(*candidates: Path) -> Path | None:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def existing_paths(*candidates: Path) -> list[Path]:
    return [candidate for candidate in candidates if candidate.exists()]


def load_paths(source_dir: Path, stampli_source_dir: Path | None = None) -> Paths:
    offline_dir = source_dir / "dashboard-partner_offline_billing 2"
    revenue_dir = source_dir / "dashboard-partner_revenue_share_reporting_-_fixed_rate"
    stampli_fx_dir = source_dir / "dashboard-stampli_fx_revenue_share_reporting"
    default_stampli_dir = Path.home() / "Downloads" / "Stampli Data"
    direct_stampli_dir = stampli_source_dir if stampli_source_dir and stampli_source_dir.exists() else (default_stampli_dir if default_stampli_dir.exists() else None)

    offline_txns = first_existing(
        offline_dir / "Partner Offline Billing.xlsx",
        offline_dir / "Partner Offline Billing.csv",
    )
    offline_reversals = first_existing(
        offline_dir / "partner_offline_billing_(reversals).xlsx",
        offline_dir / "partner_offline_billing_(reversals).csv",
    )
    offline_accounts = first_existing(
        offline_dir / "All Registered Accounts.xlsx",
        offline_dir / "all_registered_accounts.csv",
        offline_dir / "all_registered_accounts_.csv",
    )
    revenue_txns = existing_paths(
        revenue_dir / "Partner Rev Share V2.xlsx",
        revenue_dir / "partner_revenue_share_v2.csv",
        revenue_dir / "Partner Revenue Share.xlsx",
        revenue_dir / "partner_revenue_share.csv",
    )
    revenue_summary = first_existing(
        revenue_dir / "partner_revenue_summary.csv",
        revenue_dir / "Partner Revenue Summary.xlsx",
    )
    stampli_credit_complete_all = first_existing(
        *((
            direct_stampli_dir / "Credit Complete" / "All Stampli Credit Complete.xlsx",
            direct_stampli_dir / "Credit Complete" / "All Stampli Credit Complete.csv",
        ) if direct_stampli_dir else tuple())
    )
    stampli_domestic_revenue = first_existing(
        *((
            direct_stampli_dir / "Stampli Domestic Revenue.xlsx",
            direct_stampli_dir / "Stampli Domestic Revenue.csv",
        ) if direct_stampli_dir else tuple())
    )
    stampli_usd_abroad_revenue = first_existing(
        *((
            direct_stampli_dir / "Stampli USD Abroad Revenue.xlsx",
            direct_stampli_dir / "Stampli USD Abroad Revenue.csv",
        ) if direct_stampli_dir else tuple())
    )
    stampli_fx_share = first_existing(
        stampli_fx_dir / "stampli_fx_revenue_share.csv",
        stampli_fx_dir / "stampli_fx_revenue_share.xlsx",
        *((
            direct_stampli_dir / "stampli_fx_revenue_share.csv",
            direct_stampli_dir / "stampli_fx_revenue_share.xlsx",
        ) if direct_stampli_dir else tuple())
    )
    stampli_fx_reversals = first_existing(
        stampli_fx_dir / "stampli_fx_revenue_reversal.csv",
        stampli_fx_dir / "stampli_fx_revenue_reversal.xlsx",
        *((
            direct_stampli_dir / "stampli_fx_revenue_reversal.csv",
            direct_stampli_dir / "stampli_fx_revenue_reversal.xlsx",
        ) if direct_stampli_dir else tuple())
    )

    if not offline_txns or not offline_reversals or not revenue_txns:
        raise FileNotFoundError("Missing one or more required billing exports in the provided folder.")

    return Paths(
        offline_txns=offline_txns,
        offline_reversals=offline_reversals,
        offline_accounts=offline_accounts,
        revenue_txns=revenue_txns,
        revenue_summary=revenue_summary,
        stampli_credit_complete_all=stampli_credit_complete_all,
        stampli_domestic_revenue=stampli_domestic_revenue,
        stampli_usd_abroad_revenue=stampli_usd_abroad_revenue,
        stampli_fx_share=stampli_fx_share,
        stampli_fx_reversals=stampli_fx_reversals,
    )


def iter_table_rows(path: Path):
    if path.suffix.lower() == ".csv":
        with path.open(newline="") as handle:
            for row in csv.DictReader(handle):
                yield row
        return

    if path.suffix.lower() == ".xlsx" and path.stat().st_size >= STREAM_XLSX_THRESHOLD_BYTES:
        try:
            yield from iter_large_xlsx_rows(path)
            return
        except Exception:
            pass

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    raw_headers = next(iterator)
    headers = [text(header) for header in raw_headers]
    for values in iterator:
        row = {
            header: value
            for header, value in zip(headers, values)
            if header
        }
        if any(value not in (None, "") for value in row.values()):
            yield row


def read_table(path: Path) -> list[dict[str, Any]]:
    return list(iter_table_rows(path))


def build_stampli_credit_complete_lookup(path: Path | None) -> dict[str, str]:
    if not path:
        return {}
    lookup: dict[str, str] = {}
    for row in iter_table_rows(path):
        payment_id = text(
            row_value_first(
                row,
                "** Payment For Sales DV ** Payment Id",
                "Payment Payment ID",
                "Payment ID",
                patterns=("paymentid",),
            )
        )
        month = month_key(
            row_value_first(
                row,
                "Credit Complete Date",
                "Transaction Lookup Dates Credit Complete Timestamp Time",
                "Transaction Lookup Dates Credit Complete Timestamp Date",
                patterns=("creditcompletedate", "creditcompletetimestamptime", "creditcompletetimestampdate"),
            )
        )
        if payment_id and month:
            lookup[payment_id] = month
    return lookup


def build_stampli_direct_billing(
    credit_complete_path: Path | None,
    domestic_path: Path | None,
    usd_abroad_path: Path | None,
    period: str | None,
) -> tuple[list[dict[str, Any]], list[str], list[dict[str, Any]], dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], dict[str, float]] = defaultdict(lambda: {"txnCount": 0, "totalVolume": 0.0, "directInvoiceAmount": 0.0})
    detail_rows: list[dict[str, Any]] = []
    periods_seen: set[str] = set()
    sources: list[str] = []
    seen_payment_ids: set[str] = set()

    if credit_complete_path:
        sources.append(str(credit_complete_path))
        for row in iter_table_rows(credit_complete_path):
            payment_id = text(
                row_value_first(
                    row,
                    "** Payment For Sales DV ** Payment Id",
                    "Payment Payment ID",
                    "Payment ID",
                    patterns=("paymentid",),
                )
            )
            if not payment_id or payment_id in seen_payment_ids:
                continue
            credit_complete_value = row_value_first(
                row,
                "Credit Complete Date",
                "Transaction Lookup Dates Credit Complete Timestamp Date",
                "Transaction Lookup Dates Credit Complete Timestamp Time",
                patterns=("creditcompletedate", "creditcompletetimestampdate", "creditcompletetimestamptime"),
            )
            month = month_key(credit_complete_value)
            if not month or not matches_period(month, period):
                continue
            txn_type = derive_contract_txn_type(
                row_value_first(row, "** Payment For Sales DV ** Txn Type (Dom/Fx/Abroad..", "Txn Type", patterns=("txntype",)),
                payer_ccy=row_value_first(row, "** Payment For Sales DV ** Payer Amount Currency", patterns=("payeramountcurrency",)),
                payee_ccy=row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",)),
                payer_country=row_value_first(row, "** Payment For Sales DV ** Payer Country", "Payer Country", patterns=("payercountry",)),
                payee_country=row_value_first(row, "** Payment For Sales DV ** Payee Country", "Payee Country", patterns=("payeecountry",)),
                payment_type=row_value_first(row, "** Payment For Sales DV ** Payment Type", "Payment Type", patterns=("paymenttype",)),
            )
            if txn_type not in {"Domestic", "USD Abroad"}:
                continue
            seen_payment_ids.add(payment_id)
            periods_seen.add(month)
            speed_flag = normalize_speed(False, text(row.get("Is Faster Ach")) == "FasterACH")
            processing_method = normalize_processing_method(txn_type, speed_flag, set())
            key = (month, txn_type, speed_flag, processing_method)
            grouped[key]["txnCount"] += 1
            grouped[key]["totalVolume"] += money(
                row_value_first(
                    row,
                    "Payment USD Equivalent Amount",
                    "** Payment For Sales DV ** Total USD Amount Number",
                    "** Payment For Sales DV ** USD Amount Number",
                    patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"),
                )
            )
            grouped[key]["directInvoiceAmount"] += money(
                row.get("Fees")
                or row_value_by_patterns(row, "fees", "fee")
                or extract_est_revenue(row)
            )
            detail_rows.append(
                {
                    "detailCategory": "transaction",
                    "detailSource": "stampli_credit_complete_billing",
                    "partner": STAMPLI_FX_PARTNER,
                    "period": month,
                    "paymentId": payment_id,
                    "txnType": txn_type,
                    "speedFlag": speed_flag,
                    "processingMethod": processing_method,
                    "payerFunding": "",
                    "payeeFunding": "",
                    "payerCcy": "USD",
                    "payeeCcy": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))) or "USD",
                    "payerCountry": "",
                    "payeeCountry": normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payee Country", "Payee Country", patterns=("payeecountry",))),
                    "accountId": text(row_value_first(row, "** Payment For Sales DV ** Payer Account ID", "Account ID", patterns=("payeraccountid", "accountid"))),
                    "paymentType": txn_type,
                    "submissionDate": iso_value(row_value_first(row, "Date of Payment Submission", "** Payment For Sales DV ** Time Created Date", patterns=("dateofpaymentsubmission", "timecreateddate"))),
                    "creditCompleteDate": iso_value(credit_complete_value),
                    "payerEmail": text(row.get("Payer Email")),
                    "payerBusinessName": text(row.get("Payer Business Name")),
                    "payeeEmail": text(row.get("Payee Email")),
                    "payeeBusinessName": text(row.get("Payee Business Name")),
                    "payeeAmountCurrency": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))) or "USD",
                    "payeeAmount": round(money(row_value_first(row, "** Payment For Sales DV ** Payee Amount Number", "Foreign Currency Amount (Payee Amount Number)", patterns=("payeeamountnumber",))), 2),
                    "usdAmountDebited": round(money(row_value_first(row, "USD Amount Debited to the Customer", "** Payment For Sales DV ** Total USD Amount Number", "** Payment For Sales DV ** USD Amount Number", patterns=("usdamountdebited", "totalusdamountnumber", "usdamountnumber"))), 2),
                    "paymentUsdEquivalentAmount": round(money(row_value_first(row, "Payment USD Equivalent Amount", "** Payment For Sales DV ** Total USD Amount Number", "** Payment For Sales DV ** USD Amount Number", patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"))), 2),
                    "directFeeAmount": round(
                        money(
                            row.get("Fees")
                            or row_value_by_patterns(row, "fees", "fee")
                            or extract_est_revenue(row)
                        ),
                        2,
                    ),
                }
            )

        output: list[dict[str, Any]] = []
        for (month, txn_type, speed_flag, processing_method), aggregate in sorted(grouped.items()):
            txn_count = int(aggregate["txnCount"])
            total_volume = round(aggregate["totalVolume"], 2)
            direct_invoice_amount = round(aggregate["directInvoiceAmount"], 2)
            direct_invoice_rate = round(direct_invoice_amount / txn_count, 4) if txn_count else 0.0
            output.append(
                {
                    "period": month,
                    "partner": STAMPLI_FX_PARTNER,
                    "txnType": txn_type,
                    "speedFlag": speed_flag,
                    "minAmt": txn_count,
                    "maxAmt": txn_count,
                    "payerFunding": "",
                    "payeeFunding": "",
                    "payerCcy": "USD",
                    "payeeCcy": "USD",
                    "payerCountry": "",
                    "payeeCountry": "",
                    "processingMethod": processing_method,
                    "txnCount": txn_count,
                    "totalVolume": total_volume,
                    "customerRevenue": direct_invoice_amount,
                    "avgTxnSize": round(total_volume / txn_count, 2) if txn_count else 0.0,
                    "directInvoiceAmount": direct_invoice_amount,
                    "directInvoiceRate": direct_invoice_rate,
                    "directInvoiceSource": "stampli_credit_complete_billing",
                }
            )

        meta = {
            "sources": sources,
            "paymentIdsImported": len(detail_rows),
            "periods": sorted(periods_seen),
        }
        return output, sorted(periods_seen), detail_rows, meta

    configs = [
        {
            "path": domestic_path,
            "txnType": "Domestic",
            "processingMethod": "ACH",
            "detailSource": "stampli_domestic_revenue",
            "speedResolver": lambda row: normalize_speed(False, text(row.get("Is Faster Ach")) == "FasterACH"),
        },
        {
            "path": usd_abroad_path,
            "txnType": "USD Abroad",
            "processingMethod": "Wire",
            "detailSource": "stampli_usd_abroad_revenue",
            "speedResolver": lambda row: "Standard",
        },
    ]

    for config in configs:
        path = config["path"]
        if not path:
            continue
        sources.append(str(path))
        for row in iter_table_rows(path):
            credit_complete_value = row_value_first(
                row,
                "Credit Complete Date",
                "Transaction Lookup Dates Credit Complete Timestamp Date",
                "Transaction Lookup Dates Credit Complete Timestamp Time",
                patterns=("creditcompletedate", "creditcompletetimestampdate", "creditcompletetimestamptime"),
            )
            month = month_key(credit_complete_value)
            if not month:
                continue
            if not matches_period(month, period):
                continue
            payment_id = text(
                row_value_first(
                    row,
                    "** Payment For Sales DV ** Payment Id",
                    "Payment Payment ID",
                    "Payment ID",
                    patterns=("paymentid",),
                )
            )
            if not payment_id or payment_id in seen_payment_ids:
                continue
            seen_payment_ids.add(payment_id)
            periods_seen.add(month)
            speed_flag = config["speedResolver"](row)
            key = (month, config["txnType"], speed_flag, config["processingMethod"])
            grouped[key]["txnCount"] += 1
            grouped[key]["totalVolume"] += money(
                row_value_first(
                    row,
                    "Payment USD Equivalent Amount",
                    "** Payment For Sales DV ** Total USD Amount Number",
                    "** Payment For Sales DV ** USD Amount Number",
                    patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"),
                )
            )
            grouped[key]["directInvoiceAmount"] += money(
                row.get("Fees")
                or row_value_by_patterns(row, "fees", "fee")
                or extract_est_revenue(row)
            )
            detail_rows.append(
                {
                    "detailCategory": "transaction",
                    "detailSource": config["detailSource"],
                    "partner": STAMPLI_FX_PARTNER,
                    "period": month,
                    "paymentId": payment_id,
                    "txnType": config["txnType"],
                    "speedFlag": speed_flag,
                    "processingMethod": config["processingMethod"],
                    "payerFunding": "",
                    "payeeFunding": "",
                    "payerCcy": "USD",
                    "payeeCcy": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))) or "USD",
                    "payerCountry": "",
                    "payeeCountry": normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payee Country", "Payee Country", patterns=("payeecountry",))),
                    "accountId": text(row_value_first(row, "** Payment For Sales DV ** Payer Account ID", "Account ID", patterns=("payeraccountid", "accountid"))),
                    "paymentType": config["txnType"],
                    "submissionDate": iso_value(row_value_first(row, "Date of Payment Submission", "** Payment For Sales DV ** Time Created Date", patterns=("dateofpaymentsubmission", "timecreateddate"))),
                    "creditCompleteDate": iso_value(credit_complete_value),
                    "payerEmail": text(row.get("Payer Email")),
                    "payerBusinessName": text(row.get("Payer Business Name")),
                    "payeeEmail": text(row.get("Payee Email")),
                    "payeeBusinessName": text(row.get("Payee Business Name")),
                    "payeeAmountCurrency": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))) or "USD",
                    "payeeAmount": round(money(row_value_first(row, "** Payment For Sales DV ** Payee Amount Number", "Foreign Currency Amount (Payee Amount Number)", patterns=("payeeamountnumber",))), 2),
                    "usdAmountDebited": round(money(row_value_first(row, "USD Amount Debited to the Customer", "** Payment For Sales DV ** Total USD Amount Number", "** Payment For Sales DV ** USD Amount Number", patterns=("usdamountdebited", "totalusdamountnumber", "usdamountnumber"))), 2),
                    "paymentUsdEquivalentAmount": round(money(row_value_first(row, "Payment USD Equivalent Amount", "** Payment For Sales DV ** Total USD Amount Number", "** Payment For Sales DV ** USD Amount Number", patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"))), 2),
                    "directFeeAmount": round(
                        money(
                            row.get("Fees")
                            or row_value_by_patterns(row, "fees", "fee")
                            or extract_est_revenue(row)
                        ),
                        2,
                    ),
                }
            )

    output: list[dict[str, Any]] = []
    for (month, txn_type, speed_flag, processing_method), aggregate in sorted(grouped.items()):
        txn_count = int(aggregate["txnCount"])
        total_volume = round(aggregate["totalVolume"], 2)
        direct_invoice_amount = round(aggregate["directInvoiceAmount"], 2)
        direct_invoice_rate = round(direct_invoice_amount / txn_count, 4) if txn_count else 0.0
        output.append(
            {
                "period": month,
                "partner": STAMPLI_FX_PARTNER,
                "txnType": txn_type,
                "speedFlag": speed_flag,
                "minAmt": txn_count,
                "maxAmt": txn_count,
                "payerFunding": "",
                "payeeFunding": "",
                "payerCcy": "USD",
                "payeeCcy": "USD",
                "payerCountry": "",
                "payeeCountry": "",
                "processingMethod": processing_method,
                "txnCount": txn_count,
                "totalVolume": total_volume,
                "customerRevenue": direct_invoice_amount,
                "avgTxnSize": round(total_volume / txn_count, 2) if txn_count else 0.0,
                "directInvoiceAmount": direct_invoice_amount,
                "directInvoiceRate": direct_invoice_rate,
                "directInvoiceSource": "stampli_direct_billing",
            }
        )

    meta = {
        "sources": sources,
        "paymentIdsImported": len(detail_rows),
        "periods": sorted(periods_seen),
    }
    return output, sorted(periods_seen), detail_rows, meta


def build_stampli_direct_reversal_rows(path: Path, period: str | None) -> tuple[list[dict[str, Any]], list[str], list[dict[str, Any]], dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], dict[str, float]] = defaultdict(lambda: {"txnCount": 0, "totalVolume": 0.0, "directInvoiceAmount": 0.0})
    detail_rows: list[dict[str, Any]] = []
    periods_seen: set[str] = set()
    sources = [str(path)]

    for row in iter_table_rows(path):
        refund_complete_value = row_value_first(
            row,
            "Refund Complete Date",
            "Refund Completed Date",
            "Transaction Lookup Dates Refund Complete Timestamp Date",
            "Transaction Lookup Dates Refund Complete Timestamp Time",
            patterns=("refundcompletedate", "refundcompleteddate", "refundcompletetimestampdate", "refundcompletetimestamptime"),
        )
        month = month_key(refund_complete_value)
        if not month or not matches_period(month, period):
            continue
        payment_id = text(
            row_value_first(
                row,
                "Payment Payment ID",
                "Payment ID",
                patterns=("paymentpaymentid", "paymentid"),
            )
        )
        if not payment_id:
            continue
        periods_seen.add(month)
        txn_type = "USD Abroad"
        speed_flag = "Standard"
        processing_method = "Wire"
        direct_invoice_amount = -abs(
            money(
                row.get("Fees")
                or row_value_by_patterns(row, "fees", "fee")
                or extract_est_revenue(row)
            )
        )
        total_volume = money(
            row_value_first(
                row,
                "Payment USD Equivalent Amount",
                "** Payment For Sales DV ** Total USD Amount Number",
                "** Payment For Sales DV ** USD Amount Number",
                "Total USD Amount Number",
                patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"),
            )
        )
        key = (month, txn_type, speed_flag, processing_method)
        grouped[key]["txnCount"] += 1
        grouped[key]["totalVolume"] += total_volume
        grouped[key]["directInvoiceAmount"] += direct_invoice_amount
        detail_rows.append(
            {
                "detailCategory": "reversal",
                "detailSource": "stampli_usd_abroad_reversal",
                "partner": STAMPLI_FX_PARTNER,
                "period": month,
                "paymentId": payment_id,
                "txnType": txn_type,
                "speedFlag": speed_flag,
                "processingMethod": processing_method,
                "payerFunding": "",
                "payeeFunding": text(row_value_first(row, "** Payment For Sales DV ** Payee Funding Method Type", "Payee Funding Method", patterns=("payeefundingmethodtype", "payeefundingmethod"))),
                "payerCcy": text(row_value_first(row, "** Payment For Sales DV ** Payer Amount Currency", "Payer Amount Currency", patterns=("payeramountcurrency",))) or "USD",
                "payeeCcy": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))) or "USD",
                "payerCountry": normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payer Country", "Payer Country", patterns=("payercountry",))),
                "payeeCountry": normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payee Country", "Payee Country", patterns=("payeecountry",))),
                "accountId": text(row_value_first(row, "** Payment For Sales DV ** Payer Account ID", "Account ID", patterns=("payeraccountid", "accountid"))),
                "paymentType": txn_type,
                "submissionDate": iso_value(row_value_first(row, "Date of Payment Submission", "** Payment For Sales DV ** Time Created Date", "Time Created Date", patterns=("dateofpaymentsubmission", "timecreateddate"))),
                "reversalDate": iso_value(refund_complete_value),
                "payerEmail": text(row.get("Payer Email") or row_value_by_patterns(row, "payeraccountprimaryemail")),
                "payerBusinessName": text(row.get("Payer Business Name") or row_value_by_patterns(row, "payeraccountname")),
                "payeeEmail": text(row.get("Payee Email") or row_value_by_patterns(row, "payeeaccountprimaryemail")),
                "payeeBusinessName": text(row.get("Payee Business Name") or row_value_by_patterns(row, "payeeaccountname")),
                "payeeAmountCurrency": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))) or "USD",
                "payeeAmount": round(money(row_value_first(row, "** Payment For Sales DV ** Payee Amount Number", "Foreign Currency Amount (Payee Amount Number)", patterns=("payeeamountnumber",))), 2),
                "paymentUsdEquivalentAmount": round(total_volume, 2),
                "directFeeAmount": round(direct_invoice_amount, 2),
            }
        )

    output: list[dict[str, Any]] = []
    for (month, txn_type, speed_flag, processing_method), aggregate in sorted(grouped.items()):
        txn_count = int(aggregate["txnCount"])
        total_volume = round(aggregate["totalVolume"], 2)
        direct_invoice_amount = round(aggregate["directInvoiceAmount"], 2)
        direct_invoice_rate = round(direct_invoice_amount / txn_count, 4) if txn_count else 0.0
        output.append(
            {
                "period": month,
                "partner": STAMPLI_FX_PARTNER,
                "txnType": txn_type,
                "speedFlag": speed_flag,
                "minAmt": txn_count,
                "maxAmt": txn_count,
                "payerFunding": "",
                "payeeFunding": "",
                "payerCcy": "USD",
                "payeeCcy": "USD",
                "payerCountry": "",
                "payeeCountry": "",
                "processingMethod": processing_method,
                "txnCount": txn_count,
                "totalVolume": total_volume,
                "customerRevenue": direct_invoice_amount,
                "avgTxnSize": round(total_volume / txn_count, 2) if txn_count else 0.0,
                "directInvoiceAmount": direct_invoice_amount,
                "directInvoiceRate": direct_invoice_rate,
                "directInvoiceSource": "stampli_usd_abroad_reversal",
            }
        )

    meta = {
        "sources": sources,
        "paymentIdsImported": len(detail_rows),
        "periods": sorted(periods_seen),
    }
    return output, sorted(periods_seen), detail_rows, meta


def build_offline_transactions(path: Path, period: str | None) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, list[date]], dict[tuple[str, str], set[date]], list[str], list[dict[str, Any]]]:
    payment_aggs: dict[str, dict[str, Any]] = {}
    unmatched_payment_ids: set[str] = set()
    unmatched_examples: Counter[str] = Counter()
    periods_seen: set[str] = set()

    for row in iter_table_rows(path):
        credit_complete_value = row_value_first(
            row,
            "Credit Complete Date",
            "Transaction Lookup Dates Credit Complete Timestamp Date",
            "Transaction Lookup Dates Credit Complete Timestamp Time",
            patterns=(
                "creditcompletedate",
                "creditcompletetimestampdate",
                "creditcompletetimestampdate",
                "creditcompletetimestamptime",
            ),
        )
        month = month_key(credit_complete_value)
        if not month:
            continue
        if not matches_period(month, period):
            continue
        payment_id = text(
            row_value_first(
                row,
                "Payment Payment ID",
                "Payment ID",
                patterns=("paymentpaymentid", "paymentid"),
            )
        )
        if not payment_id:
            continue
        periods_seen.add(month)

        agg = payment_aggs.get(payment_id)
        if not agg:
            agg = {
                "month": month,
                "partner": infer_partner(row),
                "row": row,
                "is_rtp": boolish(row.get("Extra Info IsRTP") or row_value_by_patterns(row, "extrainfoisrtp", "isrtp")),
                "faster_ach": text(row.get("Extra Info Is Faster Ach") or row_value_by_patterns(row, "extrainfoisfasterach", "isfasterach")) == "FasterACH",
                "methods": {
                    text(row.get("Transaction Processing Method") or row_value_by_patterns(row, "transactionprocessingmethod")).lower()
                }
                if text(row.get("Transaction Processing Method") or row_value_by_patterns(row, "transactionprocessingmethod"))
                else set(),
            }
            payment_aggs[payment_id] = agg
        else:
            if not agg["partner"]:
                agg["partner"] = infer_partner(row)
            agg["is_rtp"] = agg["is_rtp"] or boolish(row.get("Extra Info IsRTP") or row_value_by_patterns(row, "extrainfoisrtp", "isrtp"))
            agg["faster_ach"] = agg["faster_ach"] or text(row.get("Extra Info Is Faster Ach") or row_value_by_patterns(row, "extrainfoisfasterach", "isfasterach")) == "FasterACH"
            method = text(row.get("Transaction Processing Method") or row_value_by_patterns(row, "transactionprocessingmethod")).lower()
            if method:
                agg["methods"].add(method)

    grouped: dict[tuple[Any, ...], dict[str, float]] = defaultdict(lambda: {"txnCount": 0, "totalVolume": 0.0, "estRevenue": 0.0})
    account_activity: dict[str, list[date]] = defaultdict(list)
    settlement_days: dict[tuple[str, str], set[date]] = defaultdict(set)
    detail_rows: list[dict[str, Any]] = []

    for payment_id, agg in payment_aggs.items():
        first = agg["row"]
        partner = agg["partner"]
        if not partner:
            unmatched_payment_ids.add(payment_id)
            example = " / ".join(
                filter(
                    None,
                    [
                        text(first.get("Payer Business Name")),
                        text(first.get("Payee Business Name")),
                    ],
                )
            ) or "Unknown"
            unmatched_examples[example] += 1
            continue

        month = agg["month"]
        raw_txn_type = row_value_first(
            first,
            "** Payment For Sales DV ** Txn Type (Dom/Fx/Abroad..",
            "Txn Type",
            "Payment Type",
            patterns=("txntype", "paymenttype"),
        )
        payment_type = text(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payment Type",
                "Payment Type",
                patterns=("paymenttype",),
            )
        )
        payer_funding = title_case_funding(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payer Funding Method Type",
                "Payer Funding Method",
                patterns=("payerfundingmethodtype", "payerfundingmethod"),
            )
        )
        payee_funding = title_case_funding(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payee Funding Method Type",
                "Payee Funding Method",
                patterns=("payeefundingmethodtype", "payeefundingmethod"),
            )
        )
        payer_ccy = text(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payer Amount Currency",
                "Payer Amount Currency",
                patterns=("payeramountcurrency",),
            )
        ) or "USD"
        payee_ccy = text(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payee Amount Currency",
                "Currency (Payee Amount Currency)",
                "Payee Amount Currency",
                patterns=("payeeamountcurrency",),
            )
        ) or "USD"
        payer_country = normalize_country_code(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payer Country",
                "Payer Country",
                patterns=("payercountry",),
            )
        )
        payee_country = normalize_country_code(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payee Country",
                "Payee Country",
                patterns=("payeecountry",),
            )
        )
        txn_type = derive_contract_txn_type(
            raw_txn_type,
            payer_ccy=payer_ccy,
            payee_ccy=payee_ccy,
            payer_country=payer_country,
            payee_country=payee_country,
            payment_type=payment_type,
        )
        speed_flag = normalize_speed(agg["is_rtp"], agg["faster_ach"])
        processing_method = normalize_processing_method(txn_type, speed_flag, agg["methods"])
        account_id = text(
            row_value_first(
                first,
                "** Payment For Sales DV ** Payer Account ID",
                "Account ID",
                patterns=("payeraccountid", "accountid"),
            )
        )
        submission_value = row_value_first(
            first,
            "Date of Payment Submission",
            "** Payment For Sales DV ** Time Created Date",
            patterns=("dateofpaymentsubmission", "timecreateddate"),
        )
        credit_complete_value = row_value_first(
            first,
            "Credit Complete Date",
            "Transaction Lookup Dates Credit Complete Timestamp Date",
            "Transaction Lookup Dates Credit Complete Timestamp Time",
            patterns=(
                "creditcompletedate",
                "creditcompletetimestampdate",
                "creditcompletetimestamptime",
            ),
        )
        payer_email = text(first.get("Payer Email") or row_value_by_patterns(first, "payeremail", "payeraccountprimaryemail"))
        payer_business_name = text(first.get("Payer Business Name") or row_value_by_patterns(first, "payerbusinessname", "payeraccountname"))
        payee_email = text(first.get("Payee Email") or row_value_by_patterns(first, "payeeemail", "payeeaccountprimaryemail"))
        payee_business_name = text(first.get("Payee Business Name") or row_value_by_patterns(first, "payeebusinessname", "payeeaccountname"))
        payee_amount = round(
            money(
                row_value_first(
                    first,
                    "** Payment For Sales DV ** Payee Amount Number",
                    "Foreign Currency Amount (Payee Amount Number)",
                    patterns=("payeeamountnumber",),
                )
            ),
            2,
        )
        usd_amount_debited = round(
            money(
                row_value_first(
                    first,
                    "USD Amount Debited to the Customer",
                    "** Payment For Sales DV ** Total USD Amount Number",
                    "** Payment For Sales DV ** USD Amount Number",
                    patterns=("usdamountdebited", "totalusdamountnumber", "usdamountnumber"),
                )
            ),
            2,
        )
        payment_usd_equivalent_amount = round(
            money(
                row_value_first(
                    first,
                    "Payment USD Equivalent Amount",
                    "** Payment For Sales DV ** Total USD Amount Number",
                    "** Payment For Sales DV ** USD Amount Number",
                    patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"),
                )
            ),
            2,
        )
        est_revenue = round(extract_est_revenue(first), 2)
        key = (
            partner,
            month,
            txn_type,
            speed_flag,
            payer_funding,
            payee_funding,
            payer_ccy,
            payee_ccy,
            payer_country,
            payee_country,
            processing_method,
        )
        grouped[key]["txnCount"] += 1
        grouped[key]["totalVolume"] += payment_usd_equivalent_amount
        grouped[key]["estRevenue"] += est_revenue

        tx_date = parse_dateish(credit_complete_value)
        if account_id and tx_date:
            account_activity[account_id].append(tx_date)
        if tx_date:
            settlement_days[(partner, month)].add(tx_date)

        detail_rows.append(
            {
                "detailCategory": "transaction",
                "detailSource": "offline_billing",
                "partner": partner,
                "period": month,
                "paymentId": payment_id,
                "txnType": txn_type,
                "speedFlag": speed_flag,
                "processingMethod": processing_method,
                "payerFunding": payer_funding,
                "payeeFunding": payee_funding,
                "payerCcy": payer_ccy,
                "payeeCcy": payee_ccy,
                "payerCountry": payer_country,
                "payeeCountry": payee_country,
                "accountId": account_id,
                "paymentType": payment_type,
                "submissionDate": iso_value(submission_value),
                "creditCompleteDate": iso_value(credit_complete_value),
                "creditCompleteMonth": month,
                "payerEmail": payer_email,
                "payerBusinessName": payer_business_name,
                "payeeEmail": payee_email,
                "payeeBusinessName": payee_business_name,
                "creditRail": text(first.get("Credit Rail") or row_value_by_patterns(first, "creditrail")),
                "transactionProcessingMethodRaw": text(first.get("Transaction Processing Method") or row_value_by_patterns(first, "transactionprocessingmethod")),
                "fundingMethodUsed": text(first.get("Funding Method Used") or row_value_by_patterns(first, "fundingmethodused")),
                "payeeAmountCurrency": payee_ccy,
                "payeeAmount": payee_amount,
                "usdAmountDebited": usd_amount_debited,
                "paymentUsdEquivalentAmount": payment_usd_equivalent_amount,
                "estRevenue": est_revenue,
                "txnTypeRaw": text(raw_txn_type),
                "isRTP": agg["is_rtp"],
                "isFasterAch": agg["faster_ach"],
                "partnerGroupSource": text(first.get("**  Initiator Customer Account ** Partner Group Source") or row_value_by_patterns(first, "partnergroupsource")),
                "initiatorStatus": text(first.get("** Payment For Sales DV ** Initiator Status") or row_value_by_patterns(first, "initiatorstatus")),
                "typeDefn": text(first.get("**  Initiator Customer Account ** Type Defn") or row_value_by_patterns(first, "typedefn")),
            }
        )

    output: list[dict[str, Any]] = []
    for key, aggregate in sorted(grouped.items()):
        (
            partner,
            month,
            txn_type,
            speed_flag,
            payer_funding,
            payee_funding,
            payer_ccy,
            payee_ccy,
            payer_country,
            payee_country,
            processing_method,
        ) = key
        txn_count = int(aggregate["txnCount"])
        total_volume = round(aggregate["totalVolume"], 2)
        est_revenue = round(aggregate["estRevenue"], 2)
        output.append(
            {
                "period": month,
                "partner": partner,
                "txnType": txn_type,
                "speedFlag": speed_flag,
                "minAmt": txn_count,
                "maxAmt": txn_count,
                "payerFunding": payer_funding,
                "payeeFunding": payee_funding,
                "payerCcy": payer_ccy,
                "payeeCcy": payee_ccy,
                "payerCountry": payer_country,
                "payeeCountry": payee_country,
                "processingMethod": processing_method,
                "txnCount": txn_count,
                "totalVolume": total_volume,
                "customerRevenue": 0.0,
                "estRevenue": est_revenue,
                "avgTxnSize": round(total_volume / txn_count, 2) if txn_count else 0.0,
            }
        )

    meta = {
        "paymentIdsProcessed": len(payment_aggs),
        "paymentIdsImported": len(output),
        "unmatchedPaymentIds": len(unmatched_payment_ids),
        "unmatchedExamples": unmatched_examples.most_common(15),
        "partners": Counter(row["partner"] for row in output),
    }
    return output, meta, account_activity, settlement_days, sorted(periods_seen), detail_rows


def build_offline_reversals(
    path: Path,
    period: str | None,
    account_partner_lookup: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any], list[str], list[dict[str, Any]]]:
    grouped: Counter[tuple[str, str, str]] = Counter()
    unmatched_examples: Counter[str] = Counter()
    periods_seen: set[str] = set()
    seen_payment_ids: set[str] = set()
    detail_rows: list[dict[str, Any]] = []
    period_fallback_count = 0
    partner_lookup = {text(account_id): normalize_partner_name(partner) for account_id, partner in (account_partner_lookup or {}).items() if text(account_id) and normalize_partner_name(partner)}

    for row in iter_table_rows(path):
        reversal_value = row_value_first(
            row,
            "Refund Complete Date",
            "Refund Completed Date",
            "Transaction Lookup Dates Refund Complete Timestamp Date",
            "Transaction Lookup Dates Refund Complete Timestamp Time",
            patterns=(
                "refundcompletedate",
                "refundcompleteddate",
                "refundcompletetimestampdate",
                "refundcompletetimestamptime",
            ),
        )
        month = month_key(
            row_value_first(
                row,
                "Refund Complete Date",
                "Refund Completed Date",
                "Transaction Lookup Dates Refund Complete Timestamp Date",
                "Transaction Lookup Dates Refund Complete Timestamp Time",
                patterns=(
                    "refundcompletedate",
                    "refundcompleteddate",
                    "refundcompletetimestampdate",
                    "refundcompletetimestamptime",
                ),
            )
            or reversal_value
        )
        if not month:
            continue
        if not matches_period(month, period):
            continue
        periods_seen.add(month)
        payment_id = text(row.get("Payment ID") or row_value_by_patterns(row, "paymentid"))
        if not payment_id:
            continue
        if payment_id and payment_id in seen_payment_ids:
            continue
        if payment_id:
            seen_payment_ids.add(payment_id)
        partner = infer_partner(row)
        if not partner:
            account_id = text(
                row_value_first(
                    row,
                    "** Payment For Sales DV ** Payer Account ID",
                    "Account ID",
                    patterns=("payeraccountid", "accountid"),
                )
            )
            partner = partner_lookup.get(account_id, "")
        if not partner:
            example = " / ".join(
                filter(
                    None,
                    [
                        text(row.get("Payer Business Name")),
                        text(row.get("Payee Business Name")),
                    ],
                )
            ) or "Unknown"
            unmatched_examples[example] += 1
            continue
        priority = text(
            row.get("Payment Priority")
            or row_value_by_patterns(
                row,
                "paymentpriority",
                "payerfundingmethodtype",
                "payerfundingmethod",
                "fundingmethodused",
            )
        ).lower()
        payer_funding = title_case_funding(priority.split(" - ")[0] if " - " in priority else priority)
        grouped[(month, partner, payer_funding)] += 1
        payer_ccy = text(row_value_first(row, "** Payment For Sales DV ** Payer Amount Currency", "Payer Amount Currency", patterns=("payeramountcurrency",))) or "USD"
        payee_ccy = text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", "Payee Amount Currency", patterns=("payeeamountcurrency",))) or "USD"
        payer_country = normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payer Country", "Payer Country", patterns=("payercountry",)))
        payee_country = normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payee Country", "Payee Country", patterns=("payeecountry",)))
        account_id = text(row_value_first(row, "** Payment For Sales DV ** Payer Account ID", "Account ID", patterns=("payeraccountid", "accountid")))
        submission_value = row_value_first(row, "Date of Payment Submission", "** Payment For Sales DV ** Time Created Date", patterns=("dateofpaymentsubmission", "timecreateddate"))
        payer_email = text(row.get("Payer Email") or row_value_by_patterns(row, "payeremail", "payeraccountprimaryemail"))
        payer_business_name = text(row.get("Payer Business Name") or row_value_by_patterns(row, "payerbusinessname", "payeraccountname"))
        payee_email = text(row.get("Payee Email") or row_value_by_patterns(row, "payeeemail", "payeeaccountprimaryemail"))
        payee_business_name = text(row.get("Payee Business Name") or row_value_by_patterns(row, "payeebusinessname", "payeeaccountname"))
        payee_amount = round(
            money(
                row.get("Foreign Currency Amount (Payee Amount Number)")
                or row_value_by_patterns(row, "payeeamountnumber")
            ),
            2,
        )
        usd_amount_debited = round(
            money(
                row.get("USD Amount Debited to the Customer")
                or row_value_by_patterns(row, "usdamountdebited", "payeramountnumber")
            ),
            2,
        )
        payment_usd_equivalent_amount = round(
            money(
                row.get("Payment USD Equivalent Amount")
                or row_value_by_patterns(row, "paymentusdequivalentamount", "usdamountnumber")
            ),
            2,
        )
        detail_rows.append(
            {
                "detailCategory": "reversal",
                "detailSource": "offline_reversal",
                "partner": partner,
                "period": month,
                "paymentId": payment_id,
                "txnType": "",
                "speedFlag": "",
                "processingMethod": "",
                "payerFunding": payer_funding,
                "payeeFunding": "",
                "payerCcy": payer_ccy,
                "payeeCcy": payee_ccy,
                "payerCountry": payer_country,
                "payeeCountry": payee_country,
                "accountId": account_id,
                "paymentType": text(row.get("Payment Type") or row_value_by_patterns(row, "paymenttype", "txntype")),
                "submissionDate": iso_value(submission_value),
                "reversalDate": iso_value(reversal_value),
                "payerEmail": payer_email,
                "payerBusinessName": payer_business_name,
                "payeeEmail": payee_email,
                "payeeBusinessName": payee_business_name,
                "paymentPriority": text(row.get("Payment Priority") or row_value_by_patterns(row, "paymentpriority")),
                "payeeAmountCurrency": payee_ccy,
                "payeeAmount": payee_amount,
                "usdAmountDebited": usd_amount_debited,
                "paymentUsdEquivalentAmount": payment_usd_equivalent_amount,
            }
        )

    output = [
        {
            "period": month,
            "partner": partner,
            "payerFunding": payer_funding,
            "reversalCount": count,
        }
        for (month, partner, payer_funding), count in sorted(grouped.items())
    ]
    return output, {
        "partners": Counter(row["partner"] for row in output),
        "unmatchedExamples": unmatched_examples.most_common(15),
        "periodFallbackCount": period_fallback_count,
    }, sorted(periods_seen), detail_rows


def revenue_payment_id(row: dict[str, Any]) -> str:
    return text(
        row_value_first(
            row,
            "Partner Revenue Share Fixed Payment ID",
            "Payment Payment ID",
            "Payment ID",
            patterns=("paymentid",),
        )
    )


def dedupe_revenue_record_list(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered: list[dict[str, Any]] = []
    seen_payment_ids: set[str] = set()
    for row in rows:
        payment_id = revenue_payment_id(row)
        if payment_id and payment_id in seen_payment_ids:
            continue
        if payment_id:
            seen_payment_ids.add(payment_id)
        ordered.append(row)
    return ordered


def dedupe_revenue_rows(row_sets: list[list[dict[str, Any]]]) -> list[dict[str, Any]]:
    ordered: list[dict[str, Any]] = []
    seen_payment_ids: set[str] = set()
    for rows in row_sets:
        for row in dedupe_revenue_record_list(rows):
            payment_id = revenue_payment_id(row)
            if payment_id and payment_id in seen_payment_ids:
                continue
            if payment_id:
                seen_payment_ids.add(payment_id)
            ordered.append(row)
    return ordered


def build_revenue_detail_transactions(rows: list[dict[str, Any]], period: str | None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows = dedupe_revenue_record_list(rows)
    grouped: dict[tuple[Any, ...], dict[str, float]] = defaultdict(
        lambda: {"txnCount": 0, "totalVolume": 0.0, "customerRevenue": 0.0, "estRevenue": 0.0}
    )
    detail_rows: list[dict[str, Any]] = []

    for row in rows:
        credit_complete_value = row_value_first(
            row,
            "Credit Complete Date",
            "Transaction Lookup Dates Credit Complete Timestamp Time",
            "Transaction Lookup Dates Credit Complete Timestamp Date",
            "Credit Complete Timestamp Date",
            patterns=("creditcompletedate", "creditcompletetimestamptime", "creditcompletetimestampdate"),
        )
        month = month_key(credit_complete_value)
        if not month:
            continue
        if not matches_period(month, period):
            continue
        partner = normalize_revenue_partner(
            row.get("Partner Group With Bank")
            or row.get("Partner Group Source")
            or row_value_by_patterns(row, "partnergroupwithbank", "partnergroupsource")
        )
        if not partner:
            continue

        raw_txn_type = row_value_first(
            row,
            "Payment Txn Type",
            "Txn Type",
            patterns=("paymenttxntype", "txntype"),
        )
        is_rtp = boolish(row.get("IsRTP") or row_value_by_patterns(row, "isrtp"))
        faster_ach = text(row.get("Is Faster Ach") or row_value_by_patterns(row, "isfasterach")) == "FasterACH"
        speed_flag = normalize_speed(is_rtp, faster_ach)
        payment_type = text(
            row_value_first(
                row,
                "Payment Payment Type",
                "Payment Type",
                patterns=("paymentpaymenttype", "paymenttype"),
            )
        )
        payer_ccy = text(row_value_first(row, "Payment Payer Amount Currency", "Payer Amount Currency", patterns=("payeramountcurrency",))) or "USD"
        payee_ccy = text(row_value_first(row, "Payment Payee Amount Currency", "Payee Amount Currency", patterns=("payeeamountcurrency",))) or "USD"
        payer_country = normalize_country_code(row_value_first(row, "Payment Payer Country", "Payer Country", patterns=("payercountry",)))
        payee_country = normalize_country_code(row_value_first(row, "Payment Payee Country", "Payee Country", patterns=("payeecountry",)))
        txn_type = derive_contract_txn_type(
            raw_txn_type,
            payer_ccy=payer_ccy,
            payee_ccy=payee_ccy,
            payer_country=payer_country,
            payee_country=payee_country,
            payment_type=payment_type,
        )
        if speed_flag == "RTP":
            processing_method = "RTP"
        elif txn_type in ("FX", "USD Abroad"):
            processing_method = "Wire"
        elif txn_type == "CAD Domestic":
            processing_method = "EFT"
        else:
            processing_method = text(
                row_value_first(
                    row,
                    "Payment Transaction Processing Method",
                    "Transaction Processing Method",
                    patterns=("transactionprocessingmethod", "processingmethod"),
                )
            ) or "ACH"

        wallet_flag = "wallet" in payment_type.lower()
        customer_revenue_value = (
            row.get("Net Revenue")
            if row.get("Net Revenue") not in (None, "")
            else row.get("Fixed Fee")
            if row.get("Fixed Fee") not in (None, "")
            else row_value_by_patterns(row, "netrevenue", "fixedfee")
        )
        est_revenue = round(extract_est_revenue(row), 2)
        revenue_basis = "net" if row.get("Net Revenue") not in (None, "") else "gross"
        key = (
            partner,
            month,
            txn_type,
            speed_flag,
            "Wallet" if wallet_flag else "",
            "Wallet" if wallet_flag else "",
            payer_ccy,
            payee_ccy,
            payer_country,
            payee_country,
            processing_method,
            revenue_basis,
        )
        grouped[key]["txnCount"] += 1
        grouped[key]["totalVolume"] += money(
            row_value_first(
                row,
                "Payment Total USD Amount Number",
                "USD Amount Number",
                "Total USD Amount Number",
                patterns=("totalusdamountnumber", "usdamountnumber"),
            )
        )
        grouped[key]["customerRevenue"] += money(customer_revenue_value)
        grouped[key]["estRevenue"] += est_revenue
        detail_rows.append(
            {
                "detailCategory": "transaction",
                "detailSource": "revenue_share",
                "partner": partner,
                "period": month,
                "paymentId": text(
                    row_value_first(
                        row,
                        "Partner Revenue Share Fixed Payment ID",
                        "Payment Payment ID",
                        "Payment ID",
                        patterns=("paymentid",),
                    )
                ),
                "txnType": txn_type,
                "speedFlag": speed_flag,
                "processingMethod": processing_method,
                "payerFunding": "Wallet" if wallet_flag else "",
                "payeeFunding": "Wallet" if wallet_flag else "",
                "payerCcy": payer_ccy,
                "payeeCcy": payee_ccy,
                "payerCountry": payer_country,
                "payeeCountry": payee_country,
                "accountId": text(row.get("Account Id") or row_value_by_patterns(row, "accountid", "payeraccountid")),
                "paymentType": payment_type,
                "submissionDate": iso_value(row_value_first(row, "Payment Time Created Date", "Time Created Date", patterns=("timecreateddate",))),
                "creditCompleteDate": iso_value(credit_complete_value),
                "txnTypeRaw": text(raw_txn_type),
                "payerName": text(row.get("Payer Name") or row_value_by_patterns(row, "payername")),
                "payeeName": text(row.get("Payee Name") or row_value_by_patterns(row, "payeename")),
                "payerEmail": text(row.get("Payer Email") or row_value_by_patterns(row, "payercustomeraccountprimaryemail")),
                "payeeEmail": text(row.get("Payee Email") or row_value_by_patterns(row, "payeecustomeraccountprimaryemail")),
                "usdAmount": round(
                    money(
                        row_value_first(
                            row,
                            "Payment Total USD Amount Number",
                            "USD Amount Number",
                            "Total USD Amount Number",
                            patterns=("totalusdamountnumber", "usdamountnumber"),
                        )
                    ),
                    2,
                ),
                "customerRevenue": round(money(customer_revenue_value), 2),
                "estRevenue": est_revenue,
                "netRevenue": round(money(row.get("Net Revenue") or row_value_by_patterns(row, "netrevenue")), 2),
                "countPricing": round(money(row.get("Count Pricing") or row_value_by_patterns(row, "countpricing")), 2),
                "isRTP": is_rtp,
                "isFasterAch": faster_ach,
                "initiatorStatus": text(row.get("Initiator Status") or row_value_by_patterns(row, "initiatorstatus")),
                "typeDefn": text(row.get("Type Defn") or row_value_by_patterns(row, "typedefn")),
                "revenueBasis": revenue_basis,
            }
        )

    output: list[dict[str, Any]] = []
    for key, aggregate in sorted(grouped.items()):
        (
            partner,
            month,
            txn_type,
            speed_flag,
            payer_funding,
            payee_funding,
            payer_ccy,
            payee_ccy,
            payer_country,
            payee_country,
            processing_method,
            revenue_basis,
        ) = key
        txn_count = int(aggregate["txnCount"])
        total_volume = round(aggregate["totalVolume"], 2)
        customer_revenue = round(aggregate["customerRevenue"], 2)
        est_revenue = round(aggregate["estRevenue"], 2)
        output.append(
            {
                "period": month,
                "partner": partner,
                "txnType": txn_type,
                "speedFlag": speed_flag,
                "minAmt": txn_count,
                "maxAmt": txn_count,
                "payerFunding": payer_funding,
                "payeeFunding": payee_funding,
                "payerCcy": payer_ccy,
                "payeeCcy": payee_ccy,
                "payerCountry": payer_country,
                "payeeCountry": payee_country,
                "processingMethod": processing_method,
                "txnCount": txn_count,
                "totalVolume": total_volume,
                "customerRevenue": customer_revenue,
                "estRevenue": est_revenue,
                "avgTxnSize": round(total_volume / txn_count, 2) if txn_count else 0.0,
                "revenueBasis": revenue_basis,
            }
        )
    return output, detail_rows


def build_revenue_share_summary(rows: list[dict[str, Any]], period: str | None, *, allow_billing_month_fallback: bool = True) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        month = month_key(
            row_value_first(
                row,
                "Credit Complete Date",
                "Transaction Lookup Dates Credit Complete Timestamp Time",
                "Transaction Lookup Dates Credit Complete Timestamp Date",
                "Credit Complete Timestamp Date",
                patterns=("creditcompletedate", "creditcompletetimestamptime", "creditcompletetimestampdate"),
            )
        )
        if month:
            if not matches_period(month, period):
                continue
            partner = normalize_revenue_partner(row.get("Partner Group Source") or row.get("Partner Group With Bank"))
            if not partner:
                continue
            output.append(
                {
                    "period": month,
                    "partner": partner,
                    "netRevenue": money(row.get("Net Revenue")),
                    "partnerRevenueShare": money(row.get("Partner Net Revenue Share")),
                    "revenueOwed": money(row.get("Revenue Owed")),
                    "monthlyMinimumRevenue": money(row.get("Monthly Minimum Revenue")),
                    "revenueSource": "summary",
                }
            )
            continue

        if not allow_billing_month_fallback:
            continue
        summary_month = month_key(row_value_by_patterns(row, "billingmonthmonth", "billingmonth", "billingmo"))
        if not summary_month or not matches_period(summary_month, period):
            continue
        partner = normalize_revenue_partner(row_value_by_patterns(row, "partnername", "partner"))
        if not partner:
            continue
        total_amount_raw = row_value_by_patterns(row, "totalamountfee", "totalamount", "totalamo")
        if total_amount_raw in (None, ""):
            continue
        total_amount = abs(money(total_amount_raw))
        billing_type = text(row_value_by_patterns(row, "billingtype", "billingty")) or "Billing Summary"
        computation = text(row_value_by_patterns(row, "computationmemo", "computation", "computati", "memo"))
        normalized_context = f"{billing_type} {computation}".lower()
        direction = "pay" if any(token in normalized_context for token in ("revsharepayout", "revshare payout", "partner net revenue share", "we pay", "veem owes", "payout")) else "charge"
        count = 0.0
        unit_amount = 0.0
        match = re.search(r"([-+]?[0-9][0-9,]*(?:\.[0-9]+)?)\s*\*\s*\$?([-+]?[0-9][0-9,]*(?:\.[0-9]+)?)", computation.replace("%", ""))
        if match:
            count = money(match.group(1))
            unit_amount = money(match.group(2))
        is_minimum_row = "minimum" in normalized_context
        output.append(
            {
                "period": summary_month,
                "partner": partner,
                "netRevenue": 0.0,
                "partnerRevenueShare": total_amount if direction == "pay" else 0.0,
                "revenueOwed": total_amount if direction == "charge" else 0.0,
                "monthlyMinimumRevenue": total_amount if direction == "charge" and is_minimum_row else 0.0,
                "revenueSource": "billing_summary",
                "summaryDirection": direction,
                "summaryBillingType": billing_type,
                "summaryLabel": billing_type.strip() or "Billing Summary",
                "summaryComputation": computation,
                "summaryCount": round(count, 2),
                "summaryUnitAmount": round(unit_amount, 6),
                "summaryLineAmount": round(total_amount, 2),
            }
        )
    return sorted(output, key=lambda row: (row["partner"], row["period"], str(row.get("summaryLabel") or row.get("revenueSource") or ""), str(row.get("summaryComputation") or "")))


def build_revenue_reversal_summary(rows: list[dict[str, Any]], period: str | None) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        month = month_key(
            row_value_first(
                row,
                "Refund Complete Date",
                "Refund Completed Date",
                "Transaction Lookup Dates Refund Complete Timestamp Time",
                "Transaction Lookup Dates Refund Complete Timestamp Date",
                patterns=("refundcompletedate", "refundcompleteddate", "refundcompletetimestamptime", "refundcompletetimestampdate"),
            )
        )
        if not month:
            continue
        if not matches_period(month, period):
            continue
        partner = normalize_revenue_partner(
            row.get("Partner Group Source")
            or row.get("Partner Group With Bank")
            or row_value_by_patterns(row, "partnergroupsource", "partnergroupwithbank")
        )
        if not partner:
            continue
        net_revenue = abs(money(row.get("Net Revenue") or row_value_by_patterns(row, "netrevenue")))
        partner_share = abs(
            money(row.get("Partner Net Revenue Share") or row_value_by_patterns(row, "partnernetrevenueshare"))
        )
        revenue_owed = abs(money(row.get("Revenue Owed") or row_value_by_patterns(row, "revenueowed")))
        if partner_share <= 0 and revenue_owed <= 0:
            rate = money(
                row.get("Partner Revenue Share Rate")
                or row_value_by_patterns(row, "inipartnerrevenuesharerate", "partnerrevenuesharerate")
            )
            if rate > 0 and net_revenue > 0:
                partner_share = round(net_revenue * rate, 2)
        output.append(
            {
                "period": month,
                "partner": partner,
                "netRevenue": -net_revenue,
                "partnerRevenueShare": -partner_share,
                "revenueOwed": -revenue_owed,
                "monthlyMinimumRevenue": 0.0,
                "revenueSource": "reversal",
            }
        )
    return sorted(output, key=lambda row: (row["partner"], row["period"]))


def collect_periods(*row_sets: list[dict[str, Any]]) -> list[str]:
    periods = {
        month_key(
            row.get("Credit Complete Month")
            or row.get("Refund Complete Date")
            or row.get("Refund Completed Date")
            or row.get("Debit Reversal Month")
            or row.get("Credit Complete Date")
            or row.get("Credit Complete Timestamp Month")
            or row.get("Credit Complete Timestamp Date")
            or row.get("Time Created Date")
        )
        for rows in row_sets
        for row in rows
    }
    return sorted(period for period in periods if period)


def month_end(period: str) -> date:
    year = int(period[:4])
    month = int(period[5:7])
    if month == 12:
        return date(year, month, 31)
    next_month = date(year, month + 1, 1)
    return next_month.fromordinal(next_month.toordinal() - 1)


def build_virtual_account_usage(
    register_rows: list[dict[str, Any]],
    account_activity: dict[str, list[date]],
    settlement_days: dict[tuple[str, str], set[date]],
    periods: list[str],
) -> list[dict[str, Any]]:
    target_partners = {
        normalize_partner_name(
            row.get("Partner Name")
            or row.get("Partner Group Source")
            or row.get("Partner Group With Bank")
            or row_value_by_patterns(row, "partnername", "partnergroupsource", "partnergroupwithbank")
        )
        for row in register_rows
        if normalize_partner_name(
            row.get("Partner Name")
            or row.get("Partner Group Source")
            or row.get("Partner Group With Bank")
            or row_value_by_patterns(row, "partnername", "partnergroupsource", "partnergroupwithbank")
        )
    }
    target_partners.update(partner for partner, _period in settlement_days.keys() if partner)
    if not target_partners:
        return []
    per_month = {
        (partner, period): {
            "period": period,
            "partner": partner,
            "newAccountsOpened": 0,
            "totalActiveAccounts": 0,
            "totalBusinessAccounts": 0,
            "totalIndividualAccounts": 0,
            "dormantAccounts": 0,
            "closedAccounts": 0,
            "newBusinessSetups": 0,
            "settlementCount": len(settlement_days.get((partner, period), set())),
        }
        for period in periods
        for partner in target_partners
    }

    accounts_by_partner: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in register_rows:
        partner = normalize_partner_name(
            row.get("Partner Name")
            or row.get("Partner Group Source")
            or row.get("Partner Group With Bank")
            or row_value_by_patterns(row, "partnername", "partnergroupsource", "partnergroupwithbank")
        )
        if partner not in target_partners:
            continue
        join_date = parse_dateish_from_row(
            row,
            "Join Date Time",
            patterns=("joindatetime",),
        )
        if not join_date:
            continue
        account_id = text(row.get("Account Id") or row.get("ACCOUNT_ID") or row_value_by_patterns(row, "accountid", "customeraccountid"))
        activities = sorted(set(account_activity.get(account_id, [])))
        accounts_by_partner[partner].append(
            {
                "accountId": account_id,
                "joinDate": join_date,
                "typeDefn": text(row.get("Type Defn") or row_value_by_patterns(row, "typedefn")),
                "status": text(row.get("Status") or row.get("STATUS") or row_value_by_patterns(row, "status")).lower(),
                "activities": activities,
            }
        )

    for partner, accounts in accounts_by_partner.items():
        for period in periods:
            entry = per_month[(partner, period)]
            period_start = date(int(period[:4]), int(period[5:7]), 1)
            period_end = month_end(period)
            dormant_cutoff = date.fromordinal(period_end.toordinal() - 90)
            for account in accounts:
                if account["joinDate"] > period_end:
                    continue
                entry["totalActiveAccounts"] += 1
                type_defn = account["typeDefn"].lower()
                if type_defn == "business":
                    entry["totalBusinessAccounts"] += 1
                elif type_defn == "individual":
                    entry["totalIndividualAccounts"] += 1
                if period_start <= account["joinDate"] <= period_end:
                    entry["newAccountsOpened"] += 1
                    if type_defn == "business":
                        entry["newBusinessSetups"] += 1
                if account["joinDate"] > dormant_cutoff:
                    continue
                activities = account["activities"]
                idx = bisect.bisect_right(activities, period_end)
                last_activity = activities[idx - 1] if idx else None
                if last_activity is None or last_activity < dormant_cutoff:
                    entry["dormantAccounts"] += 1
                if account["status"] in {"closed", "inactive", "deactivated"}:
                    entry["closedAccounts"] += 1

    return [per_month[key] for key in sorted(per_month)]


def build_stampli_fx_partner_payouts_from_detail(detail_rows: list[dict[str, Any]], periods: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_period: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        if text(row.get("detailCategory")) != "transaction":
            continue
        if text(row.get("detailSource")) != "offline_billing":
            continue
        if text(row.get("partner")) != STAMPLI_FX_PARTNER:
            continue
        payee_currency = text(row.get("payeeAmountCurrency") or row.get("payeeCcy")).upper()
        payer_currency = text(row.get("payerCcy")).upper()
        if not payee_currency or payee_currency == "USD":
            continue
        if payer_currency not in ("", "USD"):
            continue
        period = text(row.get("period"))
        if period:
            by_period[period].append(row)

    output: list[dict[str, Any]] = []
    for period in periods:
        rows = by_period.get(period, [])
        if not rows:
            continue

        avg_ratio_by_currency: dict[str, float] = {}
        ratio_totals: dict[str, dict[str, float]] = defaultdict(lambda: {"mid_usd": 0.0, "payee_amount": 0.0})
        for row in rows:
            payee_currency = text(row.get("payeeAmountCurrency")).upper()
            payee_amount = money(row.get("payeeAmount"))
            mid_market_usd = money(row.get("paymentUsdEquivalentAmount"))
            if payee_currency and payee_amount > 0 and mid_market_usd > 0:
                ratio_totals[payee_currency]["mid_usd"] += mid_market_usd
                ratio_totals[payee_currency]["payee_amount"] += payee_amount
        for payee_currency, totals in ratio_totals.items():
            if totals["payee_amount"] > 0:
                avg_ratio_by_currency[payee_currency] = totals["mid_usd"] / totals["payee_amount"]

        prepared_rows: list[dict[str, Any]] = []
        used_period_average = 0
        missing_mid_market = 0
        missing_customer_charge = 0
        for row in rows:
            payee_currency = text(row.get("payeeAmountCurrency")).upper()
            payee_amount = money(row.get("payeeAmount"))
            mid_market_usd = money(row.get("paymentUsdEquivalentAmount"))
            if mid_market_usd <= 0 and payee_amount > 0 and payee_currency in avg_ratio_by_currency:
                mid_market_usd = round(payee_amount * avg_ratio_by_currency[payee_currency], 2)
                used_period_average += 1
            if mid_market_usd <= 0:
                missing_mid_market += 1
                continue
            usd_debited = money(row.get("usdAmountDebited"))
            if usd_debited <= 0:
                missing_customer_charge += 1
                continue
            prepared_rows.append({
                "paymentId": text(row.get("paymentId")),
                "midMarketUsd": mid_market_usd,
                "usdDebited": usd_debited,
                "payeeCurrency": payee_currency,
            })

        if not prepared_rows:
            note_parts = []
            if missing_mid_market:
                note_parts.append(f"{missing_mid_market} txn(s) missing daily mid-market equivalent")
            if missing_customer_charge:
                note_parts.append(f"{missing_customer_charge} txn(s) missing customer charge amount")
            output.append({
                "partner": STAMPLI_FX_PARTNER,
                "period": period,
                "txnCount": 0,
                "partnerPayout": 0.0,
                "shareAmount": 0.0,
                "reversalAmount": 0.0,
                "shareTxnCount": 0,
                "reversalTxnCount": 0,
                "totalUsdDebited": 0.0,
                "totalMidMarketUsd": 0.0,
                "totalGrossMarkup": 0.0,
                "variableSpreadRate": 0.0,
                "totalVariableSpreadCost": 0.0,
                "totalPerTxnFee": 0.0,
                "totalCompanyMarkup": 0.0,
                "usedPeriodAverageCount": used_period_average,
                "missingMidMarketCount": missing_mid_market,
                "missingCustomerChargeCount": missing_customer_charge,
                "negativePayoutTxnCount": 0,
                "note": "; ".join(note_parts),
            })
            continue

        total_mid_market_usd = round(sum(row["midMarketUsd"] for row in prepared_rows), 2)
        variable_spread_rate = choose_fx_variable_spread_rate(total_mid_market_usd)
        totals = {
            "partnerPayout": 0.0,
            "usdDebited": 0.0,
            "grossMarkup": 0.0,
            "variableSpreadCost": 0.0,
            "perTxnFee": 0.0,
            "companyMarkup": 0.0,
        }
        negative_payout_count = 0
        fee_sources: Counter[str] = Counter()

        for row in prepared_rows:
            fixed_fee, fee_source = choose_fx_payment_fee(row["payeeCurrency"])
            fee_sources[fee_source] += 1
            gross_markup = round(row["usdDebited"] - row["midMarketUsd"], 2)
            variable_spread_cost = round(variable_spread_rate * row["midMarketUsd"], 2)
            company_markup = round(STAMPLI_COMPANY_MARKUP_BPS * row["midMarketUsd"], 2)
            partner_payout = round(gross_markup - variable_spread_cost - fixed_fee - company_markup, 2)
            if partner_payout < 0:
                negative_payout_count += 1
                partner_payout = 0.0
            totals["partnerPayout"] += partner_payout
            totals["usdDebited"] += row["usdDebited"]
            totals["grossMarkup"] += gross_markup
            totals["variableSpreadCost"] += variable_spread_cost
            totals["perTxnFee"] += fixed_fee
            totals["companyMarkup"] += company_markup

        note_parts = []
        if used_period_average:
            note_parts.append(f"Used period-average mid-market fallback for {used_period_average} txn(s)")
        if missing_mid_market:
            note_parts.append(f"Skipped {missing_mid_market} txn(s) with no daily or period-average mid-market equivalent")
        if missing_customer_charge:
            note_parts.append(f"Skipped {missing_customer_charge} txn(s) with no customer charge amount")
        if negative_payout_count:
            note_parts.append(f"Capped negative partner markup to $0 on {negative_payout_count} txn(s)")
        if fee_sources:
            fee_summary = ", ".join(f"{label}: {count}" for label, count in sorted(fee_sources.items()))
            note_parts.append(f"Per-txn fee assumptions used: {fee_summary}")

            output.append({
                "partner": STAMPLI_FX_PARTNER,
                "period": period,
                "txnCount": len(prepared_rows),
                "partnerPayout": round(totals["partnerPayout"], 2),
                "shareAmount": round(totals["partnerPayout"], 2),
                "reversalAmount": 0.0,
                "shareTxnCount": len(prepared_rows),
                "reversalTxnCount": 0,
                "totalUsdDebited": round(totals["usdDebited"], 2),
                "totalMidMarketUsd": total_mid_market_usd,
                "totalGrossMarkup": round(totals["grossMarkup"], 2),
            "variableSpreadRate": variable_spread_rate,
            "totalVariableSpreadCost": round(totals["variableSpreadCost"], 2),
            "totalPerTxnFee": round(totals["perTxnFee"], 2),
            "totalCompanyMarkup": round(totals["companyMarkup"], 2),
            "usedPeriodAverageCount": used_period_average,
            "missingMidMarketCount": missing_mid_market,
            "missingCustomerChargeCount": missing_customer_charge,
            "negativePayoutTxnCount": negative_payout_count,
            "note": "; ".join(note_parts),
        })

    return output, []


def build_stampli_fx_partner_payouts_from_feed(
    share_rows: list[dict[str, Any]],
    reversal_rows: list[dict[str, Any]],
    period: str | None,
    credit_complete_lookup: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    credit_complete_lookup = credit_complete_lookup or {}
    ratio_totals: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"mid_usd": 0.0, "payee_amount": 0.0})
    for row in share_rows:
        payment_id = text(
            row_value_first(
                row,
                "** Payment For Sales DV ** Payment Id",
                "Payment Payment ID",
                "Payment ID",
                patterns=("paymentid",),
            )
        )
        if not payment_id:
            continue
        month = month_key(
            row_value_first(
                row,
                "Credit Complete Date",
                "Transaction Lookup Dates Credit Complete Timestamp Date",
                "Transaction Lookup Dates Credit Complete Timestamp Time",
                patterns=("creditcompletedate", "creditcompletetimestampdate", "creditcompletetimestamptime"),
            )
            or credit_complete_lookup.get(payment_id, "")
        )
        if not month or not matches_period(month, period):
            continue
        payee_currency = text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))).upper()
        payee_amount = money(row_value_first(row, "** Payment For Sales DV ** Payee Amount Number", "Foreign Currency Amount (Payee Amount Number)", patterns=("payeeamountnumber",)))
        mid_market_usd = money(
            row_value_first(
                row,
                "Payment USD Equivalent Amount",
                "** Payment For Sales DV ** Total USD Amount Number",
                "** Payment For Sales DV ** USD Amount Number",
                patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"),
            )
        )
        if payee_currency and payee_amount > 0 and mid_market_usd > 0:
            ratio_totals[(month, payee_currency)]["mid_usd"] += mid_market_usd
            ratio_totals[(month, payee_currency)]["payee_amount"] += payee_amount
    for row in reversal_rows:
        payment_id = text(
            row_value_first(
                row,
                "** Payment For Sales DV ** Payment Id",
                "Payment Payment ID",
                "Payment ID",
                patterns=("paymentid",),
            )
        )
        if not payment_id:
            continue
        month = month_key(
            row_value_first(
                row,
                "Refund Complete Date",
                "Refund Completed Date",
                "Transaction Lookup Dates Refund Complete Timestamp Date",
                "Transaction Lookup Dates Refund Complete Timestamp Time",
                patterns=("refundcompletedate", "refundcompleteddate", "refundcompletetimestampdate", "refundcompletetimestamptime"),
            )
        )
        if not month or not matches_period(month, period):
            continue
        payee_currency = text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))).upper()
        payee_amount = money(row_value_first(row, "** Payment For Sales DV ** Payee Amount Number", "Foreign Currency Amount (Payee Amount Number)", patterns=("payeeamountnumber",)))
        mid_market_usd = money(
            row_value_first(
                row,
                "Payment USD Equivalent Amount",
                "** Payment For Sales DV ** Total USD Amount Number",
                "** Payment For Sales DV ** USD Amount Number",
                patterns=("paymentusdequivalentamount", "totalusdamountnumber", "usdamountnumber"),
            )
        )
        if payee_currency and payee_amount > 0 and mid_market_usd > 0:
            ratio_totals[(month, payee_currency)]["mid_usd"] += mid_market_usd
            ratio_totals[(month, payee_currency)]["payee_amount"] += payee_amount
    avg_ratio_by_month_currency = {
        key: totals["mid_usd"] / totals["payee_amount"]
        for key, totals in ratio_totals.items()
        if totals["payee_amount"] > 0
    }

    grouped: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "partner": STAMPLI_FX_PARTNER,
        "period": "",
        "txnCount": 0,
        "shareTxnCount": 0,
        "reversalTxnCount": 0,
        "partnerPayout": 0.0,
        "shareAmount": 0.0,
        "reversalAmount": 0.0,
        "totalUsdDebited": 0.0,
        "totalMidMarketUsd": 0.0,
        "shareTotalUsdDebited": 0.0,
        "shareTotalMidMarketUsd": 0.0,
        "reversalTotalUsdDebited": 0.0,
        "reversalTotalMidMarketUsd": 0.0,
        "totalGrossMarkup": 0.0,
        "totalVariableSpreadCost": 0.0,
        "totalPerTxnFee": 0.0,
        "totalCompanyMarkup": 0.0,
        "variableSpreadRate": 0.0,
        "usedPeriodAverageCount": 0,
        "missingMidMarketCount": 0,
        "missingCustomerChargeCount": 0,
        "negativePayoutTxnCount": 0,
        "validationCheckedCount": 0,
        "validationAmountMismatchCount": 0,
        "validationPctCheckedCount": 0,
        "validationPctMismatchCount": 0,
        "validationAmountDelta": 0.0,
        "skippedBlankPaymentIdCount": 0,
        "note": "Direct Stampli FX revenue-share feed",
    })
    detail_rows: list[dict[str, Any]] = []
    seen_share_ids: set[str] = set()
    seen_reversal_ids: set[str] = set()

    for row in share_rows:
        payment_id = text(
            row_value_first(
                row,
                "** Payment For Sales DV ** Payment Id",
                "Payment Payment ID",
                "Payment ID",
                patterns=("paymentid",),
            )
        )
        credit_complete_value = row_value_first(
            row,
            "Credit Complete Date",
            "Transaction Lookup Dates Credit Complete Timestamp Date",
            "Transaction Lookup Dates Credit Complete Timestamp Time",
            patterns=("creditcompletedate", "creditcompletetimestampdate", "creditcompletetimestamptime"),
        )
        month = month_key(credit_complete_value or credit_complete_lookup.get(payment_id, ""))
        if not month or not matches_period(month, period):
            continue
        if not payment_id:
            grouped[month]["period"] = month
            grouped[month]["skippedBlankPaymentIdCount"] += 1
            continue
        if payment_id in seen_share_ids:
            continue
        seen_share_ids.add(payment_id)
        entry = grouped[month]
        entry["period"] = month
        calc = calculate_stampli_markup_from_feed_row(row, month, avg_ratio_by_month_currency)
        entry["shareTxnCount"] += 1
        entry["txnCount"] += 1
        entry["shareAmount"] += calc["usedAmount"]
        entry["partnerPayout"] += calc["usedAmount"]
        entry["totalUsdDebited"] += calc["usdDebited"]
        entry["totalMidMarketUsd"] += calc["midMarketUsd"]
        entry["shareTotalUsdDebited"] += calc["usdDebited"]
        entry["shareTotalMidMarketUsd"] += calc["midMarketUsd"]
        entry["totalGrossMarkup"] += calc["grossMarkup"]
        if calc["usedPeriodAverage"]:
            entry["usedPeriodAverageCount"] += 1
        if calc["midMarketUsd"] <= 0:
            entry["missingMidMarketCount"] += 1
        if calc["usdDebited"] <= 0:
            entry["missingCustomerChargeCount"] += 1
        if calc["providedAmount"] > 0 and calc["calculatedAmount"] > 0:
            entry["validationCheckedCount"] += 1
            entry["validationAmountDelta"] += round(calc["calculatedAmount"] - calc["providedAmount"], 2)
            if not calc["amountMatchesSheet"]:
                entry["validationAmountMismatchCount"] += 1
        if calc["providedMarkupPct"] and calc["derivedMarkupPct"]:
            entry["validationPctCheckedCount"] += 1
            if not calc["pctMatchesComponents"]:
                entry["validationPctMismatchCount"] += 1
        detail_rows.append({
            "detailCategory": "transaction",
            "detailSource": "stampli_fx_revenue_share",
            "stampliFxDirection": "share",
            "partner": STAMPLI_FX_PARTNER,
            "period": month,
            "paymentId": payment_id,
            "txnType": "FX",
            "speedFlag": "Standard",
            "processingMethod": "Wire",
            "payerFunding": "Bank",
            "payeeFunding": "Bank",
            "payerCcy": "USD",
            "payeeCcy": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))).upper(),
            "payerCountry": "",
            "payeeCountry": normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payee Country", "Payee Country", patterns=("payeecountry",))),
            "accountId": text(row_value_first(row, "** Payment For Sales DV ** Payer Account ID", "Account ID", patterns=("payeraccountid", "accountid"))),
            "paymentType": "FX",
            "submissionDate": iso_value(row_value_first(row, "Date of Payment Submission", "** Payment For Sales DV ** Time Created Date", patterns=("dateofpaymentsubmission", "timecreateddate"))),
            "creditCompleteDate": iso_value(credit_complete_value),
            "payerEmail": text(row.get("Payer Email")),
            "payerBusinessName": text(row.get("Payer Business Name")),
            "payeeEmail": text(row.get("Payee Email")),
            "payeeBusinessName": text(row.get("Payee Business Name")),
            "payeeAmountCurrency": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))).upper(),
            "payeeAmount": round(money(row_value_first(row, "** Payment For Sales DV ** Payee Amount Number", "Foreign Currency Amount (Payee Amount Number)", patterns=("payeeamountnumber",))), 2),
            "usdAmountDebited": calc["usdDebited"],
            "paymentUsdEquivalentAmount": calc["midMarketUsd"],
            "openExchangeRateUsed": text(row.get("Open Exchange Rate Used for this Transaction") or row.get("Open Exchange Rate used for this Transaction")),
            "customerMarkupPct": calc["customerMarkupPct"],
            "stampliBuyRatePct": calc["stampliBuyRatePct"],
            "stampliMarkupPct": calc["providedMarkupPct"],
            "stampliMarkupPctCalculated": calc["derivedMarkupPct"],
            "stampliMarkupPctUsed": calc["effectiveMarkupPct"],
            "stampliMarkupPctMatchesSheet": calc["pctMatchesComponents"],
            "stampliMarkupAmount": calc["usedAmount"],
            "stampliMarkupAmountProvided": calc["providedAmount"],
            "stampliMarkupAmountCalculated": calc["calculatedAmount"],
            "stampliMarkupAmountMatchesSheet": calc["amountMatchesSheet"],
            "midMarketFallbackUsed": calc["usedPeriodAverage"],
        })

    for row in reversal_rows:
        refund_complete_value = row_value_first(
            row,
            "Refund Complete Date",
            "Refund Completed Date",
            "Transaction Lookup Dates Refund Complete Timestamp Date",
            "Transaction Lookup Dates Refund Complete Timestamp Time",
            patterns=("refundcompletedate", "refundcompleteddate", "refundcompletetimestampdate", "refundcompletetimestamptime"),
        )
        month = month_key(refund_complete_value)
        if not month or not matches_period(month, period):
            continue
        payment_id = text(
            row_value_first(
                row,
                "** Payment For Sales DV ** Payment Id",
                "Payment Payment ID",
                "Payment ID",
                patterns=("paymentid",),
            )
        )
        if not payment_id:
            grouped[month]["period"] = month
            grouped[month]["skippedBlankPaymentIdCount"] += 1
            continue
        if payment_id in seen_reversal_ids:
            continue
        seen_reversal_ids.add(payment_id)
        entry = grouped[month]
        entry["period"] = month
        calc = calculate_stampli_markup_from_feed_row(row, month, avg_ratio_by_month_currency)
        entry["reversalTxnCount"] += 1
        entry["txnCount"] += 1
        entry["reversalAmount"] += calc["usedAmount"]
        entry["partnerPayout"] -= calc["usedAmount"]
        entry["totalUsdDebited"] += calc["usdDebited"]
        entry["totalMidMarketUsd"] += calc["midMarketUsd"]
        entry["reversalTotalUsdDebited"] += calc["usdDebited"]
        entry["reversalTotalMidMarketUsd"] += calc["midMarketUsd"]
        entry["totalGrossMarkup"] += calc["grossMarkup"]
        if calc["usedPeriodAverage"]:
            entry["usedPeriodAverageCount"] += 1
        if calc["midMarketUsd"] <= 0:
            entry["missingMidMarketCount"] += 1
        if calc["usdDebited"] <= 0:
            entry["missingCustomerChargeCount"] += 1
        if calc["providedAmount"] > 0 and calc["calculatedAmount"] > 0:
            entry["validationCheckedCount"] += 1
            entry["validationAmountDelta"] += round(calc["calculatedAmount"] - calc["providedAmount"], 2)
            if not calc["amountMatchesSheet"]:
                entry["validationAmountMismatchCount"] += 1
        if calc["providedMarkupPct"] and calc["derivedMarkupPct"]:
            entry["validationPctCheckedCount"] += 1
            if not calc["pctMatchesComponents"]:
                entry["validationPctMismatchCount"] += 1
        detail_rows.append({
            "detailCategory": "transaction",
            "detailSource": "stampli_fx_reversal",
            "stampliFxDirection": "reversal",
            "partner": STAMPLI_FX_PARTNER,
            "period": month,
            "paymentId": payment_id,
            "txnType": "FX",
            "speedFlag": "Standard",
            "processingMethod": "Wire",
            "payerFunding": "Bank",
            "payeeFunding": "Bank",
            "payerCcy": "USD",
            "payeeCcy": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))).upper(),
            "payerCountry": "",
            "payeeCountry": normalize_country_code(row_value_first(row, "** Payment For Sales DV ** Payee Country", "Payee Country", patterns=("payeecountry",))),
            "accountId": text(row_value_first(row, "** Payment For Sales DV ** Payer Account ID", "Account ID", patterns=("payeraccountid", "accountid"))),
            "paymentType": "FX Reversal",
            "submissionDate": iso_value(row_value_first(row, "Payment Submission Date", "** Payment For Sales DV ** Time Created Date", patterns=("paymentsubmissiondate", "timecreateddate"))),
            "creditCompleteDate": iso_value(row_value_first(row, "Credit Complete Date", "Transaction Lookup Dates Credit Complete Timestamp Date", "Transaction Lookup Dates Credit Complete Timestamp Time", patterns=("creditcompletedate", "creditcompletetimestampdate", "creditcompletetimestamptime"))),
            "reversalDate": iso_value(refund_complete_value),
            "payerEmail": text(row.get("Payer Email")),
            "payerBusinessName": text(row.get("Payer Business Name")),
            "payeeEmail": text(row.get("Payee Email")),
            "payeeBusinessName": text(row.get("Payee Business Name")),
            "payeeAmountCurrency": text(row_value_first(row, "** Payment For Sales DV ** Payee Amount Currency", "Currency (Payee Amount Currency)", patterns=("payeeamountcurrency",))).upper(),
            "payeeAmount": round(money(row_value_first(row, "** Payment For Sales DV ** Payee Amount Number", "Foreign Currency Amount (Payee Amount Number)", patterns=("payeeamountnumber",))), 2),
            "usdAmountDebited": calc["usdDebited"],
            "paymentUsdEquivalentAmount": calc["midMarketUsd"],
            "openExchangeRateUsed": text(row.get("Open Exchange Rate used for this Transaction") or row.get("Open Exchange Rate Used for this Transaction")),
            "customerMarkupPct": calc["customerMarkupPct"],
            "stampliBuyRatePct": calc["stampliBuyRatePct"],
            "stampliMarkupPct": calc["providedMarkupPct"],
            "stampliMarkupPctCalculated": calc["derivedMarkupPct"],
            "stampliMarkupPctUsed": calc["effectiveMarkupPct"],
            "stampliMarkupPctMatchesSheet": calc["pctMatchesComponents"],
            "stampliMarkupAmount": -calc["usedAmount"],
            "stampliMarkupAmountProvided": -calc["providedAmount"],
            "stampliMarkupAmountCalculated": -calc["calculatedAmount"],
            "stampliMarkupAmountMatchesSheet": calc["amountMatchesSheet"],
            "midMarketFallbackUsed": calc["usedPeriodAverage"],
        })

    output = []
    for month in sorted(grouped):
        entry = grouped[month]
        note_parts = ["Direct Stampli FX revenue-share feed"]
        if entry["shareTxnCount"] or entry["reversalTxnCount"]:
            note_parts.append(f"{entry['shareTxnCount']} payout txn(s), {entry['reversalTxnCount']} reversal txn(s)")
        if entry["validationCheckedCount"]:
            matched = entry["validationCheckedCount"] - entry["validationAmountMismatchCount"]
            note_parts.append(f"Calculated markup matched sheet on {matched}/{entry['validationCheckedCount']} amount checks")
        if entry["validationPctCheckedCount"]:
            matched = entry["validationPctCheckedCount"] - entry["validationPctMismatchCount"]
            note_parts.append(f"Markup-rate formula matched on {matched}/{entry['validationPctCheckedCount']} rows")
        if entry["validationAmountMismatchCount"]:
            note_parts.append(f"Used calculated amount on {entry['validationAmountMismatchCount']} row(s) where sheet amount differed by {entry['validationAmountDelta']:.2f} total")
        if entry["usedPeriodAverageCount"]:
            note_parts.append(f"Used period-average mid-market fallback for {entry['usedPeriodAverageCount']} txn(s)")
        if entry["missingMidMarketCount"]:
            note_parts.append(f"{entry['missingMidMarketCount']} txn(s) missing mid-market equivalent")
        if entry["skippedBlankPaymentIdCount"]:
            note_parts.append(f"Skipped {entry['skippedBlankPaymentIdCount']} blank footer row(s)")
        entry["partnerPayout"] = round(entry["partnerPayout"], 2)
        entry["shareAmount"] = round(entry["shareAmount"], 2)
        entry["reversalAmount"] = round(entry["reversalAmount"], 2)
        entry["totalUsdDebited"] = round(entry["totalUsdDebited"], 2)
        entry["totalMidMarketUsd"] = round(entry["totalMidMarketUsd"], 2)
        entry["shareTotalUsdDebited"] = round(entry["shareTotalUsdDebited"], 2)
        entry["shareTotalMidMarketUsd"] = round(entry["shareTotalMidMarketUsd"], 2)
        entry["reversalTotalUsdDebited"] = round(entry["reversalTotalUsdDebited"], 2)
        entry["reversalTotalMidMarketUsd"] = round(entry["reversalTotalMidMarketUsd"], 2)
        entry["totalGrossMarkup"] = round(entry["totalGrossMarkup"], 2)
        entry["validationAmountDelta"] = round(entry["validationAmountDelta"], 2)
        entry["note"] = "; ".join(note_parts)
        output.append(entry)

    return output, detail_rows


def build_stampli_fx_partner_payouts(
    detail_rows: list[dict[str, Any]],
    periods: list[str],
    share_rows: list[dict[str, Any]] | None = None,
    reversal_rows: list[dict[str, Any]] | None = None,
    period: str | None = None,
    credit_complete_lookup: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if share_rows or reversal_rows:
        return build_stampli_fx_partner_payouts_from_feed(share_rows or [], reversal_rows or [], period, credit_complete_lookup=credit_complete_lookup)
    return build_stampli_fx_partner_payouts_from_detail(detail_rows, periods)


def write_js_module(output_path: Path, payload: dict[str, Any]) -> None:
    output_path.write_text(
        "// Generated by tools/generate_looker_import.py\n"
        f"export const importedLookerData = {json.dumps(payload, indent=2, sort_keys=True)};\n"
    )


def write_json(output_path: Path, payload: dict[str, Any]) -> None:
    output_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")


def write_detail_files(output_dir: Path, detail_rows: list[dict[str, Any]]) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    by_partner_period: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        partner = text(row.get("partner"))
        period = text(row.get("period"))
        if not partner or not period:
            continue
        by_partner_period[(partner, period)].append(row)

    manifest: dict[str, str] = {}
    for (partner, period), rows in sorted(by_partner_period.items()):
        filename = f"{slugify(partner)}-{period}-details.json"
        output_path = output_dir / filename
        rows_sorted = sorted(
            rows,
            key=lambda row: (
                text(row.get("detailCategory")),
                text(row.get("submissionDate")),
                text(row.get("creditCompleteDate")),
                text(row.get("reversalDate")),
                text(row.get("paymentId")),
                text(row.get("accountId")),
            ),
        )
        output_path.write_text(json.dumps(rows_sorted, indent=2, sort_keys=True) + "\n")
        manifest[f"{partner}|{period}"] = f"./{output_dir.name}/{filename}"
    return manifest


def write_report(output_path: Path, payload: dict[str, Any], offline_meta: dict[str, Any], selected_period: str | None) -> None:
    partner_counts = Counter(row["partner"] for row in payload["ltxn"])
    summary_partner_counts = Counter(row["partner"] for row in payload["lrs"])
    stampli_fx_counts = Counter(row["partner"] for row in payload.get("lfxp", []))
    periods = sorted({row["period"] for row in payload["ltxn"] + payload["lrev"] + payload["lrs"] + payload.get("lfxp", []) + payload["lva"] if row.get("period")})
    lines = [
        "# Looker Import Summary",
        "",
        f"- Generated at: `{payload['generatedAt']}`",
        f"- Source folder: `{payload['sourceFolder']}`",
        f"- Imported periods: `{', '.join(periods)}`",
        f"- Period mode: `{'single month' if selected_period else 'all available periods'}`",
        f"- Transaction rows loaded into workbook: `{len(payload['ltxn'])}`",
        f"- Reversal rows loaded into workbook: `{len(payload['lrev'])}`",
        f"- Revenue-share summary rows loaded into workbook: `{len(payload['lrs'])}`",
        f"- FX partner-payout summary rows loaded into workbook: `{len(payload.get('lfxp', []))}`",
        "",
        "## Transaction Partners",
        "",
    ]
    for partner, count in sorted(partner_counts.items()):
        lines.append(f"- `{partner}`: {count} grouped rows")

    lines.extend(["", "## Revenue Share Summary Partners", ""])
    if summary_partner_counts:
        for partner, count in sorted(summary_partner_counts.items()):
            lines.append(f"- `{partner}`: {count} summary row(s)")
    else:
        lines.append("- No separate monthly revenue-share summary file was found; invoices will rely on rev-share detail plus workbook contract config.")

    lines.extend(["", "## FX Partner Payout Partners", ""])
    if stampli_fx_counts:
        for partner, count in sorted(stampli_fx_counts.items()):
            lines.append(f"- `{partner}`: {count} payout row(s)")
    else:
        lines.append("- No partner FX markup payout rows were derived from the supplied exports.")

    lines.extend(["", "## Missing Manual Inputs", ""])
    for gap in payload["gaps"]:
        lines.append(f"- {gap}")

    lines.extend(["", "## Unmatched Fixed-Billing Payments", "", f"- Unmatched payment IDs: `{offline_meta['unmatchedPaymentIds']}`"])
    for example, count in offline_meta["unmatchedExamples"]:
        lines.append(f"- `{example}`: {count} payment id(s)")

    output_path.write_text("\n".join(lines) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a workbook-ready Looker import bundle.")
    parser.add_argument("source_dir", type=Path, help="Folder containing the downloaded Looker exports")
    parser.add_argument("period", nargs="?", help="Optional invoice month in YYYY-MM format. Omit to load all periods present.")
    parser.add_argument("--stampli-source-dir", type=Path, default=None, help="Optional folder containing dedicated Stampli billing exports.")
    parser.add_argument("--output-js", type=Path, default=Path("/Users/danielsinukoff/Documents/billing-workbook/looker-import.js"))
    parser.add_argument("--output-json", type=Path, default=Path("/Users/danielsinukoff/Documents/billing-workbook/reports/looker_import/looker_import_bundle.json"))
    parser.add_argument("--output-report", type=Path, default=Path("/Users/danielsinukoff/Documents/billing-workbook/reports/looker_import/looker_import_summary.md"))
    parser.add_argument("--output-detail-dir", type=Path, default=Path("/Users/danielsinukoff/Documents/billing-workbook/looker-detail-files"))
    args = parser.parse_args()

    paths = load_paths(args.source_dir, args.stampli_source_dir)
    revenue_row_sets = [read_table(path) for path in paths.revenue_txns]
    revenue_rows = dedupe_revenue_rows(revenue_row_sets)
    revenue_summary_rows = read_table(paths.revenue_summary) if paths.revenue_summary else []
    stampli_fx_share_rows = read_table(paths.stampli_fx_share) if paths.stampli_fx_share else []
    stampli_fx_reversal_rows = read_table(paths.stampli_fx_reversals) if paths.stampli_fx_reversals else []
    stampli_credit_complete_lookup = build_stampli_credit_complete_lookup(paths.stampli_credit_complete_all)

    fixed_rows, offline_meta, account_activity, settlement_days, offline_periods, offline_detail_rows = build_offline_transactions(paths.offline_txns, args.period)
    reversal_rows, reversal_meta, reversal_periods, reversal_detail_rows = build_offline_reversals(paths.offline_reversals, args.period)
    revenue_detail_rows, revenue_detail_export_rows = build_revenue_detail_transactions(revenue_rows, args.period)
    revenue_summary = build_revenue_share_summary(revenue_summary_rows, args.period)
    stampli_direct_rows, stampli_direct_periods, stampli_direct_detail_rows, stampli_direct_meta = build_stampli_direct_billing(
        paths.stampli_credit_complete_all,
        paths.stampli_domestic_revenue,
        paths.stampli_usd_abroad_revenue,
        args.period,
    )
    if stampli_direct_rows:
        covered_periods = set(stampli_direct_periods)
        fixed_rows = [row for row in fixed_rows if not (row.get("partner") == STAMPLI_FX_PARTNER and row.get("period") in covered_periods)]
        offline_detail_rows = [row for row in offline_detail_rows if not (row.get("partner") == STAMPLI_FX_PARTNER and row.get("period") in covered_periods)]
        fixed_rows.extend(stampli_direct_rows)
        offline_detail_rows.extend(stampli_direct_detail_rows)
        offline_periods = sorted({*offline_periods, *stampli_direct_periods})
    periods = sorted(period for period in {
        *(offline_periods or []),
        *(reversal_periods or []),
        *{row["period"] for row in revenue_detail_rows},
        *{row["period"] for row in revenue_summary},
        *{row["period"] for row in stampli_direct_rows},
        *{period for period in stampli_credit_complete_lookup.values()},
        *{
            month_key(
                row_value_first(
                    row,
                    "Refund Complete Date",
                    "Refund Completed Date",
                    "Transaction Lookup Dates Refund Complete Timestamp Date",
                    "Transaction Lookup Dates Refund Complete Timestamp Time",
                    patterns=("refundcompletedate", "refundcompleteddate", "refundcompletetimestampdate", "refundcompletetimestamptime"),
                )
            )
            for row in stampli_fx_reversal_rows
        },
    } if period)
    if args.period:
        periods = [args.period]
    register_rows = read_table(paths.offline_accounts) if paths.offline_accounts else []
    virtual_account_rows = build_virtual_account_usage(register_rows, account_activity, settlement_days, periods)
    stampli_fx_partner_payout_rows, stampli_fx_detail_rows = build_stampli_fx_partner_payouts(
        offline_detail_rows,
        periods,
        share_rows=stampli_fx_share_rows,
        reversal_rows=stampli_fx_reversal_rows,
        period=args.period,
        credit_complete_lookup=stampli_credit_complete_lookup,
    )
    detail_manifest = write_detail_files(args.output_detail_dir, offline_detail_rows + reversal_detail_rows + revenue_detail_export_rows + stampli_fx_detail_rows)

    gaps = [
        "Account-closing fees still need manual monthly counts. The supplied exports do not include an explicit closed-account event feed.",
        "Settlement sweeps are derived as one settlement day per partner per calendar day with transaction activity. If sweeps can happen multiple times per day, you will need a dedicated settlement export.",
        "Rev-share invoices now fall back to detail-level net revenue plus workbook contract config whenever the separate monthly summary export is absent.",
    ]
    if paths.stampli_credit_complete_all:
        gaps.append("Stampli Domestic and USD Abroad billing now uses the consolidated Stampli credit-complete export when present.")
    if paths.stampli_fx_share:
        gaps.append("Stampli FX partner-markup payout is now sourced from the dedicated Stampli FX revenue-share feed, bucketed by credit-complete lookup and netted by Refund Complete date.")
    else:
        gaps.append("Stampli FX partner-markup payout uses Payment USD Equivalent Amount as the daily mid-market USD equivalent when available, with a period-average fallback only if that field is missing.")
    if offline_meta["unmatchedPaymentIds"] > 0:
        gaps.append(
            f"{offline_meta['unmatchedPaymentIds']} fixed-billing payment IDs still do not have a usable partner mapping in the source export."
        )

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFolder": str(args.source_dir),
        "period": args.period or (periods[-1] if periods else ""),
        "periods": periods,
        "ltxn": fixed_rows + revenue_detail_rows,
        "lrev": reversal_rows,
        "lrs": revenue_summary,
        "lfxp": stampli_fx_partner_payout_rows,
        "lva": virtual_account_rows,
        "detailManifest": detail_manifest,
        "gaps": gaps,
        "meta": {
            "offline": {
                **offline_meta,
                "partners": dict(offline_meta["partners"]),
            },
            "reversals": {
                **reversal_meta,
                "partners": dict(reversal_meta["partners"]),
            },
            "revenue": {
                "detailSources": [str(path) for path in paths.revenue_txns],
                "summarySource": str(paths.revenue_summary) if paths.revenue_summary else "",
            },
            "stampliFx": {
                "creditCompleteSource": str(paths.stampli_credit_complete_all) if paths.stampli_credit_complete_all else "",
                "directDomesticSource": str(paths.stampli_domestic_revenue) if paths.stampli_domestic_revenue else "",
                "directUsdAbroadSource": str(paths.stampli_usd_abroad_revenue) if paths.stampli_usd_abroad_revenue else "",
                "shareSource": str(paths.stampli_fx_share) if paths.stampli_fx_share else "",
                "reversalSource": str(paths.stampli_fx_reversals) if paths.stampli_fx_reversals else "",
                "directBillingPeriods": stampli_direct_meta["periods"],
            },
        },
    }

    args.output_js.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_report.parent.mkdir(parents=True, exist_ok=True)
    args.output_detail_dir.mkdir(parents=True, exist_ok=True)

    write_js_module(args.output_js, payload)
    write_json(args.output_json, payload)
    write_report(args.output_report, payload, offline_meta, args.period)

    print(f"Wrote JS bundle to {args.output_js}")
    print(f"Wrote JSON bundle to {args.output_json}")
    print(f"Wrote report to {args.output_report}")
    print(f"Wrote {len(detail_manifest)} detail files to {args.output_detail_dir}")


if __name__ == "__main__":
    main()
