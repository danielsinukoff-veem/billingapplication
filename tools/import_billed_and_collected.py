from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from server.storage import SharedWorkspaceStore

DB_PATH = ROOT / "server" / "data" / "shared_workspace.db"
REPORT_DIR = ROOT / "reports" / "contract_audit"

DEFAULT_WORKBOOK = Path("/Users/danielsinukoff/Downloads/Partner Billing and Processes (2).xlsx")
SHEET_NAME = "Billed and Collected"
OPEN_STATUSES = {"CURRENT", "PAST DUE", "Wire Fees"}

PARTNER_ALIASES = {
    "alttpaynet": "Altpay",
    "alttpaynetcorp": "Altpay",
    "altypaynet": "Altpay",
    "altpaynet": "Altpay",
    "cellpay": "Cellpay",
    "gmeremit": "GME_Remit",
    "lianlian": "LianLian",
    "maplewave": "Maplewave",
    "nsave": "Nsave",
    "ovaltechgraph": "Graph Finance",
    "remittanceshub": "Remittanceshub",
    "somewhere": "Shepherd",
    "yeepay": "Yeepay",
}


def normalize_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def normalize_period(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = str(value or "").strip()
    if not text:
        return ""
    match = re.search(r"(\d{4})-(\d{2})", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    match = re.search(r"([A-Za-z]+)\s+(\d{4})", text)
    if match:
        try:
            parsed = datetime.strptime(f"{match.group(1)} {match.group(2)}", "%B %Y")
        except ValueError:
            try:
                parsed = datetime.strptime(f"{match.group(1)} {match.group(2)}", "%b %Y")
            except ValueError:
                return ""
        return parsed.strftime("%Y-%m")
    return ""


def normalize_iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, "", 0):
        return ""
    text = str(value).strip()
    if not text or text == "0":
        return ""
    for fmt in ("%Y-%m-%d", "%Y-%m", "%m/%d/%Y", "%Y/%m/%d", "%m/%d/%y"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%Y-%m":
                return f"{parsed.year:04d}-{parsed.month:02d}-01"
            return parsed.date().isoformat()
        except ValueError:
            continue
    return text


def parse_amount(value: Any) -> float:
    if value in (None, "", "None"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("$", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def round_currency(value: Any) -> float:
    return round(float(value or 0.0) + 1e-9, 2)


def match_partner(raw_partner: Any, partner_lookup: dict[str, str]) -> str:
    raw = str(raw_partner or "").strip()
    if not raw:
        return ""
    key = normalize_name(raw)
    if key in PARTNER_ALIASES:
        mapped = PARTNER_ALIASES[key]
        if normalize_name(mapped) in partner_lookup:
            return partner_lookup[normalize_name(mapped)]
    return partner_lookup.get(key, "")


def row_to_record(header: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header[index]: row[index]
        for index in range(min(len(header), len(row)))
    }


def build_imported_row(existing: dict[str, Any] | None, partner: str, period: str, payload: dict[str, Any]) -> dict[str, Any]:
    invoice_amount = round_currency(payload["invoice_amount"])
    amount_paid = round_currency(payload["amount_paid"])
    return {
        "id": existing.get("id") if existing else f"pi_{normalize_name(partner)}_{period}_receivable",
        "partner": partner,
        "period": period,
        "kind": "receivable",
        "invoiceDate": payload["invoice_date"],
        "paid": invoice_amount > 0 and amount_paid >= invoice_amount - 0.005,
        "amountPaid": amount_paid,
        "amountDueOverride": invoice_amount,
        "dueDateOverride": payload["due_date"],
        "paidDate": payload["paid_date"],
        "sourceSheet": SHEET_NAME,
        "sourceStatus": payload["source_status"],
        "sourceFeeInfo": payload["source_fee_info"],
        "note": payload["note"],
    }


def main() -> None:
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    store = SharedWorkspaceStore(DB_PATH)
    workspace = store.get_workspace()
    snapshot = workspace.get("snapshot") or {}
    partners = snapshot.get("ps") or []
    partner_lookup = {normalize_name(partner): partner for partner in partners}

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 11:
        raise SystemExit("Billed and Collected sheet is missing the expected header rows")

    header = [str(value).strip() if value is not None else f"col{index}" for index, value in enumerate(rows[10])]
    imported_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    unmatched_rows: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []

    for row in rows[11:]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        record = row_to_record(header, row)
        status = str(record.get("STATUS") or "").strip()
        if status not in OPEN_STATUSES:
            continue
        partner = match_partner(record.get("PARTNER"), partner_lookup)
        if not partner:
            unmatched_rows.append({
                "rawPartner": str(record.get("PARTNER") or "").strip(),
                "status": status,
                "period": normalize_period(record.get("BILLED PERIOD")),
                "invoiceAmount": round_currency(parse_amount(record.get("INVOICE AMOUNT"))),
                "outstandingAmount": round_currency(parse_amount(record.get("OUTSTANDING AMOUNT"))),
                "feeInfo": str(record.get("Fee Info") or "").strip(),
            })
            continue

        period = normalize_period(record.get("BILLED PERIOD"))
        if not period:
            skipped_rows.append({
                "partner": partner,
                "reason": "Missing billed period",
                "status": status,
                "feeInfo": str(record.get("Fee Info") or "").strip(),
            })
            continue

        invoice_amount = parse_amount(record.get("INVOICE AMOUNT"))
        amount_paid = parse_amount(record.get("PAID AMOUNT"))
        outstanding_amount = parse_amount(record.get("OUTSTANDING AMOUNT"))
        if invoice_amount <= 0 and outstanding_amount > 0:
            invoice_amount = outstanding_amount + amount_paid

        key = (partner, period, "receivable")
        bucket = imported_by_key.setdefault(key, {
            "invoice_amount": 0.0,
            "amount_paid": 0.0,
            "invoice_date_values": [],
            "due_date_values": [],
            "paid_date_values": [],
            "statuses": set(),
            "fee_infos": set(),
            "notes": set(),
        })
        bucket["invoice_amount"] += invoice_amount
        bucket["amount_paid"] += amount_paid
        invoice_date = normalize_iso_date(record.get("INVOICE DATE"))
        due_date = normalize_iso_date(record.get("DUE DATE"))
        paid_date = normalize_iso_date(record.get("PAID DATE"))
        if invoice_date:
            bucket["invoice_date_values"].append(invoice_date)
        if due_date:
            bucket["due_date_values"].append(due_date)
        if paid_date:
            bucket["paid_date_values"].append(paid_date)
        fee_info = str(record.get("Fee Info") or "").strip()
        if fee_info:
            bucket["fee_infos"].add(fee_info)
        note_parts = []
        note_value = str(record.get("Notes") or "").strip()
        write_off = str(record.get("WRITE OFFS") or "").strip()
        if note_value:
            note_parts.append(note_value)
        if write_off and write_off not in {"0", "0.0", "None"}:
            note_parts.append(f"Write offs: {write_off}")
        if note_parts:
            bucket["notes"].add(" | ".join(note_parts))
        bucket["statuses"].add(status)

    existing_rows = snapshot.get("pInvoices") or []
    imported_keys = set(imported_by_key)
    preserved_rows = []
    existing_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in existing_rows:
        row_partner = row.get("partner")
        row_period = normalize_period(row.get("period"))
        row_kind = "payable" if row.get("kind") == "payable" else "receivable"
        row_key = (row_partner, row_period, row_kind)
        if row.get("sourceSheet") == SHEET_NAME:
            continue
        if row_key in imported_keys and row_kind == "receivable":
            existing_by_key[row_key] = row
            continue
        preserved_rows.append(row)

    imported_rows = []
    for key, payload in sorted(imported_by_key.items()):
        partner, period, _kind = key
        imported_rows.append(build_imported_row(existing_by_key.get(key), partner, period, {
            "invoice_amount": payload["invoice_amount"],
            "amount_paid": payload["amount_paid"],
            "invoice_date": min(payload["invoice_date_values"]) if payload["invoice_date_values"] else "",
            "due_date": max(payload["due_date_values"]) if payload["due_date_values"] else "",
            "paid_date": max(payload["paid_date_values"]) if payload["paid_date_values"] else "",
            "source_status": " / ".join(sorted(payload["statuses"])),
            "source_fee_info": " | ".join(sorted(payload["fee_infos"])),
            "note": " | ".join(sorted(payload["notes"])),
        }))

    snapshot["pInvoices"] = sorted(
        preserved_rows + imported_rows,
        key=lambda row: (
            str(row.get("partner") or "").lower(),
            str(row.get("period") or ""),
            0 if row.get("kind") == "receivable" else 1,
        ),
    )
    saved_at = store.save_snapshot(snapshot)

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report = {
        "sourceWorkbook": str(workbook_path),
        "savedAt": saved_at,
        "importedCount": len(imported_rows),
        "importedRows": imported_rows,
        "unmatchedCount": len(unmatched_rows),
        "unmatchedRows": unmatched_rows,
        "skippedCount": len(skipped_rows),
        "skippedRows": skipped_rows,
    }
    report_path = REPORT_DIR / "billed_and_collected_import_report.json"
    report_path.write_text(json.dumps(report, indent=2))
    print(json.dumps({
        "savedAt": saved_at,
        "importedCount": len(imported_rows),
        "unmatchedCount": len(unmatched_rows),
        "skippedCount": len(skipped_rows),
        "reportPath": str(report_path),
    }, indent=2))


if __name__ == "__main__":
    main()
