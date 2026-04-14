from __future__ import annotations

import csv
import json
import sqlite3
import sys
from datetime import date
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill

ROOT = Path("/Users/danielsinukoff/Documents/billing-workbook")
SERVER_DIR = ROOT / "server"
if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))

import invoice_engine  # noqa: E402


ALIAS_MAP = {
    "Remittances Hub": "Remittanceshub",
    "YeePay": "Yeepay",
    "MapleWave": "Maplewave",
    "Gmeremit": "GME_Remit",
    "ALTPAYNET": "Altpay",
    "ALTPAYNET ": "Altpay",
    "CellPay": "Cellpay",
    "Oval Tech(Graph)": "Graph Finance",
    "Repay ": "Repay",
    "AJ Hanna (Whish)": "Whish",
    "Nomad Global": "Nomad",
    "Nuvion (Flutterwave)": "Nuvion",
}

PAYOUT_PERIOD_MAP = {
    "Q1 2025": ("2025-01", "2025-03"),
    "Q2 2025": ("2025-04", "2025-06"),
    "Q3 2025": ("2025-07", "2025-09"),
    "Q4 2025": ("2025-10", "2025-12"),
    "Q4-2024": ("2024-10", "2024-12"),
    "Nov & Dec2023": ("2023-11", "2023-12"),
    "Jan - April 2024": ("2024-01", "2024-04"),
    "Jul-Oct2023": ("2023-07", "2023-10"),
    "May&Jun 2024": ("2024-05", "2024-06"),
    "May & June 2024": ("2024-05", "2024-06"),
    "March and April 2024": ("2024-03", "2024-04"),
    "Apr-Oct2023": ("2023-04", "2023-10"),
    "Sep2022-Mar2023": ("2022-09", "2023-03"),
    "Oct & Nov 2024": ("2024-10", "2024-11"),
    "Aug-Oct 2025": ("2025-08", "2025-10"),
    "Nov - Dec 2025": ("2025-11", "2025-12"),
    "Partner Payout for 2020": ("2020-01", "2020-12"),
}


def normalize_partner(value: Any) -> str:
    text = str(value or "").strip()
    return ALIAS_MAP.get(text, text)


def normalize_period(value: Any) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m")
    return str(value or "").strip()


def money(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


PERIOD_RE = re.compile(r"(\d{4}-\d{2})")


def canonical_period(value: Any) -> str:
    text = str(value or "").strip()
    match = PERIOD_RE.search(text)
    return match.group(1) if match else ""


def row_overlaps_period_window(row: ComparisonRow, start_period: str, end_period: str) -> bool:
    row_start = canonical_period(row.billed_period_start) or canonical_period(row.billed_period)
    row_end = canonical_period(row.billed_period_end) or row_start
    if not row_start or not row_end:
        return False
    return row_start <= end_period and row_end >= start_period


def classify_payment_type(sheet: str, fee_info: str) -> str:
    normalized = str(fee_info or "").strip().upper()
    if sheet == "Partner Payouts":
        return "Partner Payout"
    if normalized == "IMP. FEE":
        return "Implementation Fee"
    if normalized in {"MONTHLY FEE", "MIN. FEE"}:
        return "Monthly / Recurring Fee"
    if normalized == "OTHER FEE":
        return "Other / Manual Adjustment"
    return "Other / Manual Adjustment"


def load_snapshot() -> dict[str, Any]:
    db_path = ROOT / "server" / "data" / "shared_workspace.db"
    conn = sqlite3.connect(db_path)
    row = conn.execute("select snapshot_json from workbook_snapshots order by id desc limit 1").fetchone()
    if not row:
        raise RuntimeError("No workbook snapshot found.")
    return json.loads(row[0])


def component_summary(groups: list[dict[str, Any]], direction: str) -> str:
    parts: list[str] = []
    for group in groups:
        amount = group.get("charge", 0.0) if direction == "charge" else group.get("pay", 0.0)
        if not amount:
            continue
        parts.append(f"{group.get('cat')}: {group.get('label')}={amount:,.2f}")
    return " | ".join(parts)


def source_stats(snapshot: dict[str, Any], partner: str, start_period: str, end_period: str | None = None) -> dict[str, Any]:
    periods = invoice_engine.enumerate_periods(start_period, end_period)
    stats = {
        "ltxn_rows": 0,
        "ltxn_txns": 0.0,
        "lrev_rows": 0,
        "lva_rows": 0,
        "lrs_rows": 0,
        "lfxp_rows": 0,
    }
    for period in periods:
        for key, row_key, counter_key in (
            ("ltxn", "txnCount", "ltxn_txns"),
            ("lrev", None, None),
            ("lva", None, None),
            ("lrs", None, None),
            ("lfxp", None, None),
        ):
            matches = [row for row in snapshot.get(key, []) if row.get("partner") == partner and row.get("period") == period]
            stats[f"{key}_rows"] += len(matches)
            if row_key:
                stats[counter_key] += sum(float(row.get(row_key) or 0) for row in matches)
    stats["ltxn_txns"] = round(stats["ltxn_txns"], 2)
    return stats


def current_platform_fee(snapshot: dict[str, Any], partner: str, period: str) -> float:
    total = 0.0
    for row in snapshot.get("plat", []):
        if row.get("partner") != partner:
            continue
        if invoice_engine.in_range(f"{period}-15", row.get("startDate"), row.get("endDate")):
            total += money(row.get("monthlyFee"))
    return round(total, 2)


def implementation_billing_months(snapshot: dict[str, Any], partner: str) -> list[str]:
    months = []
    for row in snapshot.get("impl", []):
        if row.get("partner") != partner or row.get("feeType") != "Implementation":
            continue
        billing_date = invoice_engine.get_implementation_billing_date(snapshot, partner, row)
        if billing_date:
            months.append(str(billing_date)[:7])
    return sorted(set(months))


@dataclass
class ComparisonRow:
    sheet: str
    partner: str
    in_current_app: bool
    payment_type: str
    fee_info: str
    billed_period: str
    billed_period_start: str
    billed_period_end: str
    invoice_amount: float
    paid_amount: float
    outstanding_amount: float
    status: str
    app_amount: float
    delta: float
    assessment: str
    review_bucket: str
    review_bucket_note: str
    reason_code: str
    explanation: str
    app_method: str
    app_calculation_detail: str
    app_components: str
    source_stats: str


def build_app_method(
    snapshot: dict[str, Any],
    partner: str,
    fee_info: str,
    sheet: str,
    start_period: str,
    end_period: str | None,
    app_invoice: dict[str, Any],
) -> str:
    periods = invoice_engine.enumerate_periods(start_period, end_period)
    stats = source_stats(snapshot, partner, start_period, end_period)
    charge_groups = [group for group in app_invoice.get("groups", []) if group.get("dir") == "charge" and group.get("charge")]
    pay_groups = [group for group in app_invoice.get("groups", []) if group.get("dir") == "pay" and group.get("pay")]
    labels = ", ".join(dict.fromkeys(str(group.get("label") or "") for group in (charge_groups if sheet == "Billed and Collected" else pay_groups) if group.get("label")))
    period_label = periods[0] if len(periods) == 1 else f"{periods[0]} to {periods[-1]}"

    if sheet == "Billed and Collected":
        if fee_info == "IMP. FEE":
            contract_start = invoice_engine.get_partner_contract_start_date(snapshot, partner) or "contract start date"
            return (
                f"Used the app's implementation-fee configuration. Implementation bills once in the contract-start month "
                f"({contract_start}). App amount comes from Impl Fee groups only."
            )
        if fee_info == "MIN. FEE":
            return (
                f"Used monthly minimum logic for {period_label}. App amount equals active Minimum groups after contract "
                f"go-live rules are applied."
            )
        if partner == "Stampli":
            return (
                f"Used the dedicated Stampli direct billing feed keyed to credit-complete month for ACH / USD Abroad charges, "
                f"plus current minimum logic where applicable. Components included: {labels or 'none'}."
            )
        if stats["lrs_rows"] > 0:
            return (
                f"Used imported Partner Revenue Summary rows for {period_label} ({stats['lrs_rows']} summary row(s)). "
                f"The app invoice engine summed recurring charge groups from those rows and then applied contract rules "
                f"for minimums / implementation separately. Components included: {labels or 'none'}."
            )
        detail_sources: list[str] = []
        if stats["ltxn_rows"]:
            detail_sources.append(f"{stats['ltxn_rows']} transaction summary row(s)")
        if stats["lrev_rows"]:
            detail_sources.append(f"{stats['lrev_rows']} reversal row(s)")
        if stats["lva_rows"]:
            detail_sources.append(f"{stats['lva_rows']} virtual-account row(s)")
        return (
            f"No revenue-summary rows were available for {period_label}, so the app fell back to raw imported activity and "
            f"contract configuration ({', '.join(detail_sources) or 'no supporting activity'}). Components included: {labels or 'none'}."
        )

    if partner == "Stampli":
        return (
            f"Used the direct Stampli FX revenue-share / reversal feeds over {period_label}. App payout equals the pay-side "
            f"groups from that feed. Components included: {labels or 'none'}."
        )
    if stats["lrs_rows"] > 0:
        return (
            f"Used imported rev-share summary / reversal rows over {period_label} ({stats['lrs_rows']} supporting row(s)). "
            f"App payout equals the pay-side invoice groups built from those source rows. Components included: {labels or 'none'}."
        )
    if stats["lfxp_rows"] > 0:
        return (
            f"Used FX payout rows over {period_label} ({stats['lfxp_rows']} row(s)). App payout equals the pay-side groups "
            f"from those FX rows. Components included: {labels or 'none'}."
        )
    return (
        f"No matching payout source rows were available for {period_label}, so the app could not reproduce the historical "
        f"payout. Current pay-side components: {labels or 'none'}."
    )


def determine_review_bucket(
    sheet: str,
    partner: str,
    fee_info: str,
    status: str,
    assessment: str,
    reason_code: str,
) -> tuple[str, str]:
    if assessment == "Match":
        return ("Match", "Historical billed amount and app-calculated amount match.")

    if reason_code == "partial_month_minimum":
        return (
            "Needs Policy Decision",
            "This is a contract-policy question about whether first-month minimums should be prorated.",
        )

    if assessment in {"App likely correct", "Historical workbook likely wrong", "Historical workbook likely manual"} or reason_code in {
        "pre_go_live_billing",
        "oson_contract_minimum",
        "historical_below_platform_fee",
        "skydo_extra_charge_row",
        "manual_other_fee",
    }:
        return (
            "Likely Historical Billing Issue",
            "The app logic aligns with the current contract/source rules more closely than the historical billing entry.",
        )

    if assessment == "App incomplete" or reason_code in {
        "missing_revenue_summary",
        "missing_revenue_summary_and_activity",
        "missing_referral_payout_logic",
        "missing_payout_source",
        "partner_not_configured",
        "missing_implementation_config",
    }:
        return (
            "Likely Still App Logic Issue",
            "The app is missing source data, configuration, or billing logic needed to reproduce this line.",
        )

    if reason_code in {
        "small_basis_difference",
        "stampli_charge_basis_difference",
        "stampli_fx_basis_difference",
        "everflow_revshare_source_difference",
        "shepherd_rollup_difference",
        "implementation_timing_difference",
    } or assessment in {"Likely source-basis difference", "Needs source review"}:
        return (
            "Likely Source-Basis Difference",
            "The app and the historical workbook appear to be using different source extracts, month basis, or rollup rules.",
        )

    if reason_code == "manual_review":
        if partner in {"Nuvion", "Magaya"}:
            return (
                "Likely Still App Logic Issue",
                "This looks like the app is missing a fee construct or payout mapping rather than simply using a different source basis.",
            )
        if partner == "Skydo" and sheet == "Partner Payouts":
            return (
                "Likely Still App Logic Issue",
                "The historical payout exists but the app has no matching payout source rows, which points to a missing app-side payout import/config.",
            )
        if partner in {"Skydo", "Whish"} and sheet == "Billed and Collected":
            return (
                "Likely Historical Billing Issue",
                "The app is using explicit billed components from imported summary/source rows, while the historical workbook appears to have waived, excluded, or manually adjusted part of those charges.",
            )
        if partner in {"Remittanceshub", "Nsave", "Maplewave"}:
            return (
                "Likely Source-Basis Difference",
                "The app is calculating from the direct billing summary source, while the historical billed amount appears to follow a different source cut or invoice treatment.",
            )
        return (
            "Likely Source-Basis Difference",
            "This mismatch looks more like a source-basis difference than a deterministic app calculation bug, but it still needs review.",
        )

    return (
        "Likely Source-Basis Difference",
        "Defaulted to source-basis difference because the row does not map to a confirmed app bug.",
    )


def calculation_detail(groups: list[dict[str, Any]], direction: str) -> str:
    details: list[str] = []
    for group in groups:
        used_line = False
        for line in group.get("lines") or []:
            if str(line.get("dir") or "") != direction:
                continue
            amount = float(line.get("amount") or 0)
            desc = str(line.get("desc") or group.get("summary") or group.get("label") or "").strip()
            details.append(f"{group.get('cat')}: {group.get('label')} -> {desc} = {amount:,.2f}")
            used_line = True
        if used_line:
            continue
        amount = float(group.get("charge") or 0) if direction == "charge" else float(group.get("pay") or 0)
        if not amount:
            continue
        summary = str(group.get("summary") or group.get("label") or "").strip()
        details.append(f"{group.get('cat')}: {group.get('label')} -> {summary} = {amount:,.2f}")
    return " | ".join(details)


def classify_receivable(
    snapshot: dict[str, Any],
    partner: str,
    period: str,
    fee_info: str,
    historical_amount: float,
    app_amount: float,
    app_invoice: dict[str, Any],
) -> tuple[str, str, str]:
    delta = round(app_amount - historical_amount, 2)
    stats = source_stats(snapshot, partner, period)
    platform_fee = current_platform_fee(snapshot, partner, period)
    recurring_live = invoice_engine.is_recurring_billing_live_for_period(snapshot, partner, period)
    impl_months = implementation_billing_months(snapshot, partner)
    if abs(delta) <= 0.01:
        return ("Match", "match", "Current app and historical workbook agree.")
    if fee_info == "IMP. FEE":
        if app_amount == 0 and impl_months:
            return (
                "App likely correct",
                "implementation_timing_difference",
                f"Historical workbook bills implementation in {period}, but current app bills implementation in {', '.join(impl_months)} based on contract start date.",
            )
        return (
            "App incomplete",
            "missing_implementation_config",
            "Historical workbook has an implementation fee, but the current app has no matching implementation fee configured for this partner/period.",
        )
    if not recurring_live and historical_amount > 0:
        go_live = invoice_engine.get_partner_go_live_date(snapshot, partner) or "not set"
        return (
            "App likely correct",
            "pre_go_live_billing",
            f"Historical workbook billed recurring fees before go-live, while the app suppresses recurring billing until the configured go-live date ({go_live}).",
        )
    if partner == "Stampli" and fee_info in {"MONTHLY FEE", "MIN. FEE"}:
        return (
            "Needs source review",
            "stampli_charge_basis_difference",
            "Stampli charge rows in the app come from the dedicated direct billing feed keyed to the current credit-complete logic. The historical workbook amount appears to come from a slightly different extract or month cut, so this is not a simple rate-calculation bug.",
        )
    if partner == "Oson" and period >= "2025-04":
        return (
            "App likely correct",
            "oson_contract_minimum",
            "Oson's contract says the monthly minimum is $7,500 from December 2024 onward. The workbook's $2,500 monthly rows and the separate $500 September 2025 other-fee row do not match that contract schedule.",
        )
    if partner == "Nomad" and period == "2025-01" and fee_info == "MIN. FEE":
        return (
            "Needs policy decision",
            "partial_month_minimum",
            "The workbook prorated Nomad's first-month minimum to $4,516. The app applies the full $10,000 minimum from the contract start month. This is a proration-policy difference, not a math bug.",
        )
    if partner == "Skydo" and period == "2025-12" and historical_amount > 10000:
        return (
            "Historical workbook likely wrong",
            "skydo_extra_charge_row",
            "The extra $22,701.82 Skydo December row does not line up with the contract-backed partner billing model in the app. It looks like payer-borne premium/card revenue was historically mixed into the partner invoice.",
        )
    if fee_info == "OTHER FEE":
        return (
            "Historical workbook likely manual",
            "manual_other_fee",
            "This row is tracked as an 'OTHER FEE' in the workbook, but there is no matching automated billing construct in the app. It appears to be a manual adjustment.",
        )
    if platform_fee and historical_amount + 1000 < platform_fee:
        return (
            "App likely correct",
            "historical_below_platform_fee",
            f"The workbook amount is materially below the configured monthly platform/subscription fee of ${platform_fee:,.2f}, so the historical billing likely reflected a concession, credit, or off-app adjustment.",
        )
    if stats["lrs_rows"] == 0 and fee_info in {"MONTHLY FEE", "MIN. FEE"}:
        if app_amount == 0:
            return (
                "App incomplete",
                "missing_revenue_summary_and_activity",
                "The current app snapshot has no revenue-summary rows and no matching imported activity supporting this billed amount, so the app cannot reproduce this historical invoice yet.",
            )
        if historical_amount > app_amount + 100:
            return (
                "App incomplete",
                "missing_revenue_summary",
                "The app has no revenue-summary rows for this partner/period, so it is falling back to raw contract/minimum logic. The historical invoice is higher and likely used a billing-summary source that is not currently loaded.",
            )
    if abs(delta) <= 200:
        return (
            "Likely source-basis difference",
            "small_basis_difference",
            "The delta is small enough that it most likely comes from a different source extract, cut-off, or rounding basis rather than a calculation bug.",
        )
    return (
        "Needs manual review",
        "manual_review",
        "This mismatch does not cleanly map to a known app bug. It needs line-by-line source review against the underlying exported report for that month.",
    )


def classify_payout(
    snapshot: dict[str, Any],
    partner: str,
    period_label: str,
    start_period: str,
    end_period: str,
    fee_info: str,
    historical_amount: float,
    app_amount: float,
) -> tuple[str, str, str]:
    delta = round(app_amount - historical_amount, 2)
    stats = source_stats(snapshot, partner, start_period, end_period)
    if abs(delta) <= 0.01:
        return ("Match", "match", "Current app and historical workbook agree.")
    if fee_info == "REFERRAL OUT" and app_amount == 0:
        return (
            "App incomplete",
            "missing_referral_payout_logic",
            "The workbook includes a referral payout, but the current app has no matching referral-payout source or configuration for this row.",
        )
    if partner == "Stampli":
        return (
            "Needs source review",
            "stampli_fx_basis_difference",
            "Stampli payouts come from the direct FX revenue-share feed. The app uses the current credit-complete/refund-complete feed, while the historical workbook appears to have used a different source extract or month basis.",
        )
    if partner == "Everflow":
        return (
            "Needs source review",
            "everflow_revshare_source_difference",
            "Everflow payout differences are consistent with revenue-share source drift. The current app uses the deduped/imported revenue-share feed, while the historical workbook likely used a different exported set or duplicate payment IDs.",
        )
    if partner == "Shepherd":
        return (
            "Needs source review",
            "shepherd_rollup_difference",
            "Shepherd historical payouts are partly quarterly/range-based. The current app rolls monthly payout rows together, so differences here are likely source-range aggregation issues rather than a simple math bug.",
        )
    if app_amount == 0 and (stats["lrs_rows"] == 0 and stats["lfxp_rows"] == 0):
        return (
            "App incomplete",
            "missing_payout_source",
            "The current app has no supporting payout source rows for this partner/period range, so it cannot reproduce the historical payout yet.",
        )
    if abs(delta) <= 100:
        return (
            "Likely source-basis difference",
            "small_basis_difference",
            "The delta is small enough that it most likely comes from a different source extract, cut-off, or rounding basis rather than a calculation bug.",
        )
    return (
        "Needs manual review",
        "manual_review",
        "This payout mismatch needs line-by-line source review against the exported payout/rev-share feed for that period.",
    )


def build_receivable_rows(snapshot: dict[str, Any], workbook_path: Path) -> list[ComparisonRow]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["Billed and Collected"]
    current_partners = set(snapshot.get("ps", []))
    invoice_cache: dict[tuple[str, str, str | None], dict[str, Any]] = {}
    comparisons: list[ComparisonRow] = []

    def get_invoice(partner: str, start_period: str, end_period: str | None = None) -> dict[str, Any]:
        key = (partner, start_period, end_period)
        if key not in invoice_cache:
            invoice_cache[key] = invoice_engine.calculate_invoice(snapshot, partner, start_period, end_period)
        return invoice_cache[key]

    rows = list(ws.iter_rows(min_row=12, values_only=True))

    def has_min_row(partner: str, period: str) -> bool:
        for row in rows:
            if normalize_partner(row[0]) != partner:
                continue
            billed_period = row[2]
            if hasattr(billed_period, "strftime") and billed_period.strftime("%Y-%m") == period and str(row[1] or "") == "MIN. FEE":
                return True
        return False

    for row in rows:
        if not row[0]:
            continue
        partner = normalize_partner(row[0])
        billed_period = row[2]
        if not hasattr(billed_period, "strftime"):
            continue
        period = billed_period.strftime("%Y-%m")
        fee_info = str(row[1] or "")
        invoice_amount = money(row[6])
        paid_amount = money(row[9])
        outstanding_amount = money(row[10])
        status = str(row[12] or "").strip()
        in_current_app = partner in current_partners
        app_amount = 0.0
        app_components = ""
        app_calculation = ""
        assessment = "Not in current app"
        reason_code = "partner_not_configured"
        explanation = "This partner is present in the historical workbook but is not currently configured as an app partner."
        if in_current_app:
            app_invoice = get_invoice(partner, period)
            minimum_groups = [group for group in app_invoice["groups"] if group["cat"] == "Minimum" and group["charge"]]
            impl_groups = [group for group in app_invoice["groups"] if group["cat"] == "Impl Fee" and group["charge"]]
            recurring_groups = [
                group
                for group in app_invoice["groups"]
                if group["dir"] == "charge" and group["charge"] and group["cat"] not in {"Minimum", "Impl Fee"}
            ]
            recurring_amount = round(
                sum(
                    group["charge"]
                    for group in recurring_groups
                ),
                2,
            )
            minimum_amount = round(sum(group["charge"] for group in minimum_groups), 2)
            implementation_amount = round(sum(group["charge"] for group in impl_groups), 2)
            selected_groups: list[dict[str, Any]]
            if fee_info == "IMP. FEE":
                app_amount = implementation_amount
                selected_groups = impl_groups
            elif fee_info == "MIN. FEE":
                app_amount = minimum_amount
                selected_groups = minimum_groups
            elif fee_info == "MONTHLY FEE":
                if has_min_row(partner, period):
                    app_amount = recurring_amount
                    selected_groups = recurring_groups
                else:
                    app_amount = round(recurring_amount + minimum_amount, 2)
                    selected_groups = [*recurring_groups, *minimum_groups]
            elif fee_info == "OTHER FEE":
                app_amount = 0.0
                selected_groups = []
            else:
                app_amount = round(app_invoice["chg"], 2)
                selected_groups = [group for group in app_invoice["groups"] if group["dir"] == "charge" and group["charge"]]
            app_components = component_summary(selected_groups, "charge")
            app_calculation = calculation_detail(selected_groups, "charge")
            app_method = build_app_method(snapshot, partner, fee_info, "Billed and Collected", period, None, app_invoice)
            assessment, reason_code, explanation = classify_receivable(
                snapshot, partner, period, fee_info, invoice_amount, app_amount, app_invoice
            )
        else:
            app_method = "Partner is not configured in the current app snapshot, so there is no app-side calculation to compare."
        review_bucket, review_bucket_note = determine_review_bucket(
            "Billed and Collected",
            partner,
            fee_info,
            status,
            assessment,
            reason_code,
        )

        comparisons.append(
            ComparisonRow(
                sheet="Billed and Collected",
                partner=partner,
                in_current_app=in_current_app,
                payment_type=classify_payment_type("Billed and Collected", fee_info),
                fee_info=fee_info,
                billed_period=period,
                billed_period_start=period,
                billed_period_end=period,
                invoice_amount=invoice_amount,
                paid_amount=paid_amount,
                outstanding_amount=outstanding_amount,
                status=status,
                app_amount=round(app_amount, 2),
                delta=round(app_amount - invoice_amount, 2),
                assessment=assessment,
                review_bucket=review_bucket,
                review_bucket_note=review_bucket_note,
                reason_code=reason_code,
                explanation=explanation,
                app_method=app_method,
                app_calculation_detail=app_calculation,
                app_components=app_components,
                source_stats=json.dumps(source_stats(snapshot, partner, period), sort_keys=True),
            )
        )
    return comparisons


def build_payout_rows(snapshot: dict[str, Any], workbook_path: Path) -> list[ComparisonRow]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["Partner Payouts"]
    current_partners = set(snapshot.get("ps", []))
    invoice_cache: dict[tuple[str, str, str | None], dict[str, Any]] = {}
    comparisons: list[ComparisonRow] = []

    def get_invoice(partner: str, start_period: str, end_period: str | None = None) -> dict[str, Any]:
        key = (partner, start_period, end_period)
        if key not in invoice_cache:
            invoice_cache[key] = invoice_engine.calculate_invoice(snapshot, partner, start_period, end_period)
        return invoice_cache[key]

    for row in ws.iter_rows(min_row=12, values_only=True):
        if not row[0]:
            continue
        partner = normalize_partner(row[0])
        period_value = row[2]
        fee_info = str(row[1] or "")
        invoice_amount = money(row[6])
        paid_amount = money(row[9])
        outstanding_amount = money(row[10])
        status = str(row[12] or "").strip()
        in_current_app = partner in current_partners
        if hasattr(period_value, "strftime"):
            start_period = period_value.strftime("%Y-%m")
            end_period = start_period
            period_label = str(period_value)
        else:
            period_label = str(period_value or "").strip()
            start_period, end_period = PAYOUT_PERIOD_MAP.get(period_label, ("", ""))
        app_amount = 0.0
        app_components = ""
        app_calculation = ""
        assessment = "Not in current app"
        reason_code = "partner_not_configured"
        explanation = "This partner is present in the historical workbook but is not currently configured as an app partner."
        if in_current_app and start_period:
            app_invoice = get_invoice(partner, start_period, end_period if end_period and end_period != start_period else None)
            app_amount = round(app_invoice["pay"], 2)
            selected_groups = [group for group in app_invoice["groups"] if group["dir"] == "pay" and group["pay"]]
            app_components = component_summary(selected_groups, "pay")
            app_calculation = calculation_detail(selected_groups, "pay")
            app_method = build_app_method(
                snapshot,
                partner,
                fee_info,
                "Partner Payouts",
                start_period,
                end_period if end_period and end_period != start_period else None,
                app_invoice,
            )
            assessment, reason_code, explanation = classify_payout(
                snapshot, partner, period_label, start_period, end_period or start_period, fee_info, invoice_amount, app_amount
            )
        elif in_current_app and not start_period:
            assessment = "Needs manual review"
            reason_code = "unparsed_historical_range"
            explanation = "The historical payout range label could not be mapped to invoice months automatically."
            app_method = "Historical payout range label could not be mapped to invoice months, so no app-side payout calculation was run."
        else:
            app_method = "Partner is not configured in the current app snapshot, so there is no app-side payout calculation to compare."
        review_bucket, review_bucket_note = determine_review_bucket(
            "Partner Payouts",
            partner,
            fee_info,
            status,
            assessment,
            reason_code,
        )
        comparisons.append(
            ComparisonRow(
                sheet="Partner Payouts",
                partner=partner,
                in_current_app=in_current_app,
                payment_type=classify_payment_type("Partner Payouts", fee_info),
                fee_info=fee_info,
                billed_period=period_label,
                billed_period_start=start_period,
                billed_period_end=end_period or start_period,
                invoice_amount=invoice_amount,
                paid_amount=paid_amount,
                outstanding_amount=outstanding_amount,
                status=status,
                app_amount=round(app_amount, 2),
                delta=round(app_amount - invoice_amount, 2),
                assessment=assessment,
                review_bucket=review_bucket,
                review_bucket_note=review_bucket_note,
                reason_code=reason_code,
                explanation=explanation,
                app_method=app_method,
                app_calculation_detail=app_calculation,
                app_components=app_components,
                source_stats=json.dumps(
                    source_stats(snapshot, partner, start_period, end_period or start_period) if start_period else {},
                    sort_keys=True,
                ),
            )
        )
    return comparisons


def write_csv(rows: list[ComparisonRow], path: Path) -> None:
    fieldnames = [
        "sheet",
        "partner",
        "in_current_app",
        "payment_type",
        "fee_info",
        "billed_period",
        "billed_period_start",
        "billed_period_end",
        "invoice_amount",
        "paid_amount",
        "outstanding_amount",
        "status",
        "app_amount",
        "delta",
        "assessment",
        "review_bucket",
        "review_bucket_note",
        "reason_code",
        "explanation",
        "app_method",
        "app_calculation_detail",
        "app_components",
        "source_stats",
    ]
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.__dict__)


def write_markdown(rows: list[ComparisonRow], path: Path) -> None:
    current_rows = [row for row in rows if row.in_current_app]
    mismatches = [row for row in current_rows if abs(row.delta) > 0.01]
    reason_counts = Counter(row.reason_code for row in mismatches)
    assessment_counts = Counter(row.assessment for row in mismatches)
    top_rows = sorted(mismatches, key=lambda row: abs(row.delta), reverse=True)[:30]
    lines = [
        "# Partner Billing Reconciliation",
        "",
        f"- Compared `{len(rows)}` historical workbook lines against the latest billing app snapshot.",
        f"- Current app partner lines reviewed: `{len(current_rows)}`",
        f"- Current app mismatches: `{len(mismatches)}`",
        "",
        "## Mismatch Assessment Counts",
        "",
    ]
    for assessment, count in assessment_counts.most_common():
        lines.append(f"- `{assessment}`: `{count}`")
    lines.extend(["", "## Top Mismatches", ""])
    for row in top_rows:
        lines.append(
            f"- `{row.sheet}` | `{row.partner}` | `{row.billed_period}` | `{row.fee_info}`: historical `{row.invoice_amount:,.2f}`, app `{row.app_amount:,.2f}`, delta `{row.delta:,.2f}`"
        )
        lines.append(f"  - Assessment: `{row.assessment}`")
        lines.append(f"  - Reason: `{row.reason_code}`")
        lines.append(f"  - Explanation: {row.explanation}")
        if row.app_components:
            lines.append(f"  - App components: {row.app_components}")
    lines.extend(["", "## Reason Codes", ""])
    for reason, count in reason_counts.most_common():
        lines.append(f"- `{reason}`: `{count}`")
    path.write_text("\n".join(lines) + "\n")


def write_xlsx(rows: list[ComparisonRow], path: Path) -> None:
    workbook = Workbook()
    all_sheet = workbook.active
    all_sheet.title = "All Lines"
    simple_totals_sheet = workbook.create_sheet("Simple Totals")
    simple_recent_totals_sheet = workbook.create_sheet("Simple Totals Dec2025+")
    totals_by_type_sheet = workbook.create_sheet("Totals By Type")
    totals_by_type_recent_sheet = workbook.create_sheet("Totals By Type Dec2025+")
    mismatch_sheet = workbook.create_sheet("Mismatches Only")
    historical_sheet = workbook.create_sheet("Likely Historical Billing")
    source_basis_sheet = workbook.create_sheet("Likely Source-Basis Diff")
    app_sheet = workbook.create_sheet("Likely App Logic Issue")
    top_app_sheet = workbook.create_sheet("Top App Logic Issues")
    policy_sheet = workbook.create_sheet("Needs Policy Decision")
    partner_totals_sheet = workbook.create_sheet("Partner Totals")
    summary_sheet = workbook.create_sheet("Summary")

    headers = [
        "Source Tab",
        "Partner",
        "In Current App",
        "Payment Type",
        "Fee Info",
        "Billed Period",
        "Billed Period Start",
        "Billed Period End",
        "Historical Billed Amount",
        "Historical Paid Amount",
        "Historical Outstanding Amount",
        "Historical Status",
        "App Calculated Amount",
        "Delta (App - Historical)",
        "Delta vs Paid",
        "Assessment",
        "Review Bucket",
        "Review Bucket Note",
        "Reason Code",
        "Explanation",
        "App Calculation Method",
        "App Calculation Detail",
        "App Components",
        "Source Stats",
    ]

    def append_table(ws, data_rows: list[ComparisonRow]) -> None:
        ws.append(headers)
        for row in data_rows:
            ws.append(
                [
                    row.sheet,
                    row.partner,
                    row.in_current_app,
                    row.payment_type,
                    row.fee_info,
                    row.billed_period,
                    row.billed_period_start,
                    row.billed_period_end,
                    row.invoice_amount,
                    row.paid_amount,
                    row.outstanding_amount,
                    row.status,
                    row.app_amount,
                    row.delta,
                    round(row.app_amount - row.paid_amount, 2),
                    row.assessment,
                    row.review_bucket,
                    row.review_bucket_note,
                    row.reason_code,
                    row.explanation,
                    row.app_method,
                    row.app_calculation_detail,
                    row.app_components,
                    row.source_stats,
                ]
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wrapped_columns = {"Q", "R", "S", "T", "U"}
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
        widths = {
            "A": 20, "B": 18, "C": 14, "D": 22, "E": 14, "F": 18, "G": 14, "H": 14, "I": 18,
            "J": 18, "K": 22, "L": 18, "M": 18, "N": 20, "O": 16, "P": 24, "Q": 45,
            "R": 28, "S": 60, "T": 70, "U": 70, "V": 60, "W": 30,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in {"R", "S", "T", "U", "V"}:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    def append_partner_totals(ws, data_rows: list[ComparisonRow]) -> None:
        total_headers = [
            "Partner",
            "In Current App",
            "Historical Charges",
            "App Charges",
            "Charge Delta",
            "Historical Payouts",
            "App Payouts",
            "Payout Delta",
            "Historical Net",
            "App Net",
            "Net Delta",
            "Mismatch Rows",
            "Mismatch Periods",
            "Likely Timing / Period Shift",
            "Timing Note",
        ]
        ws.append(total_headers)
        partner_map: dict[str, dict[str, Any]] = {}
        for row in data_rows:
            bucket = partner_map.setdefault(
                row.partner,
                {
                    "partner": row.partner,
                    "in_current_app": row.in_current_app,
                    "historical_charges": 0.0,
                    "app_charges": 0.0,
                    "historical_payouts": 0.0,
                    "app_payouts": 0.0,
                    "mismatch_rows": 0,
                    "mismatch_periods": set(),
                },
            )
            bucket["in_current_app"] = bucket["in_current_app"] or row.in_current_app
            if row.sheet == "Billed and Collected":
                bucket["historical_charges"] += row.invoice_amount
                bucket["app_charges"] += row.app_amount
            else:
                bucket["historical_payouts"] += row.invoice_amount
                bucket["app_payouts"] += row.app_amount
            if abs(row.delta) > 0.01:
                bucket["mismatch_rows"] += 1
                if row.billed_period:
                    bucket["mismatch_periods"].add(row.billed_period)

        partner_rows: list[list[Any]] = []
        for partner in sorted(partner_map):
            bucket = partner_map[partner]
            historical_charges = round(bucket["historical_charges"], 2)
            app_charges = round(bucket["app_charges"], 2)
            historical_payouts = round(bucket["historical_payouts"], 2)
            app_payouts = round(bucket["app_payouts"], 2)
            charge_delta = round(app_charges - historical_charges, 2)
            payout_delta = round(app_payouts - historical_payouts, 2)
            historical_net = round(historical_charges - historical_payouts, 2)
            app_net = round(app_charges - app_payouts, 2)
            net_delta = round(app_net - historical_net, 2)
            mismatch_rows = bucket["mismatch_rows"]
            mismatch_periods = sorted(bucket["mismatch_periods"])
            basis = max(
                abs(historical_charges),
                abs(app_charges),
                abs(historical_payouts),
                abs(app_payouts),
                abs(historical_net),
                abs(app_net),
                1.0,
            )
            timing_tolerance = round(max(100.0, basis * 0.01), 2)
            likely_timing = (
                mismatch_rows > 0
                and abs(net_delta) <= timing_tolerance
                and abs(charge_delta) <= timing_tolerance
                and abs(payout_delta) <= timing_tolerance
            )
            if likely_timing:
                timing_note = (
                    "Overall partner totals are close even though individual months differ. "
                    "Review billing-month assignment, source cut, and rollup rules before treating this as a true billing error."
                )
            elif mismatch_rows > 0:
                timing_note = (
                    "Overall partner totals are still materially different, so this is more than a simple timing shift."
                )
            else:
                timing_note = "Totals match."
            partner_rows.append(
                [
                    partner,
                    bucket["in_current_app"],
                    historical_charges,
                    app_charges,
                    charge_delta,
                    historical_payouts,
                    app_payouts,
                    payout_delta,
                    historical_net,
                    app_net,
                    net_delta,
                    mismatch_rows,
                    ", ".join(mismatch_periods),
                    "Yes" if likely_timing else "No",
                    timing_note,
                ]
            )

        partner_rows.sort(key=lambda row: abs(float(row[10])), reverse=True)
        for row in partner_rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
        widths = {
            "A": 22,
            "B": 14,
            "C": 18,
            "D": 18,
            "E": 16,
            "F": 18,
            "G": 18,
            "H": 16,
            "I": 16,
            "J": 16,
            "K": 14,
            "L": 14,
            "M": 30,
            "N": 20,
            "O": 95,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in {"M", "N", "O"}:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        return partner_rows

    def append_simple_totals(ws, partner_rows: list[list[Any]]) -> None:
        headers = ["Partner", "Billing Total", "App Total", "Delta"]
        ws.append(headers)
        simple_rows = []
        for row in partner_rows:
            simple_rows.append([row[0], row[8], row[9], row[10]])
        simple_rows.sort(key=lambda row: abs(float(row[3])), reverse=True)
        for row in simple_rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
        widths = {"A": 22, "B": 16, "C": 16, "D": 16}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def append_totals_by_type(ws, data_rows: list[ComparisonRow]) -> None:
        headers = ["Partner", "Payment Type", "Billing Total", "App Total", "Delta"]
        ws.append(headers)
        buckets: dict[tuple[str, str], dict[str, Any]] = {}
        for row in data_rows:
            key = (row.partner, row.payment_type)
            bucket = buckets.setdefault(
                key,
                {
                    "partner": row.partner,
                    "payment_type": row.payment_type,
                    "billing_total": 0.0,
                    "app_total": 0.0,
                },
            )
            bucket["billing_total"] += row.invoice_amount
            bucket["app_total"] += row.app_amount
        out_rows = []
        for _, bucket in buckets.items():
            billing_total = round(bucket["billing_total"], 2)
            app_total = round(bucket["app_total"], 2)
            delta = round(app_total - billing_total, 2)
            out_rows.append([bucket["partner"], bucket["payment_type"], billing_total, app_total, delta])
        out_rows.sort(key=lambda row: (row[0], row[1]))
        for row in out_rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
        widths = {"A": 22, "B": 24, "C": 16, "D": 16, "E": 16}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    append_table(all_sheet, rows)
    mismatches = [row for row in rows if abs(row.delta) > 0.01]
    append_table(mismatch_sheet, mismatches)

    likely_historical = [
        row
        for row in mismatches
        if row.review_bucket == "Likely Historical Billing Issue"
    ]
    likely_source_basis = [
        row
        for row in mismatches
        if row.review_bucket == "Likely Source-Basis Difference"
    ]
    likely_app = [
        row
        for row in mismatches
        if row.review_bucket == "Likely Still App Logic Issue"
    ]
    top_app = sorted(likely_app, key=lambda row: abs(row.delta), reverse=True)[:20]
    policy_review = [row for row in mismatches if row.review_bucket == "Needs Policy Decision"]
    append_table(historical_sheet, likely_historical)
    append_table(source_basis_sheet, likely_source_basis)
    append_table(app_sheet, likely_app)
    append_table(top_app_sheet, top_app)
    append_table(policy_sheet, policy_review)
    partner_rows = append_partner_totals(partner_totals_sheet, rows)
    append_simple_totals(simple_totals_sheet, partner_rows)
    append_totals_by_type(totals_by_type_sheet, rows)
    current_year_periods = [
        canonical_period(row.billed_period_start) or canonical_period(row.billed_period_end) or canonical_period(row.billed_period)
        for row in rows
    ]
    current_year_periods = [period for period in current_year_periods if period.startswith("2026-")]
    recent_end_period = max(current_year_periods) if current_year_periods else date.today().strftime("%Y-%m")
    recent_rows = [row for row in rows if row_overlaps_period_window(row, "2025-12", recent_end_period)]
    recent_partner_rows = append_partner_totals(workbook.create_sheet("_tmp_recent_partner_totals"), recent_rows)
    workbook.remove(workbook["_tmp_recent_partner_totals"])
    append_simple_totals(simple_recent_totals_sheet, recent_partner_rows)
    append_totals_by_type(totals_by_type_recent_sheet, recent_rows)

    current_rows = [row for row in rows if row.in_current_app]
    reason_counts = Counter(row.reason_code for row in mismatches)
    assessment_counts = Counter(row.assessment for row in mismatches)
    bucket_counts = Counter(row.review_bucket for row in mismatches)
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Historical workbook lines reviewed", len(rows)])
    summary_sheet.append(["Current app partner lines reviewed", len(current_rows)])
    summary_sheet.append(["Mismatches", len(mismatches)])
    summary_sheet.append(["Likely historical billing error", len(likely_historical)])
    summary_sheet.append(["Likely source-basis difference", len(likely_source_basis)])
    summary_sheet.append(["Likely app logic / missing logic", len(likely_app)])
    summary_sheet.append(["Top app logic issues tab rows", len(top_app)])
    summary_sheet.append(["Needs policy decision", len(policy_review)])
    summary_sheet.append([])
    summary_sheet.append(["Review Bucket", "Count"])
    for bucket, count in bucket_counts.most_common():
        summary_sheet.append([bucket, count])
    summary_sheet.append([])
    summary_sheet.append(["Assessment", "Count"])
    for assessment, count in assessment_counts.most_common():
        summary_sheet.append([assessment, count])
    summary_sheet.append([])
    summary_sheet.append(["Reason Code", "Count"])
    for reason, count in reason_counts.most_common():
        summary_sheet.append([reason, count])
    summary_sheet.freeze_panes = "A2"
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in summary_sheet[6]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in summary_sheet[summary_sheet.max_row - len(reason_counts)]:
        pass
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 18

    workbook.save(path)


def main() -> int:
    workbook_path = Path("/Users/danielsinukoff/Downloads/Partner Billing and Processes (3).xlsx")
    snapshot = load_snapshot()
    rows = build_receivable_rows(snapshot, workbook_path) + build_payout_rows(snapshot, workbook_path)
    out_dir = ROOT / "reports" / "billing_reconciliation"
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / "partner_billing_processes_3_reconciliation.csv"
    md_path = out_dir / "partner_billing_processes_3_reconciliation.md"
    xlsx_path = out_dir / "partner_billing_processes_3_delta_review.xlsx"
    write_csv(rows, csv_path)
    write_markdown(rows, md_path)
    write_xlsx(rows, xlsx_path)
    print(csv_path)
    print(md_path)
    print(xlsx_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
