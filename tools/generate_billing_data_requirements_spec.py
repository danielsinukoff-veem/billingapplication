#!/usr/bin/env python3

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/danielsinukoff/Documents/billing-workbook")
OUT_DIR = ROOT / "reports" / "data_requirements"
OUT_XLSX = OUT_DIR / "billing_data_requirements_spec.xlsx"
OUT_HANDOFF_XLSX = OUT_DIR / "billing_data_requirements_engineering_handoff.xlsx"


def row(
    category: str,
    field_name: str,
    required_for: str,
    source_report: str,
    required_level: str,
    notes: str,
    example: str = "",
) -> dict[str, str]:
    return {
        "Category": category,
        "Field Name": field_name,
        "Required For": required_for,
        "Source Report": source_report,
        "Required / Recommended": required_level,
        "Notes": notes,
        "Example": example,
    }


LOOKER_SYNC_FIELDS = [
    row("Identity", "partner", "All fee calculations", "All reports", "Required", "Canonical partner name used by the workbook. All alias fields must map back to this.", "Remittanceshub"),
    row("Identity", "Partner Name", "Partner mapping", "All Registered Accounts / any partner report", "Required if raw report uses it", "Raw partner name from the source report. Needed if the canonical name is not emitted directly.", "Multigate Network Pay Inc."),
    row("Identity", "Partner Group Source", "Partner mapping", "Partner Offline Billing / Partner Revenue Summary", "Required if present in source", "Used by existing importer to infer the canonical partner.", "Remittanceshub"),
    row("Identity", "Partner Group With Bank", "Partner mapping", "Partner Revenue Share / Partner Revenue Reversal", "Required if present in source", "Used by revenue-share imports to map rows to the workbook partner.", "Blindpay | Bank"),
    row("Identity", "Partner Offline Billing PARTNER", "Partner mapping", "Partner Offline Billing", "Required if present in source", "Alternative partner field in offline exports.", "Skydo"),
    row("Identity", "paymentId", "Transaction, reversal, audit, dedupe", "Partner Offline Billing / Reversals / Revenue Share / Stampli FX", "Required", "Primary payment-level identifier. Needed for dedupe, reversal matching, and detail evidence.", "PAY_12345"),
    row("Identity", "accountId", "VA fees, account tracking, detail evidence", "Partner Offline Billing / All Registered Accounts / Stampli feeds", "Required if account-based product exists", "Needed to connect transactions and account lifecycle behavior.", "ACC_001"),
    row("Identity", "typeDefn", "Business vs individual pricing", "All Registered Accounts / Offline Billing detail", "Required if contract charges differ by account type", "Must distinguish Business vs Individual accounts.", "Business"),
    row("Identity", "initiatorStatus", "Audit / troubleshooting", "Partner Offline Billing / Revenue Share", "Recommended", "Helps explain billing anomalies and status edge cases.", "Credit Complete"),
    row("Identity", "payerEmail", "Audit / partner support", "Partner Offline Billing / Reversals / Stampli FX", "Recommended", "Used for traceability and issue resolution.", "ops@example.com"),
    row("Identity", "payeeEmail", "Audit / partner support", "Partner Offline Billing / Reversals / Stampli FX", "Recommended", "Used for traceability and issue resolution.", "vendor@example.com"),
    row("Identity", "payerBusinessName", "Partner mapping and audit", "Partner Offline Billing / Reversals / Stampli FX", "Recommended", "Helpful for unmatched partner resolution.", "RemittancesHub LLC"),
    row("Identity", "payeeBusinessName", "Partner mapping and audit", "Partner Offline Billing / Reversals / Stampli FX", "Recommended", "Helpful for unmatched partner resolution.", "Supplier Ltd."),
    row("Identity", "payerName", "Audit / detail review", "Partner Revenue Share", "Recommended", "Human-readable name from rev-share detail.", "ACME Corp"),
    row("Identity", "payeeName", "Audit / detail review", "Partner Revenue Share", "Recommended", "Human-readable name from rev-share detail.", "John Smith"),

    row("Timing", "period", "All billing aggregation", "All reports", "Required", "Billing month bucket in YYYY-MM. Every imported record must resolve to one period.", "2026-04"),
    row("Timing", "submissionDateTime", "As-of analysis, evidence, settlement inference", "Partner Offline Billing / Reversals / Revenue Share / Stampli FX", "Required", "Payment submission timestamp. Needed for detailed audit and some fallback coverage logic.", "2026-04-07T13:24:58Z"),
    row("Timing", "creditCompleteDateTime", "Month bucketing and detail evidence", "Partner Offline Billing / Revenue Share / Stampli credit-complete / Stampli FX", "Required where available", "Most reliable event date for payment completion and period assignment.", "2026-04-07T15:10:02Z"),
    row("Timing", "creditCompleteMonth", "Month bucketing", "Partner Offline Billing", "Required if timestamp not available", "Used when the report emits only a month bucket instead of a timestamp.", "2026-04"),
    row("Timing", "timeCreatedDate", "Revenue-share bucketing fallback", "Partner Revenue Share", "Required if credit complete timestamp absent", "Used by the importer as a fallback period source.", "2026-04-03T11:00:00Z"),
    row("Timing", "paymentSubmissionDate", "Stampli FX reversals", "Stampli FX Reversal", "Required if separate from submissionDateTime", "Needed in reversal detail rows for audit.", "2026-04-03"),
    row("Timing", "creditInitiatedDateTime", "Stampli direct and FX detail", "Stampli Domestic Revenue / Stampli FX Share / Stampli FX Reversal", "Required for Stampli", "Used as the transactional event timestamp in Stampli feeds.", "2026-04-01T09:11:00Z"),
    row("Timing", "refundCompletedDateTime", "Reversal month bucketing", "Partner Revenue Reversal / Stampli FX Reversal", "Required for reversal-based fee logic", "Needed to bucket reversals into the correct invoice month.", "2026-04-08T12:00:00Z"),
    row("Timing", "refundCompletedMonth", "Reversal month bucketing", "Partner Revenue Reversal", "Required if timestamp not available", "Fallback month key for reversal summaries.", "2026-04"),
    row("Timing", "debitReversalDateTime", "Offline reversal detail", "Partner Offline Billing (Reversals)", "Required if available", "Exact reversal date for evidence and exact coverage.", "2026-04-02T10:02:00Z"),
    row("Timing", "debitReversalMonth", "Offline reversal month bucketing", "Partner Offline Billing (Reversals)", "Required if exact date absent", "Fallback month key for offline reversals.", "2026-04"),
    row("Timing", "joinDateTime", "VA opening, business setup, annual account fees", "All Registered Accounts", "Required if any VA or account-setup fee exists", "Account open timestamp.", "2026-01-12T00:00:00Z"),
    row("Timing", "closeDateTime", "Account closing fees", "All Registered Accounts", "Required if any closing fee exists", "Needed to assess one-time account closing charges in the correct month.", "2026-12-14T00:00:00Z"),
    row("Timing", "lastInboundTransactionDateTime", "Dormancy fees", "All Registered Accounts or dedicated account activity feed", "Required if dormancy fees exist", "Used to determine 90-day inactivity and ongoing dormancy billing.", "2026-01-01T15:00:00Z"),
    row("Timing", "settlementSweepDateTime", "Daily settlement fee", "Dedicated settlement feed or All Registered Accounts enrichment", "Required if daily settlement fee exists", "Need one event per wallet-to-bank settlement sweep; a simple monthly count is only acceptable if it is exact.", "2026-04-05T18:45:00Z"),

    row("Classification", "txnType", "Offline, FX, wire, local leg, rev-share, surcharge logic", "Partner Offline Billing / Revenue Share", "Required", "Canonical transaction family. Examples: Domestic, USD Abroad, FX, Payin, Payout.", "Domestic"),
    row("Classification", "paymentType", "Billing interpretation and evidence", "Partner Offline Billing / Reversals / Revenue Share / Stampli FX", "Required", "Raw payment-type label from source.", "Incoming US"),
    row("Classification", "txnTypeRaw", "Audit / troubleshooting", "Partner Revenue Share", "Recommended", "Raw txn-type value before normalization.", "Incoming US"),
    row("Classification", "speedFlag", "ACH vs Faster ACH vs RTP pricing", "Partner Offline Billing / Revenue Share", "Required", "Standard, FasterACH, RTP, etc.", "FasterACH"),
    row("Classification", "processingMethod", "ACH / Wire / RTP / EFT / Card routing", "Partner Offline Billing / Revenue Share", "Required", "Used directly by pricing row matching.", "Wire"),
    row("Classification", "creditRail", "Audit / processing interpretation", "Partner Offline Billing", "Recommended", "Helpful for rail-specific debugging.", "Wire"),
    row("Classification", "fundingMethodUsed", "Audit / source interpretation", "Partner Offline Billing", "Recommended", "Helpful when payerFunding/payeeFunding are incomplete.", "Bank"),
    row("Classification", "payerFunding", "Offline row matching, reversal fee matching", "Partner Offline Billing / Reversals / Revenue Share", "Required", "Bank, Wallet, Card, etc.", "Bank"),
    row("Classification", "payeeFunding", "Offline / push-to-debit / card payout logic", "Partner Offline Billing / Revenue Share", "Required if contract differentiates by destination rail", "Bank, Wallet, Card, etc.", "Card"),
    row("Classification", "payeeCardType", "Debit vs credit pricing", "Partner Offline Billing / Revenue Share", "Required if card pricing exists", "Needed for credit-card vs debit-card fee rows.", "Debit"),
    row("Classification", "isRTP", "RTP identification", "Partner Offline Billing / Revenue Share", "Required if RTP fees exist", "Boolean flag used to normalize RTP pricing.", "TRUE"),
    row("Classification", "isFasterAch", "Faster ACH identification", "Partner Offline Billing / Revenue Share", "Required if Faster ACH exists", "Used to distinguish same-day/expedited ACH from standard ACH.", "TRUE"),

    row("Geography & Currency", "payerCcy", "FX, local payment legs, country/currency matching", "Partner Offline Billing / Revenue Share / Stampli FX", "Required", "Originating currency.", "USD"),
    row("Geography & Currency", "payeeCcy", "FX, local payment legs, wire vs local payout pricing", "Partner Offline Billing / Revenue Share / Stampli FX", "Required", "Destination currency.", "EUR"),
    row("Geography & Currency", "payeeAmountCurrency", "FX validation and Stampli markup math", "Partner Offline Billing / Stampli FX", "Required for FX", "Often same as payeeCcy, but keep as explicit source field.", "EUR"),
    row("Geography & Currency", "payerCountry", "Local payment leg logic", "Partner Offline Billing / Revenue Share", "Required", "Needed to identify domestic vs cross-border and EEA/UK/CA/AU local legs.", "US"),
    row("Geography & Currency", "payeeCountry", "Local payment leg logic", "Partner Offline Billing / Reversals / Revenue Share / Stampli FX", "Required", "Needed to identify domestic vs cross-border and EEA/UK/CA/AU local legs.", "DE"),

    row("Amounts & Counts", "txnCount", "Per-txn fees, caps, summaries", "All transaction and summary reports", "Required", "Transaction count for grouped rows.", "64"),
    row("Amounts & Counts", "totalVolume", "Volume fees, FX fees, minimum revenue checks", "Partner Offline Billing / Revenue Share", "Required", "USD-equivalent total volume for the grouped row.", "535891.47"),
    row("Amounts & Counts", "avgTxnSize", "FX tier matching", "Derived or provided", "Recommended", "Can be derived from totalVolume / txnCount, but source-provided is fine.", "8373.30"),
    row("Amounts & Counts", "payeeAmount", "FX validation and local currency audit", "Partner Offline Billing / Stampli FX", "Required for FX", "Foreign-currency payee amount.", "1000"),
    row("Amounts & Counts", "usdAmountDebited", "FX markup detection and validation", "Partner Offline Billing / Stampli FX", "Required for FX", "Customer USD amount charged / debited.", "1015"),
    row("Amounts & Counts", "paymentUsdEquivalentAmount", "Volume pricing, FX markup math", "Partner Offline Billing / Stampli feeds", "Required", "USD-equivalent volume amount for the payment.", "1000"),
    row("Amounts & Counts", "customerRevenue", "Rev-share and imported direct billing", "Partner Revenue Share / direct billing capable feeds", "Required if source computes revenue directly", "Revenue already attributed to the transaction row.", "31.08"),
    row("Amounts & Counts", "estRevenue", "Double-charge prevention and monthly minimum logic", "Partner Offline Billing", "Required", "Revenue already collected at transaction time. Must be included in minimum tests but excluded from duplicate invoice lines.", "42.55"),
    row("Amounts & Counts", "netRevenue", "Revenue-share summary and reversals", "Partner Revenue Share / Revenue Summary / Revenue Reversal", "Required if reported", "Net revenue basis from revenue-share sources.", "12000.15"),
    row("Amounts & Counts", "countPricing", "Revenue-share detail support", "Partner Revenue Share", "Recommended", "Component pricing value from rev-share detail.", "1320"),
    row("Amounts & Counts", "directInvoiceAmount", "Authoritative imported charge line", "Stampli direct billing or any feed with explicit billed fee", "Required if source directly states the charge", "Lets the app use feed-provided charges instead of recomputing.", "1071.78"),
    row("Amounts & Counts", "directInvoiceRate", "Audit / per-txn explanation", "Derived or source-provided", "Recommended", "Helpful for transparency in imported direct-billing rows.", "1.50"),

    row("Reversals", "reversalCount", "Reversal fee billing", "Partner Offline Billing (Reversals)", "Required if reversal fees exist", "Count of reversal events for the grouped row.", "12"),
    row("Reversals", "paymentPriority", "Funding-type fallback for reversal fees", "Partner Offline Billing (Reversals)", "Recommended", "Useful when payerFunding is ambiguous in reversal feeds.", "Bank - ACH"),

    row("Revenue Summary", "revenueOwed", "Charges to partner", "Partner Revenue Summary / Partner Revenue Reversal", "Required if summary feed is used", "Amount partner owes Veem.", "7517.29"),
    row("Revenue Summary", "partnerRevenueShare", "Payouts to partner", "Partner Revenue Summary / Partner Revenue Reversal", "Required if revenue share exists", "Amount Veem owes partner.", "14985.00"),
    row("Revenue Summary", "monthlyMinimumRevenue", "Monthly minimum commitment calculation", "Partner Revenue Summary", "Required if partner has minimum", "Monthly minimum amount from the summary source.", "5000"),
    row("Revenue Summary", "billingType", "Interpret billing-summary rows", "Partner Revenue Summary", "Required if using billing-summary style exports", "Examples: MONTHLY_SUBSCRIPTION_FEE, TXN_COUNT, VOLUME, REVERSAL.", "MONTHLY_SUBSCRIPTION_FEE FOR FEB"),
    row("Revenue Summary", "summaryLabel", "Display / matching of summary lines", "Partner Revenue Summary", "Recommended", "Normalized summary label.", "MONTHLY_SUBSCRIPTION_FEE FOR FEB"),
    row("Revenue Summary", "summaryComputation", "Audit / math explanation", "Partner Revenue Summary", "Required if summary rows are used", "Text like '74.4 * 5000' or '132 * 10'.", "64 * 10"),
    row("Revenue Summary", "summaryCount", "Summary math", "Partner Revenue Summary", "Recommended", "Parsed count component from summaryComputation.", "64"),
    row("Revenue Summary", "summaryUnitAmount", "Summary math", "Partner Revenue Summary", "Recommended", "Parsed unit amount from summaryComputation.", "10"),
    row("Revenue Summary", "summaryLineAmount", "Summary math", "Partner Revenue Summary", "Required if summary rows are used", "Total line amount for the summary row.", "640"),
    row("Revenue Summary", "summaryDirection", "Know whether line is charge or payout", "Partner Revenue Summary", "Required if summary rows are used", "charge or pay", "charge"),

    row("Virtual Accounts", "newAccountsOpened", "Account opening fees", "All Registered Accounts", "Required if account opening fee exists", "Monthly count of newly opened accounts.", "35"),
    row("Virtual Accounts", "totalActiveAccounts", "Monthly active account fees", "All Registered Accounts", "Required if monthly active fee exists", "Total active/open accounts in the period.", "412"),
    row("Virtual Accounts", "totalBusinessAccounts", "Year-end per-business account setup fees", "All Registered Accounts", "Required if per-business setup fee exists", "Critical for contracts charging per active business account.", "120"),
    row("Virtual Accounts", "totalIndividualAccounts", "Year-end per-individual account setup fees", "All Registered Accounts", "Required if per-individual setup fee exists", "Critical for contracts charging per active individual account.", "950"),
    row("Virtual Accounts", "dormantAccounts", "Dormancy fees", "All Registered Accounts", "Required if dormancy fee exists", "Count of accounts with 90+ days without inbound activity.", "40"),
    row("Virtual Accounts", "closedAccounts", "Account closing fees", "All Registered Accounts", "Required if closing fee exists", "Count of accounts closed during the period.", "12"),
    row("Virtual Accounts", "newBusinessSetups", "Business-setup fees", "All Registered Accounts", "Required if one-time business setup fee exists", "Count of newly opened business accounts in the period.", "18"),
    row("Virtual Accounts", "settlementCount", "Daily settlement fee", "Settlement feed or All Registered Accounts enrichment", "Required if daily settlement fee exists", "Exact number of wallet-to-bank settlement sweeps in the period.", "27"),
    row("Virtual Accounts", "accountActiveStatus", "VA active / dormant / closed logic", "All Registered Accounts", "Recommended", "Explicit account status if source can emit it.", "Active"),

    row("Stampli FX", "openExchangeRateUsed", "Stampli markup audit", "Stampli FX Share / Stampli FX Reversal", "Required for Stampli FX", "Source exchange rate used on the transaction.", "1.0842"),
    row("Stampli FX", "customerMarkupPct", "Stampli markup validation", "Stampli FX Share / Stampli FX Reversal", "Required for Stampli FX", "Customer markup percent from feed.", "0.0125"),
    row("Stampli FX", "stampliBuyRatePct", "Stampli markup validation", "Stampli FX Share / Stampli FX Reversal", "Required for Stampli FX", "Buy-rate percentage used to derive partner payout.", "0.0035"),
    row("Stampli FX", "stampliMarkupPct", "Stampli markup validation", "Stampli FX Share / Stampli FX Reversal", "Required for Stampli FX", "Reported markup percent from source.", "0.0090"),
    row("Stampli FX", "stampliMarkupAmount", "Stampli FX payout", "Stampli FX Share / Stampli FX Reversal", "Required for Stampli FX", "Source markup amount before partner payout logic.", "45.22"),

    row("Audit & Coverage", "sourceReport", "Upload traceability", "All reports", "Recommended", "Human-readable report source name for debugging.", "Partner Offline Billing"),
    row("Audit & Coverage", "dashboardId", "Upload traceability", "Workflow metadata", "Recommended", "Useful for tracking which Looker tile/look produced the file.", "1009"),
    row("Audit & Coverage", "tileId", "Upload traceability", "Workflow metadata", "Recommended", "Useful for tracking which Looker dashboard tile produced the file.", "9356"),
    row("Audit & Coverage", "lookId", "Upload traceability", "Workflow metadata", "Recommended", "Needed when report is driven by a Look instead of a dashboard tile.", "6836"),
    row("Audit & Coverage", "fetchedAt", "Freshness and troubleshooting", "Workflow metadata", "Recommended", "When the export was fetched from Looker.", "2026-04-08T22:48:00Z"),
    row("Audit & Coverage", "savedAt", "Freshness and troubleshooting", "Import audit", "Recommended", "When the workbook saved the upload.", "2026-04-08T22:48:21Z"),
    row("Audit & Coverage", "filterRangeStart", "Coverage reporting", "Workflow metadata / source report", "Recommended", "Start of requested query window.", "2026-04-01"),
    row("Audit & Coverage", "filterRangeEnd", "Coverage reporting", "Workflow metadata / source report", "Recommended", "End of requested query window.", "2026-04-30"),
    row("Audit & Coverage", "currentThroughDate", "Coverage reporting", "Derived from row-level dates", "Required if you want exact 'data through' reporting", "Best actual covered date from imported rows, not only the filter month.", "2026-04-08"),
    row("Audit & Coverage", "byteCount", "Debugging empty/partial files", "Workflow metadata", "Recommended", "Helps detect empty or malformed downloads.", "1048576"),
    row("Audit & Coverage", "warningCount", "Workflow health", "Import audit", "Recommended", "Count or flag for warnings produced during import.", "1"),
]


NON_LOOKER_FIELDS = [
    row("Contract / Billing Config", "effectiveDate", "Contract versioning and one-time fee timing", "Contract parse / partner config", "Required", "Contract effective date from the signed agreement.", "2026-01-12"),
    row("Contract / Billing Config", "contractStartDate", "Invoice schedule, implementation billing timing", "Partner Billing config", "Required", "Start date used for recurring fee schedules and implementation timing.", "2026-01-12"),
    row("Contract / Billing Config", "goLiveDate", "Recurring billing activation", "Partner Billing config", "Required", "Go-live date gates recurring billing for not-yet-live partners.", "2026-03-15"),
    row("Contract / Billing Config", "notYetLive", "Recurring billing gating", "Partner Billing config", "Required", "If true, only implementation-style charges should bill until go live.", "TRUE"),
    row("Contract / Billing Config", "billingFreq", "Invoice schedule", "Partner Billing config", "Required", "Monthly / Quarterly / Annual / Custom.", "Monthly"),
    row("Contract / Billing Config", "payBy", "Due-date logic", "Partner Billing config", "Required", "Contract terms like 'Due in 7 days' or 'Monthly settlement / setoff'.", "Due in 7 days"),
    row("Contract / Billing Config", "dueDays", "Automation and expected payment date", "Partner Billing config", "Required", "Numeric due-days value derived from payBy.", "7"),
    row("Contract / Billing Config", "billingDay", "Expected send date", "Partner Billing config", "Required for automated invoicing", "Day of month invoice is expected to go out.", "10"),
    row("Contract / Billing Config", "preferredBillingTiming", "Fallback invoice schedule inference", "Partner Billing config", "Recommended", "Human-readable billing timing used when exact billing day is not set.", "1st week of following month"),
    row("Contract / Billing Config", "contactEmails", "Invoice send / reminders", "Partner Billing config", "Required for automation", "Partner billing contacts.", "billing@example.com, finance@example.com"),
    row("Contract / Billing Config", "integrationStatus", "Lifecycle visibility", "Partner Billing config", "Recommended", "Operational status like onboarding / live.", "Integration Underway (Partners Onboarding)"),
    row("Contract / Billing Config", "lateFeePercentMonthly", "Late fee notices and accrual", "Partner Billing config", "Required if contract has late fee", "Monthly late fee rate.", "1.5"),
    row("Contract / Billing Config", "lateFeeStartDays", "Late fee timing", "Partner Billing config", "Required if contract has late fee", "Days after due date when late fees apply.", "30"),
    row("Contract / Billing Config", "serviceSuspensionDays", "Collections / service suspension alerting", "Partner Billing config", "Required if contract has suspension clause", "Days after due date when service suspension becomes eligible.", "30"),
    row("Contract / Billing Config", "lateFeeTerms", "Collections communication", "Partner Billing config", "Recommended", "Text summary of contractual late fee terms.", "1.5% monthly interest begins after 30 days overdue"),
]


FEE_REQUIREMENT_MAP = [
    {
        "Fee Type": "Implementation Fee",
        "Needs These Fields": "partner, effectiveDate OR contractStartDate",
        "Primary Source Reports": "Contract / Partner Billing config",
        "Notes": "One-time at signing/effective date. No Looker activity required.",
    },
    {
        "Fee Type": "Offline Fixed Transaction Fees",
        "Needs These Fields": "partner, paymentId, period, txnType, speedFlag, processingMethod, payerFunding, payeeFunding, payerCcy, payeeCcy, payerCountry, payeeCountry, txnCount, totalVolume",
        "Primary Source Reports": "Partner Offline Billing",
        "Notes": "Per-transaction fees for ACH, Faster ACH, wires, local payment legs, etc.",
    },
    {
        "Fee Type": "Volume % Fees",
        "Needs These Fields": "partner, period, txnType, speedFlag, processingMethod, payerFunding, payeeFunding, payeeCardType, payerCcy, payeeCcy, payerCountry, payeeCountry, txnCount, totalVolume, avgTxnSize",
        "Primary Source Reports": "Partner Offline Billing / Revenue Share detail",
        "Notes": "Used for RTP, card, FX, and tiered volume pricing.",
    },
    {
        "Fee Type": "RTP Cap Logic",
        "Needs These Fields": "partner, period, txnType, speedFlag, processingMethod, txnCount, totalVolume, avgTxnSize",
        "Primary Source Reports": "Partner Offline Billing",
        "Notes": "RTP % fee is applied to volume and then capped per transaction or per fee-cap rule.",
    },
    {
        "Fee Type": "Local Payment Legs (CAD / GBP / EUR / AUD / EEA)",
        "Needs These Fields": "partner, paymentId, payerCcy, payeeCcy, payerCountry, payeeCountry, processingMethod, speedFlag, txnCount",
        "Primary Source Reports": "Partner Offline Billing",
        "Notes": "Must distinguish same-country domestic legs from general cross-border flows.",
    },
    {
        "Fee Type": "SWIFT / USD Abroad Wire Fees",
        "Needs These Fields": "partner, paymentId, txnType, processingMethod, payerCcy, payeeCcy, payeeCountry, txnCount",
        "Primary Source Reports": "Partner Offline Billing",
        "Notes": "USD abroad / wire fees typically apply when processingMethod is Wire.",
    },
    {
        "Fee Type": "FX Conversion Fees",
        "Needs These Fields": "partner, paymentId, payerCcy, payeeCcy, payerCountry, payeeCountry, totalVolume, avgTxnSize, txnCount",
        "Primary Source Reports": "Partner Offline Billing / Stampli FX",
        "Notes": "Needs currency pair and volume for majors/minors/tertiary tiering.",
    },
    {
        "Fee Type": "Double-Charge Protection via Est Revenue",
        "Needs These Fields": "partner, paymentId, period, txnType, processingMethod, totalVolume, estRevenue",
        "Primary Source Reports": "Partner Offline Billing",
        "Notes": "If estRevenue already contains the transaction-time fee, the invoice line should be reduced or suppressed, but minimum revenue still includes it.",
    },
    {
        "Fee Type": "Minimum Monthly Revenue Commitment",
        "Needs These Fields": "partner, period, revenueOwed, partnerRevenueShare, monthlyMinimumRevenue, netRevenue, estRevenue",
        "Primary Source Reports": "Partner Revenue Summary + Partner Offline Billing",
        "Notes": "Minimum is the delta between actual earned revenue and the contractual monthly target.",
    },
    {
        "Fee Type": "Platform / Subscription Fee",
        "Needs These Fields": "partner, period",
        "Primary Source Reports": "Contract / Partner Billing config or Partner Revenue Summary if summary-driven",
        "Notes": "No transaction data required unless a summary feed overrides it.",
    },
    {
        "Fee Type": "Reversal Fees",
        "Needs These Fields": "partner, paymentId, period, payerFunding, reversalCount, reversalDateTime",
        "Primary Source Reports": "Partner Offline Billing (Reversals)",
        "Notes": "Requires reversal counts and funding type matching.",
    },
    {
        "Fee Type": "Revenue Share Payout / Reversal",
        "Needs These Fields": "partner, period, netRevenue, partnerRevenueShare, revenueOwed, billingType, summaryLineAmount, refundCompletedDateTime",
        "Primary Source Reports": "Partner Revenue Summary / Partner Revenue Share / Partner Revenue Reversal",
        "Notes": "Needed for partner payout invoices and net settlement logic.",
    },
    {
        "Fee Type": "Stampli FX Payout / Reversal",
        "Needs These Fields": "partner, paymentId, period, payeeAmount, payeeAmountCurrency, usdAmountDebited, paymentUsdEquivalentAmount, customerMarkupPct, stampliBuyRatePct, stampliMarkupPct, stampliMarkupAmount, refundCompletedDateTime",
        "Primary Source Reports": "Stampli FX Revenue Share / Stampli FX Revenue Reversal / All Stampli Credit Complete",
        "Notes": "Needs exact payout and reversal detail plus month bucketing.",
    },
    {
        "Fee Type": "VA Account Opening",
        "Needs These Fields": "partner, period, accountId, joinDateTime, newAccountsOpened",
        "Primary Source Reports": "All Registered Accounts",
        "Notes": "One-time per newly opened account.",
    },
    {
        "Fee Type": "VA Monthly Active",
        "Needs These Fields": "partner, period, totalActiveAccounts",
        "Primary Source Reports": "All Registered Accounts",
        "Notes": "Monthly charge per active/open account.",
    },
    {
        "Fee Type": "VA Dormancy",
        "Needs These Fields": "partner, period, dormantAccounts, lastInboundTransactionDateTime",
        "Primary Source Reports": "All Registered Accounts or dedicated account activity feed",
        "Notes": "Monthly fee after 90 days without inbound activity.",
    },
    {
        "Fee Type": "VA Account Closing",
        "Needs These Fields": "partner, period, closedAccounts, closeDateTime",
        "Primary Source Reports": "All Registered Accounts",
        "Notes": "One-time closing fee on qualifying account closures.",
    },
    {
        "Fee Type": "Account Setup Per Business",
        "Needs These Fields": "partner, period, totalBusinessAccounts",
        "Primary Source Reports": "All Registered Accounts",
        "Notes": "Year-end fee on active business accounts.",
    },
    {
        "Fee Type": "Account Setup Per Individual",
        "Needs These Fields": "partner, period, totalIndividualAccounts",
        "Primary Source Reports": "All Registered Accounts",
        "Notes": "Year-end fee on active individual accounts.",
    },
    {
        "Fee Type": "Daily Settlement",
        "Needs These Fields": "partner, period, settlementCount OR settlementSweepDateTime",
        "Primary Source Reports": "Dedicated settlement feed or All Registered Accounts enrichment",
        "Notes": "Charge once per wallet-to-bank settlement sweep.",
    },
    {
        "Fee Type": "Other / Misc Fees (1099, 1042S, NSF, etc.)",
        "Needs These Fields": "partner, period, event count for each applicable fee type",
        "Primary Source Reports": "Dedicated operational report(s)",
        "Notes": "These are parsed from contracts but currently need explicit event/count feeds if they will be invoiced automatically.",
    },
]


CRITICAL_FIELD_NAMES = {
    "partner",
    "paymentId",
    "accountId",
    "period",
    "txnType",
    "processingMethod",
    "payerFunding",
    "payeeFunding",
    "payerCcy",
    "payeeCcy",
    "payerCountry",
    "payeeCountry",
    "submissionDateTime",
    "creditCompleteDateTime",
    "refundCompletedDateTime",
    "paymentUsdEquivalentAmount",
    "txnCount",
    "totalVolume",
    "estRevenue",
    "revenueOwed",
    "monthlyMinimumRevenue",
    "totalBusinessAccounts",
    "totalIndividualAccounts",
    "dormantAccounts",
    "closedAccounts",
    "settlementCount",
}


def write_sheet(ws, rows: list[dict[str, str]], tab_color: str | None = None) -> None:
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="EBD9B9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for record in rows:
        ws.append([record.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, _header in enumerate(headers, start=1):
        column_letter = get_column_letter(idx)
        max_len = max(len(str(ws[f"{column_letter}{row}"].value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 14), 48)


def build_summary_rows() -> list[dict[str, str]]:
    return [
        {
            "Section": "Looker Sync Fields",
            "Count": str(len(LOOKER_SYNC_FIELDS)),
            "Purpose": "All raw and derived fields that should exist in Looker exports or workflow metadata to support invoice calculation, auditability, and freshness reporting.",
        },
        {
            "Section": "Non-Looker App Fields",
            "Count": str(len(NON_LOOKER_FIELDS)),
            "Purpose": "Partner billing/admin metadata that still must exist in the app even though it does not come from Looker.",
        },
        {
            "Section": "Fee Requirement Map",
            "Count": str(len(FEE_REQUIREMENT_MAP)),
            "Purpose": "Maps each fee family to the exact fields and source reports required to calculate it safely.",
        },
        {
            "Section": "Critical Must-Haves",
            "Count": str(sum(1 for item in LOOKER_SYNC_FIELDS if item["Field Name"] in CRITICAL_FIELD_NAMES)),
            "Purpose": "Smallest set of fields that cannot be missing without risking underbilling, double billing, or broken minimum-revenue logic.",
        },
    ]


def to_snake_case(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    chars: list[str] = []
    prev_lower_or_digit = False
    for char in text:
        if char.isupper() and prev_lower_or_digit:
            chars.append("_")
        if char.isalnum():
            chars.append(char.lower())
            prev_lower_or_digit = char.islower() or char.isdigit()
        else:
            if chars and chars[-1] != "_":
                chars.append("_")
            prev_lower_or_digit = False
    slug = "".join(chars).strip("_")
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug


def build_engineering_handoff_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record in LOOKER_SYNC_FIELDS:
        field_name = record["Field Name"]
        rows.append(
            {
                "Section": "Looker Sync",
                "Category": record["Category"],
                "Canonical App Field": field_name,
                "Looker Field Name to Implement": to_snake_case(field_name),
                "Source Report": record["Source Report"],
                "Required For": record["Required For"],
                "Required / Recommended": record["Required / Recommended"],
                "Looker Needed?": "Yes",
                "Notes": record["Notes"],
                "Example": record["Example"],
            }
        )
    for record in NON_LOOKER_FIELDS:
        field_name = record["Field Name"]
        rows.append(
            {
                "Section": "App Config",
                "Category": record["Category"],
                "Canonical App Field": field_name,
                "Looker Field Name to Implement": "N/A - app config",
                "Source Report": record["Source Report"],
                "Required For": record["Required For"],
                "Required / Recommended": record["Required / Recommended"],
                "Looker Needed?": "No",
                "Notes": record["Notes"],
                "Example": record["Example"],
            }
        )
    return rows


def build_handoff_summary_rows() -> list[dict[str, str]]:
    return [
        {
            "Section": "Looker Sync Fields",
            "Count": str(len(LOOKER_SYNC_FIELDS)),
            "Purpose": "Fields engineering should expose in Looker exports or workflow metadata.",
        },
        {
            "Section": "App Config Fields",
            "Count": str(len(NON_LOOKER_FIELDS)),
            "Purpose": "Fields that still live in the billing app and do not need to come from Looker.",
        },
        {
            "Section": "Fee Mapping Rows",
            "Count": str(len(FEE_REQUIREMENT_MAP)),
            "Purpose": "Reference map of which fee families depend on which fields.",
        },
    ]


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    write_sheet(summary_ws, build_summary_rows(), tab_color="0F5A52")

    looker_ws = wb.create_sheet("Looker Sync Fields")
    write_sheet(looker_ws, LOOKER_SYNC_FIELDS, tab_color="1C7C7C")

    app_ws = wb.create_sheet("App Non-Looker Fields")
    write_sheet(app_ws, NON_LOOKER_FIELDS, tab_color="5A7D9A")

    fee_ws = wb.create_sheet("Fee Requirement Map")
    write_sheet(fee_ws, FEE_REQUIREMENT_MAP, tab_color="A36F2E")

    critical_rows = [record for record in LOOKER_SYNC_FIELDS if record["Field Name"] in CRITICAL_FIELD_NAMES]
    critical_ws = wb.create_sheet("Critical Must Haves")
    write_sheet(critical_ws, critical_rows, tab_color="B74134")

    summary_ws["A7"] = "Notes"
    summary_ws["A7"].font = Font(bold=True)
    summary_ws["A8"] = "1. The strongest implementation is payment-level detail plus account-level lifecycle data, not only monthly partner totals."
    summary_ws["A9"] = "2. Est Revenue is required to prevent double charging while still supporting monthly minimum calculations."
    summary_ws["A10"] = "3. Account setup, dormancy, and daily settlement fees cannot be trusted unless business/individual/closed account counts and exact settlement events are available."
    summary_ws["A11"] = "4. Revenue Summary and Stampli FX feeds are required for minimum revenue, payout, and markup-reversal logic."
    summary_ws.column_dimensions["A"].width = 120

    wb.save(OUT_XLSX)

    handoff_wb = Workbook()
    handoff_summary_ws = handoff_wb.active
    handoff_summary_ws.title = "Summary"
    write_sheet(handoff_summary_ws, build_handoff_summary_rows(), tab_color="0F5A52")

    handoff_fields_ws = handoff_wb.create_sheet("Field Handoff")
    write_sheet(handoff_fields_ws, build_engineering_handoff_rows(), tab_color="1C7C7C")

    handoff_fee_ws = handoff_wb.create_sheet("Fee Mapping")
    write_sheet(handoff_fee_ws, FEE_REQUIREMENT_MAP, tab_color="A36F2E")

    handoff_summary_ws["A6"] = "Notes"
    handoff_summary_ws["A6"].font = Font(bold=True)
    handoff_summary_ws["A7"] = "1. 'Looker Field Name to Implement' is a suggested standardized export field name for engineering."
    handoff_summary_ws["A8"] = "2. 'Canonical App Field' is the field name the billing workbook logic expects conceptually."
    handoff_summary_ws["A9"] = "3. Rows marked 'N/A - app config' should be maintained in the app, not in Looker."
    handoff_summary_ws.column_dimensions["A"].width = 120

    handoff_wb.save(OUT_HANDOFF_XLSX)
    print(OUT_XLSX)
    print(OUT_HANDOFF_XLSX)


if __name__ == "__main__":
    main()
