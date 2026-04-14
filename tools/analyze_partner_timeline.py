from __future__ import annotations

import csv
import json
import re
import sqlite3
from calendar import monthrange
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path("/Users/danielsinukoff/Documents/billing-workbook")
DB_PATH = ROOT / "server" / "data" / "shared_workspace.db"
LIFECYCLE_PATH = ROOT / "reports" / "contract_audit" / "partner_lifecycle_dates_report.json"
HUBSPOT_PATH = ROOT / "reports" / "contract_audit" / "partner_hubspot_status_report.json"
OUTPUT_DIR = ROOT / "reports" / "investor_timeline"
CSV_PATH = OUTPUT_DIR / "partner_timeline_analysis.csv"
MD_PATH = OUTPUT_DIR / "partner_timeline_analysis.md"
XLSX_PATH = OUTPUT_DIR / "partner_timeline_analysis.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="E8F2FF")
GOOD_FILL = PatternFill("solid", fgColor="DCFCE7")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
BAD_FILL = PatternFill("solid", fgColor="FEE2E2")
MUTED_FILL = PatternFill("solid", fgColor="F3F4F6")


def normalize_partner(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def parse_date(value: str) -> date | None:
    return datetime.strptime(value, "%Y-%m-%d").date() if value else None


def month_key(value: str) -> str:
    return value[:7] if value else ""


def add_months(value: date, months: int) -> date:
    year = value.year + (value.month - 1 + months) // 12
    month = (value.month - 1 + months) % 12 + 1
    return date(year, month, min(value.day, monthrange(year, month)[1]))


def month_diff(start_period: str, end_period: str) -> int:
    start_year, start_month = map(int, start_period.split("-"))
    end_year, end_month = map(int, end_period.split("-"))
    return (end_year - start_year) * 12 + (end_month - start_month)


def money(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


@dataclass
class PartnerTimeline:
    partner: str
    contract_sign_date: str
    go_live_date: str
    integration_status: str
    not_yet_live: bool
    onboarding_days: int | None
    onboarding_months: float | None
    met_3mo_onboarding: bool | None
    floor_model: str
    floor_model_label: str
    first_post_go_live_period: str
    floor_reference: str
    first_ramp_period: str
    ramp_months_from_live: int | None
    approx_sign_to_ramp_months: float | None
    met_6mo_ramp: bool | None
    overall_followed_timeline: bool | None
    current_assessment: str
    explanation: str
    monthly_trace: str


def load_snapshot() -> dict[str, Any]:
    with sqlite3.connect(DB_PATH) as conn:
        row = conn.execute("select snapshot_json from workbook_snapshots order by id desc limit 1").fetchone()
    if not row:
        raise RuntimeError("No workbook snapshot found.")
    return json.loads(row[0])


def load_json(path: Path) -> Any:
    return json.loads(path.read_text())


def is_explicit_live(config: dict[str, Any], hubspot_row: dict[str, Any]) -> bool:
    if config.get("goLiveDate") or hubspot_row.get("goLiveDate"):
        return True
    status = (config.get("integrationStatus") or hubspot_row.get("integrationStatus") or "").lower()
    return status.startswith("live")


def applicable_platform(platform_rows: list[dict[str, Any]], period: str) -> float:
    for row in platform_rows:
        start = month_key(row.get("startDate") or "") or "0000-00"
        end = month_key(row.get("endDate") or "") or "9999-99"
        if start <= period <= end:
            return money(row.get("monthlyFee"))
    return 0.0


def applicable_minimum(
    minimum_rows: list[dict[str, Any]],
    volume: float,
    period: str,
) -> float:
    candidates: list[tuple[float, float]] = []
    for row in minimum_rows:
        start = month_key(row.get("startDate") or "") or "0000-00"
        end = month_key(row.get("endDate") or "") or "9999-99"
        min_vol = float(row.get("minVol") or 0)
        max_vol = float(row.get("maxVol") or 10**18)
        if start <= period <= end and min_vol <= volume <= max_vol:
            candidates.append((min_vol, money(row.get("minAmount"))))
    if not candidates:
        return 0.0
    candidates.sort()
    return candidates[-1][1]


def floor_model_label(model: str) -> str:
    return {
        "minimum_floor": "True minimum / floor",
        "fixed_subscription_floor": "Fixed subscription / platform proxy",
        "none": "No comparable floor",
    }.get(model, model)


def assessment_label(onboarding_met: bool | None, ramp_met: bool | None, ramp_status: str) -> str:
    if ramp_status == "not_live":
        return "Not live yet"
    if ramp_status == "no_post_go_live_revenue":
        return "Live, but no post-go-live revenue loaded"
    if ramp_status == "no_floor_in_contract":
        return "Live, but no comparable floor in contract"
    if ramp_status == "in_progress":
        return "In progress"
    if onboarding_met is True and ramp_met is True:
        return "Followed assumption"
    if onboarding_met is False and ramp_met is True:
        return "Missed onboarding, met ramp"
    if onboarding_met is True and ramp_met is False:
        return "Met onboarding, missed ramp"
    if onboarding_met is False and ramp_met is False:
        return "Missed onboarding and ramp"
    return "Needs review"


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells[: min(len(column_cells), 150)]) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 14), 50)
    sheet.freeze_panes = "A2"


def build_timelines() -> tuple[list[PartnerTimeline], dict[str, Any]]:
    snapshot = load_snapshot()
    lifecycle_rows = load_json(LIFECYCLE_PATH)["rows"]
    hubspot_rows = load_json(HUBSPOT_PATH)["updatedPartners"]

    lifecycle_by_partner = {normalize_partner(row["partner"]): row for row in lifecycle_rows}
    hubspot_by_partner = {normalize_partner(row["partner"]): row for row in hubspot_rows}
    billing_by_partner = {normalize_partner(row["partner"]): row for row in snapshot["pBilling"] if row.get("partner")}

    minimum_by_partner: dict[str, list[dict[str, Any]]] = defaultdict(list)
    platform_by_partner: dict[str, list[dict[str, Any]]] = defaultdict(list)
    lrs_by_partner: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    ltxn_by_partner: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))

    for row in snapshot["mins"]:
        if row.get("partner"):
            minimum_by_partner[normalize_partner(row["partner"])].append(row)
    for row in snapshot["plat"]:
        if row.get("partner"):
            platform_by_partner[normalize_partner(row["partner"])].append(row)
    for row in snapshot["lrs"]:
        if row.get("partner") and row.get("period") and (row.get("summaryDirection") or "charge") == "charge":
            lrs_by_partner[normalize_partner(row["partner"])][row["period"]].append(row)
    for row in snapshot["ltxn"]:
        if row.get("partner") and row.get("period"):
            ltxn_by_partner[normalize_partner(row["partner"])][row["period"]].append(row)

    latest_period = max(period for partner_rows in lrs_by_partner.values() for period in partner_rows.keys())
    timelines: list[PartnerTimeline] = []

    for partner_key, config in billing_by_partner.items():
        partner = config["partner"]
        lifecycle = lifecycle_by_partner.get(partner_key, {})
        hubspot = hubspot_by_partner.get(partner_key, {})

        contract_sign_date = lifecycle.get("contractStartDate") or config.get("contractStartDate") or ""
        explicit_go_live = config.get("goLiveDate") or hubspot.get("goLiveDate") or ""
        not_yet_live = (
            bool(config.get("notYetLive"))
            if config.get("notYetLive") is not None
            else bool(hubspot.get("notYetLive"))
        )
        integration_status = config.get("integrationStatus") or hubspot.get("integrationStatus") or ""
        lifecycle_go_live = lifecycle.get("goLiveDate") or ""
        go_live_date = explicit_go_live or (
            lifecycle_go_live if (not not_yet_live and is_explicit_live(config, hubspot)) else ""
        )

        sign_date = parse_date(contract_sign_date)
        live_date = parse_date(go_live_date)
        periods = sorted(lrs_by_partner.get(partner_key, {}))
        post_periods = [period for period in periods if go_live_date and period >= month_key(go_live_date)]

        has_minimum_rows = any(
            sum(
                money(line.get("revenueOwed"))
                for line in lrs_by_partner[partner_key][period]
                if (line.get("summaryBillingType") or "").upper() == "MINIMUM_MONTHLY_REVENUE"
            ) > 0.01
            for period in post_periods
        )
        has_applicable_minimum = any(
            applicable_minimum(
                minimum_by_partner.get(partner_key, []),
                sum(money(row.get("totalVolume")) for row in ltxn_by_partner.get(partner_key, {}).get(period, [])),
                period,
            )
            > 0
            for period in post_periods
        )
        has_platform = any(
            applicable_platform(platform_by_partner.get(partner_key, []), period) > 0 for period in post_periods
        )
        has_subscription_lines = any(
            any(
                (line.get("summaryBillingType") or "").upper().startswith("MONTHLY_SUBSCRIPTION_FEE")
                for line in lrs_by_partner[partner_key][period]
            )
            for period in post_periods
        )

        if has_minimum_rows or has_applicable_minimum:
            floor_model = "minimum_floor"
        elif has_platform or has_subscription_lines:
            floor_model = "fixed_subscription_floor"
        else:
            floor_model = "none"

        onboarding_days = (live_date - sign_date).days if sign_date and live_date else None
        onboarding_months = round(onboarding_days / 30.44, 2) if onboarding_days is not None else None
        met_3mo_onboarding = live_date <= add_months(sign_date, 3) if sign_date and live_date else None

        first_ramp_period = ""
        ramp_months_from_live: int | None = None
        met_6mo_ramp: bool | None = None
        floor_reference = ""
        explanation = ""
        monthly_trace_parts: list[str] = []

        if not sign_date:
            current_assessment = "Missing contract date"
            explanation = "No contract sign date was available, so the onboarding and ramp timeline could not be measured."
        elif not live_date:
            current_assessment = "Not live yet"
            explanation = (
                f"No confirmed go-live date is loaded. Current status: {integration_status or 'Unknown'}."
            )
        elif not post_periods:
            current_assessment = "No post-go-live revenue"
            explanation = "A go-live date exists, but there are no post-go-live revenue rows in the workbook yet."
        elif floor_model == "none":
            current_assessment = "No comparable floor in contract"
            explanation = (
                "The partner is live, but the current contract data does not include a comparable monthly minimum or "
                "fixed subscription/platform floor, so the 6-month ramp assumption is not directly testable."
            )
        else:
            ramp_status = "in_progress"
            for period in post_periods:
                lines = lrs_by_partner[partner_key][period]
                total = round(sum(money(line.get("revenueOwed")) for line in lines), 2)
                if floor_model == "minimum_floor":
                    floor = applicable_minimum(
                        minimum_by_partner.get(partner_key, []),
                        sum(money(row.get("totalVolume")) for row in ltxn_by_partner.get(partner_key, {}).get(period, [])),
                        period,
                    )
                    minimum_component = round(
                        sum(
                            money(line.get("revenueOwed"))
                            for line in lines
                            if (line.get("summaryBillingType") or "").upper() == "MINIMUM_MONTHLY_REVENUE"
                        ),
                        2,
                    )
                    activity_revenue = round(max(0.0, total - minimum_component), 2)
                    floor_reference = floor_reference or (f"{floor:,.2f}" if floor else "")
                    monthly_trace_parts.append(
                        f"{period}: activity {activity_revenue:,.2f} vs floor {floor:,.2f}"
                    )
                    if floor and activity_revenue >= floor - 1:
                        first_ramp_period = period
                        ramp_months_from_live = month_diff(month_key(go_live_date), period)
                        met_6mo_ramp = ramp_months_from_live <= 6
                        ramp_status = "ramped"
                        explanation = (
                            f"True minimum-floor partner. In {period}, activity revenue reached {activity_revenue:,.2f} "
                            f"against a contractual floor of {floor:,.2f}, so the partner exited the minimum "
                            f"{ramp_months_from_live} month(s) after go-live."
                        )
                        break
                else:
                    floor = applicable_platform(platform_by_partner.get(partner_key, []), period)
                    threshold = round(floor * 1.05, 2)
                    floor_reference = floor_reference or (f"{floor:,.2f}" if floor else "")
                    monthly_trace_parts.append(
                        f"{period}: recurring {total:,.2f} vs proxy floor {floor:,.2f} (5% threshold {threshold:,.2f})"
                    )
                    if floor and total >= threshold - 0.01:
                        first_ramp_period = period
                        ramp_months_from_live = month_diff(month_key(go_live_date), period)
                        met_6mo_ramp = ramp_months_from_live <= 6
                        ramp_status = "ramped_proxy"
                        explanation = (
                            f"Fixed subscription/platform partner. Using a proxy ramp threshold of 105% of the fixed "
                            f"monthly floor, recurring revenue reached {total:,.2f} against a {floor:,.2f} floor in "
                            f"{period}, which is {ramp_months_from_live} month(s) after go-live."
                        )
                        break

            if not explanation:
                elapsed_months = month_diff(month_key(go_live_date), latest_period)
                ramp_months_from_live = elapsed_months
                if elapsed_months >= 6:
                    met_6mo_ramp = False
                    ramp_status = "missed"
                    if floor_model == "minimum_floor":
                        explanation = (
                            f"True minimum-floor partner. The available post-go-live months never show activity revenue "
                            f"reaching the contractual floor within the first 6 months; {elapsed_months} months of "
                            f"history are loaded."
                        )
                    else:
                        explanation = (
                            f"Fixed subscription/platform partner. Using the 105% proxy threshold, recurring revenue "
                            f"has not cleared the floor within the first 6 months; {elapsed_months} months of history "
                            f"are loaded."
                        )
                else:
                    explanation = (
                        f"Only {elapsed_months} post-go-live month(s) of revenue history are loaded, so the 6-month "
                        f"ramp assumption is still in progress."
                    )

            current_assessment = assessment_label(met_3mo_onboarding, met_6mo_ramp, ramp_status)

        overall_followed_timeline = (
            bool(met_3mo_onboarding and met_6mo_ramp)
            if met_3mo_onboarding is not None and met_6mo_ramp is not None
            else None
        )
        approx_sign_to_ramp_months = (
            round((onboarding_days / 30.44) + ramp_months_from_live, 2)
            if onboarding_days is not None and ramp_months_from_live is not None and first_ramp_period
            else None
        )

        timelines.append(
            PartnerTimeline(
                partner=partner,
                contract_sign_date=contract_sign_date,
                go_live_date=go_live_date,
                integration_status=integration_status,
                not_yet_live=not_yet_live,
                onboarding_days=onboarding_days,
                onboarding_months=onboarding_months,
                met_3mo_onboarding=met_3mo_onboarding,
                floor_model=floor_model,
                floor_model_label=floor_model_label(floor_model),
                first_post_go_live_period=post_periods[0] if post_periods else "",
                floor_reference=floor_reference,
                first_ramp_period=first_ramp_period,
                ramp_months_from_live=ramp_months_from_live,
                approx_sign_to_ramp_months=approx_sign_to_ramp_months,
                met_6mo_ramp=met_6mo_ramp,
                overall_followed_timeline=overall_followed_timeline,
                current_assessment=current_assessment,
                explanation=explanation,
                monthly_trace=" | ".join(monthly_trace_parts[:12]),
            )
        )

    summary = {
        "latest_period": latest_period,
        "live_rows": [row for row in timelines if row.onboarding_days is not None],
        "minimum_floor_rows": [row for row in timelines if row.go_live_date and row.floor_model == "minimum_floor"],
        "fixed_fee_rows": [row for row in timelines if row.go_live_date and row.floor_model == "fixed_subscription_floor"],
        "no_floor_rows": [row for row in timelines if row.go_live_date and row.floor_model == "none"],
        "not_live_rows": [row for row in timelines if not row.go_live_date],
    }
    return timelines, summary


def write_csv(rows: list[PartnerTimeline]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    headers = [
        "Partner",
        "Contract Sign Date",
        "Go Live Date",
        "Integration Status",
        "Not Yet Live",
        "Onboarding Days",
        "Onboarding Months",
        "Met 3-Month Onboarding",
        "Floor Model",
        "First Post-Go-Live Period",
        "Floor Reference",
        "First Ramp Period",
        "Ramp Months From Live",
        "Approx Total Months Sign To Ramp",
        "Met 6-Month Ramp",
        "Overall Followed Timeline",
        "Current Assessment",
        "Explanation",
        "Monthly Trace",
    ]
    with CSV_PATH.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(
                [
                    row.partner,
                    row.contract_sign_date,
                    row.go_live_date,
                    row.integration_status,
                    row.not_yet_live,
                    row.onboarding_days,
                    row.onboarding_months,
                    row.met_3mo_onboarding,
                    row.floor_model_label,
                    row.first_post_go_live_period,
                    row.floor_reference,
                    row.first_ramp_period,
                    row.ramp_months_from_live,
                    row.approx_sign_to_ramp_months,
                    row.met_6mo_ramp,
                    row.overall_followed_timeline,
                    row.current_assessment,
                    row.explanation,
                    row.monthly_trace,
                ]
            )


def write_markdown(rows: list[PartnerTimeline], summary: dict[str, Any]) -> None:
    live_rows = summary["live_rows"]
    minimum_rows = summary["minimum_floor_rows"]
    fixed_rows = summary["fixed_fee_rows"]
    no_floor_rows = summary["no_floor_rows"]
    resolved_rows = [row for row in minimum_rows + fixed_rows if row.overall_followed_timeline is not None]

    avg_onboarding_days = (
        round(sum(row.onboarding_days or 0 for row in live_rows) / len(live_rows), 1) if live_rows else 0
    )
    avg_onboarding_months = (
        round(sum(row.onboarding_months or 0 for row in live_rows) / len(live_rows), 2) if live_rows else 0
    )
    avg_minimum_ramp = round(
        sum(row.ramp_months_from_live or 0 for row in minimum_rows if row.met_6mo_ramp is not None and row.first_ramp_period)
        / max(1, len([row for row in minimum_rows if row.first_ramp_period])),
        2,
    )
    avg_fixed_ramp = round(
        sum(row.ramp_months_from_live or 0 for row in fixed_rows if row.first_ramp_period)
        / max(1, len([row for row in fixed_rows if row.first_ramp_period])),
        2,
    )
    avg_full_sign_to_ramp = round(
        sum(row.approx_sign_to_ramp_months or 0 for row in rows if row.approx_sign_to_ramp_months is not None)
        / max(1, len([row for row in rows if row.approx_sign_to_ramp_months is not None])),
        2,
    )
    avg_full_sign_to_ramp = round(
        sum(row.approx_sign_to_ramp_months or 0 for row in rows if row.approx_sign_to_ramp_months is not None)
        / max(1, len([row for row in rows if row.approx_sign_to_ramp_months is not None])),
        2,
    )

    lines = [
        "# Partner Timeline Analysis",
        "",
        f"As of loaded revenue period: `{summary['latest_period']}`",
        "",
        "## Headline",
        "",
        f"- Live partners with contract sign and go-live dates: **{len(live_rows)}**",
        f"- Met the 3-month onboarding assumption: **{sum(row.met_3mo_onboarding is True for row in live_rows)}**",
        f"- Missed the 3-month onboarding assumption: **{sum(row.met_3mo_onboarding is False for row in live_rows)}**",
        f"- Average actual onboarding timeline: **{avg_onboarding_days} days / {avg_onboarding_months} months**",
        f"- Average full sign-to-ramp timeline (partners that have actually ramped): **{avg_full_sign_to_ramp} months**",
        f"- Fully assessable partners for the combined 3+6 assumption: **{len(resolved_rows)}**",
        f"- Followed the full timeline: **{sum(row.overall_followed_timeline is True for row in resolved_rows)}**",
        f"- Did not follow the full timeline: **{sum(row.overall_followed_timeline is False for row in resolved_rows)}**",
        "",
        "## Ramp Cohorts",
        "",
        f"- True minimum-floor partners with live dates: **{len(minimum_rows)}**",
        f"  - Ramped within 6 months: **{sum(row.met_6mo_ramp is True for row in minimum_rows)}**",
        f"  - Missed 6 months: **{sum(row.current_assessment == 'Met onboarding, missed ramp' or row.current_assessment == 'Missed onboarding and ramp' for row in minimum_rows)}**",
        f"  - In progress: **{sum(row.current_assessment == 'In progress' for row in minimum_rows)}**",
        f"  - Average actual ramp (ramped only): **{avg_minimum_ramp} months**",
        f"- Fixed subscription/platform proxy partners with live dates: **{len(fixed_rows)}**",
        f"  - Exceeded the fixed floor within 6 months (proxy): **{sum(row.met_6mo_ramp is True for row in fixed_rows)}**",
        f"  - Average proxy ramp (ramped only): **{avg_fixed_ramp} months**",
        f"- Live partners without a comparable floor in contract: **{len(no_floor_rows)}**",
        "",
        "## Methodology",
        "",
        "- Contract sign date: taken from the contract lifecycle backfill.",
        "- Go-live date: taken from the current partner billing record or HubSpot status feed; the lifecycle fallback is only used when the partner is otherwise marked live.",
        "- Onboarding timeline: days from contract sign to go-live; the 3-month assumption is tested using a true calendar 3-month window.",
        "- True minimum-floor contracts: ramp is the first month where activity revenue met or exceeded the contractual floor.",
        "- Fixed subscription/platform contracts: there is no true minimum-top-up to exit, so the report uses a proxy threshold of 105% of the fixed monthly floor.",
        "- Partners without a comparable floor in contract are excluded from the ramp count.",
        "",
        "## Live Partners",
        "",
    ]

    for row in sorted(live_rows, key=lambda item: item.partner.lower()):
        lines.extend(
            [
                f"### {row.partner}",
                "",
                f"- Contract sign: `{row.contract_sign_date}`",
                f"- Go-live: `{row.go_live_date}`",
                f"- Onboarding: `{row.onboarding_days}` days ({row.onboarding_months} months) -> `{row.met_3mo_onboarding}`",
                f"- Floor model: `{row.floor_model_label}`",
                f"- First ramp period: `{row.first_ramp_period or 'n/a'}`",
                f"- Ramp months from live: `{row.ramp_months_from_live if row.ramp_months_from_live is not None else 'n/a'}`",
                f"- Assessment: **{row.current_assessment}**",
                f"- Explanation: {row.explanation}",
                "",
            ]
        )

    MD_PATH.write_text("\n".join(lines))


def write_workbook(rows: list[PartnerTimeline], summary: dict[str, Any]) -> None:
    live_rows = summary["live_rows"]
    minimum_rows = summary["minimum_floor_rows"]
    fixed_rows = summary["fixed_fee_rows"]
    no_floor_rows = summary["no_floor_rows"]
    not_live_rows = summary["not_live_rows"]
    resolved_rows = [row for row in minimum_rows + fixed_rows if row.overall_followed_timeline is not None]

    avg_onboarding_days = (
        round(sum(row.onboarding_days or 0 for row in live_rows) / len(live_rows), 1) if live_rows else 0
    )
    avg_onboarding_months = (
        round(sum(row.onboarding_months or 0 for row in live_rows) / len(live_rows), 2) if live_rows else 0
    )
    avg_minimum_ramp = round(
        sum(row.ramp_months_from_live or 0 for row in minimum_rows if row.first_ramp_period)
        / max(1, len([row for row in minimum_rows if row.first_ramp_period])),
        2,
    )
    avg_fixed_ramp = round(
        sum(row.ramp_months_from_live or 0 for row in fixed_rows if row.first_ramp_period)
        / max(1, len([row for row in fixed_rows if row.first_ramp_period])),
        2,
    )
    avg_full_sign_to_ramp = round(
        sum(row.approx_sign_to_ramp_months or 0 for row in rows if row.approx_sign_to_ramp_months is not None)
        / max(1, len([row for row in rows if row.approx_sign_to_ramp_months is not None])),
        2,
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    metrics = [
        ["Metric", "Value"],
        ["As-of revenue period", summary["latest_period"]],
        ["Live partners with sign + go-live dates", len(live_rows)],
        ["Met 3-month onboarding", sum(row.met_3mo_onboarding is True for row in live_rows)],
        ["Missed 3-month onboarding", sum(row.met_3mo_onboarding is False for row in live_rows)],
        ["Average onboarding days", avg_onboarding_days],
        ["Average onboarding months", avg_onboarding_months],
        ["Average full sign-to-ramp (ramped partners only)", avg_full_sign_to_ramp],
        ["True minimum-floor partners (live)", len(minimum_rows)],
        ["True minimum-floor ramped within 6 months", sum(row.met_6mo_ramp is True for row in minimum_rows)],
        ["True minimum-floor missed 6 months", sum(row.met_6mo_ramp is False for row in minimum_rows)],
        ["True minimum-floor in progress", sum(row.current_assessment == "In progress" for row in minimum_rows)],
        ["Average true minimum ramp (ramped only)", avg_minimum_ramp],
        ["Fixed subscription/platform proxy partners (live)", len(fixed_rows)],
        ["Fixed subscription proxy ramped within 6 months", sum(row.met_6mo_ramp is True for row in fixed_rows)],
        ["Average fixed subscription proxy ramp (ramped only)", avg_fixed_ramp],
        ["Live partners with no comparable floor", len(no_floor_rows)],
        ["Fully assessable partners for full 3+6 assumption", len(resolved_rows)],
        ["Followed full timeline", sum(row.overall_followed_timeline is True for row in resolved_rows)],
        ["Did not follow full timeline", sum(row.overall_followed_timeline is False for row in resolved_rows)],
    ]

    for row in metrics:
        summary_sheet.append(row)
    for cell in summary_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    summary_sheet.column_dimensions["A"].width = 44
    summary_sheet.column_dimensions["B"].width = 18

    summary_sheet["D1"] = "Methodology"
    summary_sheet["D1"].fill = HEADER_FILL
    summary_sheet["D1"].font = HEADER_FONT
    methodology_lines = [
        "Contract sign date comes from the contract lifecycle backfill.",
        "Go-live comes from the current partner billing row or HubSpot; the lifecycle fallback is only used when the partner is otherwise marked live.",
        "3-month onboarding uses a calendar 3-month window.",
        "True minimum-floor ramp: first month activity revenue met/exceeded the contractual floor.",
        "Fixed subscription/platform ramp proxy: first month recurring revenue reached at least 105% of the fixed monthly floor.",
        "Partners with no comparable floor in contract are excluded from the ramp count.",
    ]
    for index, line in enumerate(methodology_lines, start=2):
        summary_sheet[f"D{index}"] = line
    summary_sheet.column_dimensions["D"].width = 80

    headers = [
        "Partner",
        "Contract Sign Date",
        "Go Live Date",
        "Integration Status",
        "Not Yet Live",
        "Onboarding Days",
        "Onboarding Months",
        "Met 3-Month Onboarding",
        "Floor Model",
        "First Post-Go-Live Period",
        "Floor Reference",
        "First Ramp Period",
        "Ramp Months From Live",
        "Approx Total Months Sign To Ramp",
        "Met 6-Month Ramp",
        "Overall Followed Timeline",
        "Current Assessment",
        "Explanation",
        "Monthly Trace",
    ]

    def rows_for(items: list[PartnerTimeline]) -> list[list[Any]]:
        return [
            [
                row.partner,
                row.contract_sign_date,
                row.go_live_date,
                row.integration_status,
                row.not_yet_live,
                row.onboarding_days,
                row.onboarding_months,
                row.met_3mo_onboarding,
                row.floor_model_label,
                row.first_post_go_live_period,
                row.floor_reference,
                row.first_ramp_period,
                row.ramp_months_from_live,
                row.approx_sign_to_ramp_months,
                row.met_6mo_ramp,
                row.overall_followed_timeline,
                row.current_assessment,
                row.explanation,
                row.monthly_trace,
            ]
            for row in items
        ]

    add_sheet(workbook, "Partner Timelines", headers, rows_for(sorted(rows, key=lambda item: item.partner.lower())))
    add_sheet(workbook, "Minimum Floor", headers, rows_for(sorted(minimum_rows, key=lambda item: item.partner.lower())))
    add_sheet(workbook, "Fixed Fee Proxy", headers, rows_for(sorted(fixed_rows, key=lambda item: item.partner.lower())))
    add_sheet(workbook, "Live No Floor", headers, rows_for(sorted(no_floor_rows, key=lambda item: item.partner.lower())))
    add_sheet(workbook, "Not Live Missing", headers, rows_for(sorted(not_live_rows, key=lambda item: item.partner.lower())))

    for sheet in workbook.worksheets[1:]:
        for row in sheet.iter_rows(min_row=2):
            assessment = str(row[15].value or "")
            fill = None
            if assessment == "Followed assumption":
                fill = GOOD_FILL
            elif "Missed" in assessment:
                fill = BAD_FILL
            elif assessment in {"In progress", "Not live yet"}:
                fill = WARN_FILL
            elif "No comparable floor" in assessment or "No post-go-live" in assessment:
                fill = MUTED_FILL
            if fill:
                for cell in row:
                    cell.fill = fill
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(XLSX_PATH)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows, summary = build_timelines()
    write_csv(rows)
    write_markdown(rows, summary)
    write_workbook(rows, summary)

    live_rows = summary["live_rows"]
    resolved_rows = [
        row
        for row in summary["minimum_floor_rows"] + summary["fixed_fee_rows"]
        if row.overall_followed_timeline is not None
    ]
    print(
        json.dumps(
            {
                "csv": str(CSV_PATH),
                "markdown": str(MD_PATH),
                "workbook": str(XLSX_PATH),
                "livePartners": len(live_rows),
                "metOnboarding": sum(row.met_3mo_onboarding is True for row in live_rows),
                "missedOnboarding": sum(row.met_3mo_onboarding is False for row in live_rows),
                "resolvedTimelinePartners": len(resolved_rows),
                "followedTimeline": sum(row.overall_followed_timeline is True for row in resolved_rows),
                "missedTimeline": sum(row.overall_followed_timeline is False for row in resolved_rows),
                "assessmentCounts": Counter(row.current_assessment for row in rows),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
