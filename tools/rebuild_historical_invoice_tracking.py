from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path("/Users/danielsinukoff/Documents/billing-workbook")
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from server.storage import SharedWorkspaceStore

DB_PATH = ROOT / "server" / "data" / "shared_workspace.db"
REPORTS_DIR = ROOT / "reports" / "historical_invoice_rebuild"
WORKBOOK_PATH = Path("/Users/danielsinukoff/Downloads/Partner Billing and Processes (4).xlsx")
PRE_2026_CUTOFF = "2026-01"

SHEET_CONFIG = {
    "Billed and Collected": {"kind": "receivable", "note_col": 19},
    "Partner Payouts": {"kind": "payable", "note_col": 17},
}

PARTNER_ALIASES = {
    "AJ Hanna (Whish)": "Whish",
    "ALTPAYNET": "Altpay",
    "ALTPAYNET ": "Altpay",
    "CellPay": "Cellpay",
    "Gmeremit": "GME_Remit",
    "M-DAQ": "M-Daq",
    "MapleWave": "Maplewave",
    "Oval Tech(Graph)": "Graph Finance",
    "Remittances Hub": "Remittanceshub",
    "Repay ": "Repay",
    "YeePay": "Yeepay",
}


@dataclass
class SourceRow:
    sheet: str
    row_number: int
    partner: str
    period: str
    kind: str
    fee_info: str
    billed_label: str
    billed_is_date: bool
    invoice_date: str
    due_date: str
    paid_date: str
    source_status: str
    workbook_note: str
    invoice_amount: float
    effective_due: float
    amount_paid: float
    writeoff_flag: bool


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Replace pre-2026 invoice tracking rows from historical workbook tabs.")
    parser.add_argument("--workbook", type=Path, default=WORKBOOK_PATH)
    parser.add_argument("--apply", action="store_true", help="Persist the rebuilt historical rows to the shared workbook.")
    return parser.parse_args()


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return ""


def month_key(value: Any) -> str:
    if isinstance(value, datetime):
        return f"{value.year:04d}-{value.month:02d}"
    return ""


def as_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def optional_text(value: Any) -> str:
    raw = text(value)
    return "" if raw in {"0", "0.0", "#VALUE!"} else raw


def looks_like_writeoff(status: str, writeoff_value: Any) -> bool:
    status_text = status.upper()
    writeoff_text = text(writeoff_value).upper()
    return "WRITE OFF" in status_text or "WRITE OFF" in writeoff_text or status_text == "WAIVED"


def normalize_partner(raw_partner: str, known_partners: set[str]) -> str | None:
    partner = text(raw_partner)
    mapped = PARTNER_ALIASES.get(partner, partner)
    if mapped in known_partners:
        return mapped
    return None


def build_period(billed_period: Any, invoice_date: Any) -> tuple[str, bool]:
    if isinstance(billed_period, datetime):
        return month_key(billed_period), True
    if isinstance(invoice_date, datetime):
        return month_key(invoice_date), False
    return "", False


def build_source_rows(workbook_path: Path, known_partners: set[str]) -> tuple[list[SourceRow], list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, data_only=True)
    rows: list[SourceRow] = []
    skipped: list[dict[str, Any]] = []

    for sheet_name, config in SHEET_CONFIG.items():
        ws = wb[sheet_name]
        for r in range(12, ws.max_row + 1):
            raw_partner = ws.cell(r, 1).value
            if raw_partner is None:
                continue

            billed_period = ws.cell(r, 3).value
            invoice_date = ws.cell(r, 4).value
            period, billed_is_date = build_period(billed_period, invoice_date)
            if not period or period >= PRE_2026_CUTOFF:
                continue

            normalized_partner = normalize_partner(text(raw_partner), known_partners)
            if not normalized_partner:
                skipped.append(
                    {
                        "sheet": sheet_name,
                        "row": r,
                        "partner": text(raw_partner),
                        "reason": "No matching current app partner",
                        "feeInfo": text(ws.cell(r, 2).value),
                        "billedPeriod": text(billed_period),
                        "invoiceDate": iso_date(invoice_date),
                        "amount": as_float(ws.cell(r, 7).value),
                    }
                )
                continue

            fee_info = text(ws.cell(r, 2).value)
            due_date = ws.cell(r, 6).value
            invoice_amount = as_float(ws.cell(r, 7).value)
            paid_date = ws.cell(r, 8).value
            paid_amount = as_float(ws.cell(r, 10).value)
            outstanding = as_float(ws.cell(r, 11).value)
            writeoff_value = ws.cell(r, 12).value
            source_status = optional_text(ws.cell(r, 13).value or writeoff_value)
            workbook_note = optional_text(ws.cell(r, config["note_col"]).value)
            writeoff_flag = looks_like_writeoff(source_status, writeoff_value)

            if writeoff_flag:
                effective_due = 0.0
            elif paid_amount or outstanding:
                effective_due = paid_amount + outstanding
            else:
                effective_due = invoice_amount

            rows.append(
                SourceRow(
                    sheet=sheet_name,
                    row_number=r,
                    partner=normalized_partner,
                    period=period,
                    kind=config["kind"],
                    fee_info=fee_info,
                    billed_label=text(billed_period),
                    billed_is_date=billed_is_date,
                    invoice_date=iso_date(invoice_date),
                    due_date=iso_date(due_date),
                    paid_date=iso_date(paid_date),
                    source_status=source_status,
                    workbook_note=workbook_note,
                    invoice_amount=round(invoice_amount, 2),
                    effective_due=round(effective_due, 2),
                    amount_paid=round(paid_amount, 2),
                    writeoff_flag=writeoff_flag,
                )
            )

    return rows, skipped


def build_note(group_rows: list[SourceRow], total_invoice_amount: float, total_effective_due: float) -> str:
    note_parts: list[str] = []
    billed_labels = [row.billed_label for row in group_rows if row.billed_label]
    if billed_labels:
        unique_labels = list(dict.fromkeys(billed_labels))
        if len(unique_labels) > 1 or any(not row.billed_is_date for row in group_rows):
            note_parts.append("Workbook billed periods: " + "; ".join(unique_labels))

    workbook_notes = list(dict.fromkeys(row.workbook_note for row in group_rows if row.workbook_note))
    if workbook_notes:
        note_parts.append("Workbook notes: " + " | ".join(workbook_notes))

    if any(row.writeoff_flag for row in group_rows) and total_invoice_amount > total_effective_due:
        note_parts.append(
            f"Historical workbook invoice total {total_invoice_amount:,.2f} reduced to {total_effective_due:,.2f} after write-offs."
        )

    return " ".join(part.strip() for part in note_parts if part).strip()


def deterministic_id(partner: str, period: str, kind: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", partner.lower()).strip("_")
    return f"hist_{slug}_{period}_{kind}"


def aggregate_rows(source_rows: list[SourceRow]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[SourceRow]] = defaultdict(list)
    for row in source_rows:
        grouped[(row.partner, row.period, row.kind)].append(row)

    rebuilt: list[dict[str, Any]] = []
    for (partner, period, kind), rows in sorted(grouped.items()):
        rows.sort(key=lambda row: (row.invoice_date or "", row.row_number))
        latest_invoice_date = max((row.invoice_date for row in rows if row.invoice_date), default="")
        latest_due_date = max((row.due_date for row in rows if row.due_date), default="")
        latest_paid_date = max((row.paid_date for row in rows if row.paid_date), default="")
        total_invoice_amount = round(sum(row.invoice_amount for row in rows), 2)
        total_effective_due = round(sum(row.effective_due for row in rows), 2)
        total_paid = round(sum(row.amount_paid for row in rows), 2)
        balance = round(total_effective_due - total_paid, 2)
        source_statuses = [row.source_status for row in rows if row.source_status]
        fee_infos = [row.fee_info for row in rows if row.fee_info]

        rebuilt.append(
            {
                "id": deterministic_id(partner, period, kind),
                "partner": partner,
                "period": period,
                "kind": kind,
                "invoiceDate": latest_invoice_date,
                "paid": balance <= 0.005,
                "amountPaid": total_paid,
                "amountDueOverride": total_effective_due,
                "dueDateOverride": latest_due_date,
                "paidDate": latest_paid_date,
                "sourceSheet": rows[0].sheet,
                "sourceStatus": " | ".join(dict.fromkeys(source_statuses)),
                "sourceFeeInfo": " | ".join(dict.fromkeys(fee_infos)),
                "note": build_note(rows, total_invoice_amount, total_effective_due),
            }
        )

    return rebuilt


def save_report(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    args = parse_args()
    store = SharedWorkspaceStore(DB_PATH)
    workspace = store.get_workspace()
    snapshot = workspace.get("snapshot") or {}
    known_partners = set(snapshot.get("ps") or [])
    current_rows = list(snapshot.get("pInvoices") or [])
    preserved_rows = [row for row in current_rows if str(row.get("period", "")) >= PRE_2026_CUTOFF]
    replaced_rows = [row for row in current_rows if str(row.get("period", "")) < PRE_2026_CUTOFF]

    source_rows, skipped = build_source_rows(args.workbook, known_partners)
    rebuilt_rows = aggregate_rows(source_rows)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = REPORTS_DIR / f"pre2026-pInvoices-backup-{timestamp}.json"
    report_path = REPORTS_DIR / f"historical-rebuild-report-{timestamp}.json"
    save_report(backup_path, {"rows": replaced_rows})

    report_payload = {
        "workbook": str(args.workbook),
        "cutoff": PRE_2026_CUTOFF,
        "existingPre2026Rows": len(replaced_rows),
        "preservedPost2025Rows": len(preserved_rows),
        "sourceRowsUsed": len(source_rows),
        "rebuiltRows": len(rebuilt_rows),
        "rebuiltRowsData": rebuilt_rows,
        "skippedRows": skipped,
        "sampleRebuiltRows": rebuilt_rows[:20],
        "backup": str(backup_path),
        "applied": args.apply,
    }
    save_report(report_path, report_payload)

    if args.apply:
        snapshot["pInvoices"] = preserved_rows + rebuilt_rows
        saved_at = store.save_snapshot(snapshot)
        print(json.dumps({"savedAt": saved_at, "backup": str(backup_path), "report": str(report_path), "rebuiltRows": len(rebuilt_rows), "skippedRows": len(skipped)}, indent=2))
    else:
        print(json.dumps({"backup": str(backup_path), "report": str(report_path), "rebuiltRows": len(rebuilt_rows), "skippedRows": len(skipped)}, indent=2))


if __name__ == "__main__":
    main()
