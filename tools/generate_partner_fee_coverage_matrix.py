from __future__ import annotations

import json
import sqlite3
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path("/Users/danielsinukoff/Documents/billing-workbook")
DB_PATH = ROOT / "server/data/shared_workspace.db"
CONFIG_PATH = ROOT / "docs/looker-direct-reports.json"
OUT_DIR = ROOT / "reports/coverage"
OUT_XLSX = OUT_DIR / "partner_fee_coverage_matrix.xlsx"
OUT_CSV = OUT_DIR / "partner_fee_coverage_missing.csv"


MATRIX_COLUMNS = [
    "Implementation Fee",
    "Monthly Platform / Subscription",
    "Minimum Monthly Revenue Commitment",
    "Offline Fixed Txn Fees",
    "Volume % Fees",
    "FX Conversion Fees",
    "RTP Fee Caps",
    "Reversal Fees",
    "Surcharges",
    "Rev Share Payout",
    "Stampli FX Payout",
    "VA Account Opening",
    "VA Monthly Active",
    "VA Dormancy",
    "VA Account Closing",
    "Account Setup Per Business",
    "Account Setup Per Individual",
    "Daily Settlement",
    "New Business Setup",
]


STATUS_PRIORITY = {
    "Configured - source data available": 0,
    "Configured - no data required": 1,
    "Configured - source feed stale vs latest workflow period": 2,
    "Configured - latest feed imported 0 rows": 3,
    "Configured - required field not populated in imported data": 4,
    "Configured - partner has no imported source rows": 5,
    "Configured - source report not in active workflow": 6,
    "No fee configured": 7,
}


@dataclass
class FeedStatus:
    name: str
    active: bool
    period: str
    saved_at: str
    section_counts: dict[str, int]
    warnings: list[str]
    is_current: bool
    has_rows: bool


@dataclass
class FeeAssessment:
    partner: str
    fee_type: str
    status: str
    detail: str
    required_feed: str
    latest_workflow_period: str
    latest_feed_period: str
    latest_feed_saved_at: str
    latest_feed_section_counts: str
    warnings: str


def latest_snapshot() -> dict[str, Any]:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    row = cur.execute("SELECT snapshot_json FROM workbook_snapshots ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        raise RuntimeError("No workbook snapshot found.")
    return json.loads(row[0])


def load_active_reports() -> set[str]:
    cfg = json.loads(CONFIG_PATH.read_text())
    return {report["fileType"] for report in cfg.get("reports", [])}


def section_total(section_counts: dict[str, Any]) -> int:
    total = 0
    for value in (section_counts or {}).values():
        try:
            total += int(value or 0)
        except (TypeError, ValueError):
            continue
    return total


def partner_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for row in rows:
        partner = str(row.get("partner") or "").strip()
        if partner:
            counts[partner] += 1
    return counts


def partner_counts_where(rows: list[dict[str, Any]], predicate) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for row in rows:
        partner = str(row.get("partner") or "").strip()
        if partner and predicate(row):
            counts[partner] += 1
    return counts


def has_impl_fee_type(rows: list[dict[str, Any]], partner: str, fee_type: str) -> bool:
    return any(row.get("partner") == partner and row.get("feeType") == fee_type for row in rows)


def has_va_fee_type(rows: list[dict[str, Any]], partner: str, fee_type: str) -> bool:
    return any(row.get("partner") == partner and row.get("feeType") == fee_type for row in rows)


def any_note_contains(rows: list[dict[str, Any]], partner: str, fee_type: str, text: str) -> bool:
    text = text.lower()
    return any(
        row.get("partner") == partner
        and row.get("feeType") == fee_type
        and text in str(row.get("note") or "").lower()
        for row in rows
    )


def format_counts(section_counts: dict[str, int]) -> str:
    if not section_counts:
        return ""
    return ", ".join(f"{key}={value}" for key, value in sorted(section_counts.items()))


def make_feed_statuses(by_file_type: dict[str, Any], active_reports: set[str], latest_period: str) -> dict[str, FeedStatus]:
    statuses: dict[str, FeedStatus] = {}
    for name in {
        "partner_revenue_summary",
        "partner_offline_billing",
        "partner_offline_billing_reversals",
        "partner_rev_share_v2",
        "partner_revenue_share",
        "partner_revenue_reversal",
        "all_registered_accounts",
        "stampli_fx_revenue_share",
        "stampli_fx_revenue_reversal",
    }:
        record = by_file_type.get(name, {}) or {}
        section_counts = record.get("sectionCounts", {}) or {}
        period = str(record.get("period") or "")
        statuses[name] = FeedStatus(
            name=name,
            active=name in active_reports,
            period=period,
            saved_at=str(record.get("savedAt") or ""),
            section_counts={k: int(v or 0) for k, v in section_counts.items()},
            warnings=[str(item) for item in (record.get("warnings", []) or [])],
            is_current=(not latest_period) or period == latest_period,
            has_rows=section_total(section_counts) > 0,
        )
    return statuses


def assess_fee(
    *,
    partner: str,
    fee_type: str,
    configured: bool,
    required_feed: str,
    latest_period: str,
    feed_status: FeedStatus | None = None,
    no_data_needed: bool = False,
    partner_source_count: int = 0,
    field_supported: bool = True,
    detail_suffix: str = "",
) -> FeeAssessment:
    if not configured:
        status = "No fee configured"
        detail = "This fee type is not configured for the partner in the workbook."
    elif no_data_needed:
        status = "Configured - no data required"
        detail = "This is a contract-only fee and does not depend on imported Looker rows."
    elif not feed_status or not feed_status.active:
        status = "Configured - source report not in active workflow"
        detail = f"The fee depends on `{required_feed}`, which is not in the current direct Looker workflow."
    elif not feed_status.is_current:
        status = "Configured - source feed stale vs latest workflow period"
        detail = (
            f"The fee depends on `{required_feed}`, but the latest saved feed period is `{feed_status.period or 'n/a'}` "
            f"while the latest workflow period is `{latest_period or 'n/a'}`."
        )
    elif not feed_status.has_rows:
        status = "Configured - latest feed imported 0 rows"
        detail = f"`{required_feed}` is active, but its latest saved import has no rows."
    elif not field_supported:
        status = "Configured - required field not populated in imported data"
        detail = f"`{required_feed}` exists, but the specific field needed for this fee is not populated in the imported data."
    elif partner_source_count <= 0:
        status = "Configured - partner has no imported source rows"
        detail = f"`{required_feed}` has current rows, but none for this partner."
    else:
        status = "Configured - source data available"
        detail = f"`{required_feed}` has {partner_source_count} current row(s) for this partner."

    if detail_suffix:
        detail = f"{detail} {detail_suffix}".strip()

    return FeeAssessment(
        partner=partner,
        fee_type=fee_type,
        status=status,
        detail=detail,
        required_feed=required_feed,
        latest_workflow_period=latest_period,
        latest_feed_period=feed_status.period if feed_status else "",
        latest_feed_saved_at=feed_status.saved_at if feed_status else "",
        latest_feed_section_counts=format_counts(feed_status.section_counts) if feed_status else "",
        warnings=" | ".join(feed_status.warnings) if feed_status and feed_status.warnings else "",
    )


def append_sheet_with_header(ws, headers: list[str], fill: str) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=fill)


def auto_size(ws, min_width: int = 16, max_width: int = 70) -> None:
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, min_width), max_width)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    state = latest_snapshot()
    active_reports = load_active_reports()
    audit = state.get("lookerImportAudit", {}) or {}
    by_file_type = audit.get("byFileType", {}) or {}
    latest_run = audit.get("latestRun", {}) or {}
    latest_period = str(latest_run.get("period") or "")
    feed_statuses = make_feed_statuses(by_file_type, active_reports, latest_period)

    partners = sorted(state.get("ps", []))
    off = state.get("off", [])
    vol = state.get("vol", [])
    fx_rates = state.get("fxRates", [])
    caps = state.get("cap", [])
    mins = state.get("mins", [])
    revf = state.get("revf", [])
    plat = state.get("plat", [])
    impl = state.get("impl", [])
    va_fees = state.get("vaFees", [])
    surch = state.get("surch", [])
    rs = state.get("rs", [])
    ltxn = state.get("ltxn", [])
    lrev = state.get("lrev", [])
    lva = state.get("lva", [])
    lrs = state.get("lrs", [])
    lfxp = state.get("lfxp", [])

    ltxn_counts = partner_counts(ltxn)
    lrev_counts = partner_counts(lrev)
    lva_counts = partner_counts(lva)
    lrs_counts = partner_counts(lrs)
    lfxp_counts = partner_counts(lfxp)
    rev_txn_counts = partner_counts_where(ltxn, lambda row: bool(row.get("revenueBasis")))

    total_business_supported = any(float(row.get("totalBusinessAccounts") or 0) > 0 for row in lva)
    total_individual_supported = any(float(row.get("totalIndividualAccounts") or 0) > 0 for row in lva)
    closed_accounts_supported = any(float(row.get("closedAccounts") or 0) > 0 for row in lva)

    rows_out: list[dict[str, str]] = []
    detail_rows: list[FeeAssessment] = []

    for partner in partners:
        partner_row = {"Partner": partner}

        partner_assessments = [
            assess_fee(
                partner=partner,
                fee_type="Implementation Fee",
                configured=has_impl_fee_type(impl, partner, "Implementation"),
                required_feed="",
                latest_period=latest_period,
                no_data_needed=True,
                detail_suffix=f"Configured rows: {sum(1 for row in impl if row.get('partner') == partner and row.get('feeType') == 'Implementation')}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="Monthly Platform / Subscription",
                configured=any(row.get("partner") == partner for row in plat),
                required_feed="",
                latest_period=latest_period,
                no_data_needed=True,
                detail_suffix=f"Configured rows: {sum(1 for row in plat if row.get('partner') == partner)}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="Minimum Monthly Revenue Commitment",
                configured=any(row.get("partner") == partner for row in mins),
                required_feed="partner_revenue_summary",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_revenue_summary"],
                partner_source_count=lrs_counts.get(partner, 0),
                detail_suffix=f"Configured minimum rows: {sum(1 for row in mins if row.get('partner') == partner)}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="Offline Fixed Txn Fees",
                configured=any(row.get("partner") == partner for row in off),
                required_feed="partner_offline_billing",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_offline_billing"],
                partner_source_count=ltxn_counts.get(partner, 0),
                detail_suffix=f"Configured offline rows: {sum(1 for row in off if row.get('partner') == partner)}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="Volume % Fees",
                configured=any(row.get("partner") == partner for row in vol),
                required_feed="partner_offline_billing",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_offline_billing"],
                partner_source_count=ltxn_counts.get(partner, 0),
                detail_suffix=f"Configured volume rows: {sum(1 for row in vol if row.get('partner') == partner)}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="FX Conversion Fees",
                configured=any(row.get("partner") == partner for row in fx_rates),
                required_feed="partner_offline_billing",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_offline_billing"],
                partner_source_count=ltxn_counts.get(partner, 0),
                detail_suffix=f"Configured FX rows: {sum(1 for row in fx_rates if row.get('partner') == partner)}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="RTP Fee Caps",
                configured=any(row.get("partner") == partner and str(row.get("productType") or "").upper() == "RTP" for row in caps),
                required_feed="partner_offline_billing",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_offline_billing"],
                partner_source_count=ltxn_counts.get(partner, 0),
                detail_suffix=f"Configured cap rows: {sum(1 for row in caps if row.get('partner') == partner and str(row.get('productType') or '').upper() == 'RTP')}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="Reversal Fees",
                configured=any(row.get("partner") == partner for row in revf),
                required_feed="partner_offline_billing_reversals",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_offline_billing_reversals"],
                partner_source_count=lrev_counts.get(partner, 0),
                detail_suffix=f"Configured reversal rows: {sum(1 for row in revf if row.get('partner') == partner)}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="Surcharges",
                configured=any(row.get("partner") == partner for row in surch),
                required_feed="partner_offline_billing",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_offline_billing"],
                partner_source_count=ltxn_counts.get(partner, 0),
                detail_suffix=f"Configured surcharge rows: {sum(1 for row in surch if row.get('partner') == partner)}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="Rev Share Payout",
                configured=any(row.get("partner") == partner for row in rs),
                required_feed="partner_revenue_summary / partner_revenue_share",
                latest_period=latest_period,
                feed_status=feed_statuses["partner_revenue_summary"],
                partner_source_count=max(lrs_counts.get(partner, 0), rev_txn_counts.get(partner, 0), lfxp_counts.get(partner, 0)),
                detail_suffix=(
                    f"Configured rev-share rows: {sum(1 for row in rs if row.get('partner') == partner)}. "
                    f"Latest detail-feed rows: revenue_share={section_total(feed_statuses['partner_revenue_share'].section_counts)}, "
                    f"revenue_reversal={section_total(feed_statuses['partner_revenue_reversal'].section_counts)}, "
                    f"rev_share_v2={section_total(feed_statuses['partner_rev_share_v2'].section_counts)}."
                ),
            ),
            assess_fee(
                partner=partner,
                fee_type="Stampli FX Payout",
                configured=partner == "Stampli",
                required_feed="stampli_fx_revenue_share / stampli_fx_revenue_reversal",
                latest_period=latest_period,
                feed_status=feed_statuses["stampli_fx_revenue_share"],
                partner_source_count=lfxp_counts.get(partner, 0),
                detail_suffix=(
                    f"Reversal feed latest period: {feed_statuses['stampli_fx_revenue_reversal'].period or 'n/a'}. "
                    f"Combined current rows for partner: {lfxp_counts.get(partner, 0)}."
                ),
            ),
            assess_fee(
                partner=partner,
                fee_type="VA Account Opening",
                configured=has_va_fee_type(va_fees, partner, "Account Opening"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                detail_suffix=f"Configured VA opening rows: {sum(1 for row in va_fees if row.get('partner') == partner and row.get('feeType') == 'Account Opening')}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="VA Monthly Active",
                configured=has_va_fee_type(va_fees, partner, "Monthly Active"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                detail_suffix=f"Configured VA monthly-active rows: {sum(1 for row in va_fees if row.get('partner') == partner and row.get('feeType') == 'Monthly Active')}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="VA Dormancy",
                configured=has_va_fee_type(va_fees, partner, "Dormancy"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                detail_suffix=f"Configured VA dormancy rows: {sum(1 for row in va_fees if row.get('partner') == partner and row.get('feeType') == 'Dormancy')}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="VA Account Closing",
                configured=has_va_fee_type(va_fees, partner, "Account Closing"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                field_supported=closed_accounts_supported,
                detail_suffix="This fee depends on `closedAccounts`, which is currently never populated above zero." if not closed_accounts_supported else "",
            ),
            assess_fee(
                partner=partner,
                fee_type="Account Setup Per Business",
                configured=any_note_contains(impl, partner, "Account Setup", "per business"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                field_supported=total_business_supported,
                detail_suffix="This fee depends on `totalBusinessAccounts`, which is currently never populated above zero." if not total_business_supported else "",
            ),
            assess_fee(
                partner=partner,
                fee_type="Account Setup Per Individual",
                configured=any_note_contains(impl, partner, "Account Setup", "per individual"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                field_supported=total_individual_supported,
                detail_suffix="This fee depends on `totalIndividualAccounts`, which is currently never populated above zero." if not total_individual_supported else "",
            ),
            assess_fee(
                partner=partner,
                fee_type="Daily Settlement",
                configured=has_impl_fee_type(impl, partner, "Daily Settlement"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                detail_suffix=f"Configured daily-settlement rows: {sum(1 for row in impl if row.get('partner') == partner and row.get('feeType') == 'Daily Settlement')}.",
            ),
            assess_fee(
                partner=partner,
                fee_type="New Business Setup",
                configured=has_impl_fee_type(impl, partner, "Account Setup"),
                required_feed="all_registered_accounts",
                latest_period=latest_period,
                feed_status=feed_statuses["all_registered_accounts"],
                partner_source_count=lva_counts.get(partner, 0),
                detail_suffix=f"Configured account-setup rows: {sum(1 for row in impl if row.get('partner') == partner and row.get('feeType') == 'Account Setup')}.",
            ),
        ]

        for assessment in partner_assessments:
            partner_row[assessment.fee_type] = assessment.status
            detail_rows.append(assessment)

        rows_out.append(partner_row)

    blocked_rows = [
        row for row in detail_rows
        if row.status not in {"No fee configured", "Configured - no data required", "Configured - source data available"}
    ]
    blocked_rows.sort(key=lambda row: (STATUS_PRIORITY.get(row.status, 99), row.partner, row.fee_type))

    summary_counter = Counter((row.fee_type, row.status) for row in detail_rows if row.status != "No fee configured")

    wb = Workbook()

    ws = wb.active
    ws.title = "Partner Coverage Matrix"
    matrix_headers = ["Partner"] + MATRIX_COLUMNS
    append_sheet_with_header(ws, matrix_headers, "D9EAF7")
    for row in rows_out:
        ws.append([row.get(header, "") for header in matrix_headers])
    auto_size(ws, min_width=18, max_width=42)

    ws2 = wb.create_sheet("Coverage Detail")
    detail_headers = [
        "Partner",
        "Fee Type",
        "Status",
        "Detail",
        "Required Feed",
        "Latest Workflow Period",
        "Latest Feed Period",
        "Latest Feed Saved At",
        "Latest Feed Section Counts",
        "Warnings",
    ]
    append_sheet_with_header(ws2, detail_headers, "E2F0D9")
    for row in sorted(detail_rows, key=lambda item: (item.partner, item.fee_type)):
        ws2.append([
            row.partner,
            row.fee_type,
            row.status,
            row.detail,
            row.required_feed,
            row.latest_workflow_period,
            row.latest_feed_period,
            row.latest_feed_saved_at,
            row.latest_feed_section_counts,
            row.warnings,
        ])
    auto_size(ws2, min_width=18, max_width=80)

    ws3 = wb.create_sheet("Missing Or Blocked")
    append_sheet_with_header(ws3, detail_headers, "FBE5D6")
    for row in blocked_rows:
        ws3.append([
            row.partner,
            row.fee_type,
            row.status,
            row.detail,
            row.required_feed,
            row.latest_workflow_period,
            row.latest_feed_period,
            row.latest_feed_saved_at,
            row.latest_feed_section_counts,
            row.warnings,
        ])
    auto_size(ws3, min_width=18, max_width=80)

    ws4 = wb.create_sheet("Workflow Feeds")
    feed_headers = ["Feed", "In Active Workflow", "Latest Saved Period", "Latest Saved At", "Current vs Latest Run", "Rows Imported", "Section Counts", "Warnings"]
    append_sheet_with_header(ws4, feed_headers, "FFF2CC")
    for name in sorted(feed_statuses):
        status = feed_statuses[name]
        ws4.append([
            name,
            "Yes" if status.active else "No",
            status.period,
            status.saved_at,
            "Current" if status.is_current else "Stale",
            "Yes" if status.has_rows else "No",
            format_counts(status.section_counts),
            " | ".join(status.warnings),
        ])
    auto_size(ws4, min_width=18, max_width=60)

    ws5 = wb.create_sheet("Status Summary")
    summary_headers = ["Fee Type", "Status", "Partner Count"]
    append_sheet_with_header(ws5, summary_headers, "D9EAD3")
    for fee_type in MATRIX_COLUMNS:
        for status, count in sorted(
            ((status, count) for (fee, status), count in summary_counter.items() if fee == fee_type),
            key=lambda item: STATUS_PRIORITY.get(item[0], 99),
        ):
            ws5.append([fee_type, status, count])
    auto_size(ws5, min_width=18, max_width=60)

    ws6 = wb.create_sheet("Legend")
    legend_rows = [
        ["Configured - no data required", "Contract-only fee; the app can bill this without Looker rows."],
        ["Configured - source data available", "The fee is configured and the needed imported source rows exist for that partner."],
        ["Configured - source feed stale vs latest workflow period", "The feed is in the workflow, but its latest saved import is older than the latest workflow period."],
        ["Configured - latest feed imported 0 rows", "The feed is in the workflow, but its latest saved import returned 0 rows overall."],
        ["Configured - required field not populated in imported data", "The feed exists, but the specific field needed for the fee is not populated in the saved data."],
        ["Configured - partner has no imported source rows", "The feed is current and non-empty overall, but this partner has no matching rows."],
        ["Configured - source report not in active workflow", "The fee depends on a Looker report that is not part of the current direct-import config."],
        ["No fee configured", "That fee family is not configured for the partner in the workbook."],
    ]
    append_sheet_with_header(ws6, ["Status", "Meaning"], "F4CCCC")
    for legend_row in legend_rows:
        ws6.append(legend_row)
    auto_size(ws6, min_width=24, max_width=110)

    wb.save(OUT_XLSX)

    csv_headers = [
        "Partner",
        "Fee Type",
        "Status",
        "Detail",
        "Required Feed",
        "Latest Workflow Period",
        "Latest Feed Period",
        "Latest Feed Saved At",
        "Latest Feed Section Counts",
        "Warnings",
    ]
    csv_lines = [",".join(csv_headers)]
    for row in blocked_rows:
        values = [
            row.partner,
            row.fee_type,
            row.status,
            row.detail,
            row.required_feed,
            row.latest_workflow_period,
            row.latest_feed_period,
            row.latest_feed_saved_at,
            row.latest_feed_section_counts,
            row.warnings,
        ]
        csv_lines.append(",".join(f"\"{str(value).replace('\"', '\"\"')}\"" for value in values))
    OUT_CSV.write_text("\n".join(csv_lines) + "\n")

    print(json.dumps({
        "xlsx": str(OUT_XLSX),
        "csv": str(OUT_CSV),
        "partners": len(partners),
        "coverageRows": len(detail_rows),
        "missingOrBlockedRows": len(blocked_rows),
        "latestWorkflowPeriod": latest_period,
    }, indent=2))


if __name__ == "__main__":
    main()
